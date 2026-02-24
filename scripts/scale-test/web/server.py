#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import io
import json
import mimetypes
import os
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import parse_qs, quote, unquote, urlparse


WEB_DIR = Path(__file__).resolve().parent
SCALE_TEST_ROOT_DEFAULT = (WEB_DIR / "..").resolve()


@dataclass(frozen=True)
class RunRef:
    task: str
    suite: str
    run_id: str


@dataclass(frozen=True)
class RunInfo:
    ref: RunRef
    run_dir: Path
    analysis_dir: Path
    mtime: float


# Best-effort metadata cache: {run_dir: (mtime, meta_dict)}
_RUN_META_CACHE: Dict[str, Tuple[float, Dict[str, str]]] = {}


def _e(s: Any) -> str:
    return html.escape(str(s), quote=True)


def _fmt_dt(ts: float) -> str:
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(ts)


def _safe_relpath(p: Path, root: Path) -> str:
    try:
        return p.relative_to(root).as_posix()
    except Exception:
        return p.as_posix()


def _sanitize_download_name(name: str, *, default_ext: Optional[str] = None, max_len: int = 200) -> str:
    """Sanitize a user-provided download filename.

    - Removes path separators
    - Keeps alnum, '.', '-', '_' and replaces others with '_'
    - Optionally enforces an extension
    """
    s = (name or "").strip()
    # drop any path component
    s = s.split("/")[-1].split("\\")[-1]
    if not s:
        s = "download"
    out = []
    for ch in s:
        if ch.isalnum() or ch in {".", "-", "_"}:
            out.append(ch)
        else:
            out.append("_")
    s2 = "".join(out)
    while "__" in s2:
        s2 = s2.replace("__", "_")
    s2 = s2.strip("._-") or "download"
    if default_ext:
        ext = default_ext if default_ext.startswith(".") else f".{default_ext}"
        if not s2.lower().endswith(ext.lower()):
            s2 = s2 + ext
    if len(s2) > max_len:
        # keep extension if present
        suf = ""
        if "." in s2:
            base, dot, ext = s2.rpartition(".")
            if base and ext:
                suf = dot + ext
                s2 = base
        s2 = s2[: max(1, max_len - len(suf))] + suf
    return s2


def discover_runs(scale_test_root: Path) -> List[RunInfo]:
    runs: List[RunInfo] = []
    for task_dir in sorted(scale_test_root.iterdir()):
        if not task_dir.is_dir():
            continue
        task = task_dir.name
        if task in {"web", "__pycache__"}:
            continue
        result_dir = task_dir / "result"
        if not result_dir.exists() or not result_dir.is_dir():
            continue

        for suite_dir in sorted(result_dir.iterdir()):
            if not suite_dir.is_dir():
                continue
            suite = suite_dir.name

            for run_dir in sorted(suite_dir.iterdir()):
                if not run_dir.is_dir():
                    continue
                run_id = run_dir.name
                analysis_dir = run_dir / "analysis"
                if not analysis_dir.exists() or not analysis_dir.is_dir():
                    continue

                try:
                    mtime = run_dir.stat().st_mtime
                except Exception:
                    mtime = 0.0

                runs.append(
                    RunInfo(
                        ref=RunRef(task=task, suite=suite, run_id=run_id),
                        run_dir=run_dir,
                        analysis_dir=analysis_dir,
                        mtime=mtime,
                    )
                )

    # Newest first
    runs.sort(key=lambda r: r.mtime, reverse=True)
    return runs


def _read_text(path: Path, max_bytes: int = 256_000) -> str:
    data = path.read_bytes()[:max_bytes]
    try:
        return data.decode("utf-8")
    except Exception:
        return data.decode("utf-8", errors="replace")


def _list_analysis_files(analysis_dir: Path) -> Tuple[List[Path], List[Path]]:
    pngs: List[Path] = []
    csvs: List[Path] = []
    for p in sorted(analysis_dir.iterdir()):
        if not p.is_file():
            continue
        suf = p.suffix.lower()
        if suf == ".png":
            pngs.append(p)
        elif suf == ".csv":
            csvs.append(p)
    return pngs, csvs


def _extract_run_meta(run: RunInfo) -> Dict[str, str]:
    """Best-effort extract run metadata for display/filter.

    Sources (in priority order):
    - model: run_dir/aggregate.csv (column: model or model_id)
    - model fallback: */auto_test_config.generated.json (jobs[0].env.MODEL / MODEL_ID)
    """
    key = str(run.run_dir)
    cached = _RUN_META_CACHE.get(key)
    if cached and cached[0] == run.mtime:
        return cached[1]

    meta: Dict[str, str] = {"model": "", "cpu": "", "cpu_cores": "", "memory": "", "mode": ""}

    def _fmt_mem_bytes(n: int) -> str:
        if n <= 0:
            return ""
        gib = float(n) / float(1024 ** 3)
        if gib >= 1024.0:
            return f"{(gib / 1024.0):.2f} TiB"
        return f"{gib:.1f} GiB"

    def _parse_memtotal_bytes_from_meminfo(s: str) -> int:
        try:
            for line in (s or "").splitlines():
                if line.lower().startswith("memtotal:"):
                    m = re.search(r"MemTotal:\s*(\d+)\s*kB", line, re.IGNORECASE)
                    if not m:
                        continue
                    kb = int(m.group(1))
                    if kb > 0:
                        return kb * 1024
        except Exception:
            return 0
        return 0

    def _parse_mem_bytes_from_lscpu_text(s: str) -> int:
        # Newer util-linux may print a line like: "Memory: 1.5 TiB".
        unit_pow = {
            "b": 0,
            "kb": 10,
            "kib": 10,
            "mb": 20,
            "mib": 20,
            "gb": 30,
            "gib": 30,
            "tb": 40,
            "tib": 40,
        }
        try:
            for line in (s or "").splitlines():
                if not line.lower().startswith("memory:"):
                    continue
                _, _, tail = line.partition(":")
                txt = tail.strip()
                m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*([KMGT]?i?B)", txt, re.IGNORECASE)
                if not m:
                    continue
                val = float(m.group(1))
                unit = m.group(2).lower()
                p = unit_pow.get(unit)
                if p is None:
                    continue
                n = int(val * (2 ** p))
                if n > 0:
                    return n
        except Exception:
            return 0
        return 0

    def _cpu_model_from_lscpu_text(s: str) -> str:
        # Typical line: "Model name:                           Intel(R) Xeon(R) ..."
        try:
            for line in (s or "").splitlines():
                if line.lower().startswith("model name"):
                    # Split at the first ':'
                    _, _, tail = line.partition(":")
                    v = tail.strip()
                    if v:
                        return v
        except Exception:
            pass
        return ""

    def _add_cpu_from_lscpu_path(p: Path) -> None:
        if meta.get("cpu"):
            return
        try:
            if p.exists() and p.is_file():
                txt = _read_text(p, max_bytes=256_000)
                v = _cpu_model_from_lscpu_text(txt)
                if v:
                    meta["cpu"] = v
        except Exception:
            return

    def _add_mem_from_server_info(base_dir: Path) -> None:
        if meta.get("memory"):
            return
        try:
            meminfo = (base_dir / "meminfo.txt").resolve()
            if _is_within(meminfo, run.run_dir) and meminfo.exists() and meminfo.is_file():
                n = _parse_memtotal_bytes_from_meminfo(_read_text(meminfo, max_bytes=256_000))
                if n > 0:
                    meta["memory"] = _fmt_mem_bytes(n)
                    return
        except Exception:
            pass
        try:
            lscpu = (base_dir / "lscpu.txt").resolve()
            if _is_within(lscpu, run.run_dir) and lscpu.exists() and lscpu.is_file():
                n = _parse_mem_bytes_from_lscpu_text(_read_text(lscpu, max_bytes=256_000))
                if n > 0:
                    meta["memory"] = _fmt_mem_bytes(n)
        except Exception:
            pass

    # 1) model from aggregate.csv
    try:
        agg = run.run_dir / "aggregate.csv"
        if agg.exists():
            with agg.open("r", encoding="utf-8", errors="replace", newline="") as f:
                reader = csv.DictReader(f)
                for i, r in enumerate(reader):
                    if i > 300:
                        break
                    m = (r.get("model") or "").strip()
                    mid = (r.get("model_id") or "").strip()
                    mode = (r.get("mode") or "").strip()
                    cores = (r.get("resource_cpu_count") or r.get("resource_cpu") or "").strip()

                    if mode and not meta.get("mode"):
                        meta["mode"] = mode

                    # For cpu_cores, take the max observed across the first rows.
                    if cores:
                        try:
                            c = int(float(cores))
                            prev = int(meta.get("cpu_cores") or "0") if (meta.get("cpu_cores") or "").isdigit() else 0
                            if c > prev:
                                meta["cpu_cores"] = str(c)
                        except Exception:
                            pass
                    if m:
                        meta["model"] = m
                        break
                    if mid and not meta["model"]:
                        meta["model"] = mid
                # If still empty, fall back to header-only rows.
    except Exception:
        pass

    # 2) model fallback from auto_test_config.generated.json (first job env)
    def _try_model_from(p: Path) -> None:
        if meta.get("model"):
            return
        try:
            obj = json.loads(p.read_text(encoding="utf-8"))
            jobs = obj.get("jobs") or []
            if not jobs:
                return
            env = (jobs[0].get("env") or {}) if isinstance(jobs[0], dict) else {}
            model2 = str(env.get("MODEL") or "").strip()
            model_id2 = str(env.get("MODEL_ID") or "").strip()
            mode2 = str(env.get("MODE") or "").strip()
            if mode2 and not meta.get("mode"):
                meta["mode"] = mode2
            if not meta.get("model"):
                if model2:
                    meta["model"] = model2
                elif model_id2:
                    meta["model"] = model_id2
        except Exception:
            return

    try:
        # Prefer variants.json order if present.
        v = run.run_dir / "variants.json"
        if v.exists():
            obj = json.loads(v.read_text(encoding="utf-8"))
            variants = obj.get("variants") or []
            if variants and isinstance(variants, list):
                first = variants[0]
                d = first.get("dir") if isinstance(first, dict) else ""
                if d:
                    p = Path(str(d)) / "auto_test_config.generated.json"
                    if p.exists():
                        _try_model_from(p)
        if not meta.get("model"):
            # Fallback: find any generated config under the run.
            for p in sorted(run.run_dir.glob("**/auto_test_config.generated.json")):
                _try_model_from(p)
                if meta.get("model"):
                    break
    except Exception:
        pass

    # 3) CPU model from captured lscpu output.
    # Prefer per-run server_info for local runs, else per-host server_info.
    try:
        _add_cpu_from_lscpu_path(run.run_dir / "server_info" / "lscpu.txt")
        _add_mem_from_server_info(run.run_dir / "server_info")
        if not meta.get("cpu"):
            # Remote dispatch: <run>/hosts/<host>/server_info/lscpu.txt
            for p in sorted(run.run_dir.glob("hosts/*/server_info/lscpu.txt")):
                _add_cpu_from_lscpu_path(p)
                if meta.get("cpu"):
                    break
        if not meta.get("memory"):
            for d in sorted(run.run_dir.glob("hosts/*/server_info")):
                if not d.is_dir():
                    continue
                _add_mem_from_server_info(d)
                if meta.get("memory"):
                    break
    except Exception:
        pass

    _RUN_META_CACHE[key] = (run.mtime, meta)
    return meta


def _apply_home_sort(runs: List[RunInfo], *, sort_spec: str) -> List[RunInfo]:
    """Stable multi-key sort for the home page.

    sort_spec: comma-separated keys, optionally prefixed with '-'.
    Supported keys: task, suite, run, model, cpu, memory, mtime
    Default (empty/invalid): task,suite,run,model,cpu,memory,mtime
    """

    spec = (sort_spec or "").strip()
    if not spec:
        spec = "task,suite,run,model,cpu,memory,mtime"

    items: List[Tuple[str, bool]] = []
    alias = {
        "timestamp": "mtime",
        "time": "mtime",
        "cpu_model": "cpu",
        "cpu": "cpu",
        "mem": "memory",
        "ram": "memory",
    }

    for raw in [x.strip() for x in spec.split(",") if x.strip()]:
        desc = raw.startswith("-")
        k = raw[1:] if desc else raw
        k = k.strip().lower()
        k2 = alias.get(k, k)
        if k2 in {"task", "suite", "run", "model", "cpu", "memory", "mtime"}:
            items.append((k2, desc))

    if not items:
        items = [("task", False), ("suite", False), ("run", False), ("model", False), ("cpu", False), ("memory", False), ("mtime", False)]

    out = list(runs)

    def key_fn(key: str):
        if key == "task":
            return lambda r: r.ref.task
        if key == "suite":
            return lambda r: r.ref.suite
        if key == "run":
            return lambda r: r.ref.run_id
        if key == "mtime":
            return lambda r: float(r.mtime or 0.0)

        # meta-backed keys
        def _meta_get(r: RunInfo) -> str:
            m = _extract_run_meta(r)
            return (m.get(key) or "").lower()

        return _meta_get

    # Apply stable sorts from least significant to most significant.
    for k, desc in reversed(items):
        out.sort(key=key_fn(k), reverse=bool(desc))
    return out


def _maybe_autogen_analysis(*, scale_test_root: Path, run: RunInfo, required: List[Path]) -> Tuple[bool, str]:
    """Ensure required analysis artifacts exist.

    If missing, try to run task-specific analyze_run.py:
      scripts/scale-test/<task>/analyze_run.py <run_dir>

    Returns (ok, message). Always best-effort; never raises.
    """
    # If everything exists, skip.
    if all(p.exists() for p in required):
        return True, ""

    autogen = os.environ.get("SCALE_TEST_WEB_AUTOGEN", "1").strip().lower() not in {"0", "false", "no"}
    if not autogen:
        return False, "autogen disabled (set SCALE_TEST_WEB_AUTOGEN=1 to enable)"

    analyzer = (scale_test_root / run.ref.task / "analyze_run.py").resolve()
    if not analyzer.exists():
        return False, f"analyzer not found: {analyzer}"

    lock = run.analysis_dir / ".autogen.lock"
    # Acquire lock (best-effort). If another thread/process is generating, wait briefly.
    lock_fd: Optional[int] = None
    for _ in range(30):
        try:
            lock_fd = os.open(str(lock), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(lock_fd, str(os.getpid()).encode("utf-8"))
            break
        except FileExistsError:
            time.sleep(0.25)
            if all(p.exists() for p in required):
                return True, ""
        except Exception as e:
            return False, f"lock error: {e}"

    if lock_fd is None:
        # Timed out waiting for another generator.
        return all(p.exists() for p in required), "autogen busy (lock held)"

    try:
        # Ensure analysis dir exists.
        run.analysis_dir.mkdir(parents=True, exist_ok=True)
        cmd = [sys.executable, str(analyzer), str(run.run_dir)]
        proc = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            timeout=600,
        )
        ok = proc.returncode == 0 and all(p.exists() for p in required)
        msg = "" if ok else (proc.stdout[-4000:] if proc.stdout else f"analyze rc={proc.returncode}")
        # Save last autogen log for debugging.
        try:
            (run.analysis_dir / "web_autogen.log").write_text(proc.stdout or "", encoding="utf-8")
        except Exception:
            pass
        return ok, msg
    except subprocess.TimeoutExpired:
        return False, "analyze_run.py timed out"
    except Exception as e:
        return False, f"autogen error: {e}"
    finally:
        try:
            os.close(lock_fd)
        except Exception:
            pass
        try:
            lock.unlink()
        except Exception:
            pass


def _csv_numeric_summary(path: Path, *, max_rows: int = 200_000) -> Dict[str, Any]:
    """Dependency-free numeric summary for a CSV.

    Returns:
      {
        rows: int,
        cols: int,
        numeric: {col: {count, mean, min, max}}
      }
    """
    text = _read_text(path, max_bytes=20_000_000)
    buf = io.StringIO(text)
    reader = csv.reader(buf)
    header = next(reader, None)
    if not header:
        return {"rows": 0, "cols": 0, "numeric": {}}

    ncols = len(header)
    # Welford-ish mean + min/max
    count: Dict[int, int] = {}
    mean: Dict[int, float] = {}
    vmin: Dict[int, float] = {}
    vmax: Dict[int, float] = {}
    rows = 0

    for row in reader:
        rows += 1
        if rows > max_rows:
            break
        # pad
        if len(row) < ncols:
            row = row + [""] * (ncols - len(row))
        for i, cell in enumerate(row[:ncols]):
            s = str(cell).strip()
            if not s:
                continue
            try:
                x = float(s)
            except Exception:
                continue
            c = count.get(i, 0) + 1
            count[i] = c
            if c == 1:
                mean[i] = x
                vmin[i] = x
                vmax[i] = x
            else:
                mean[i] = mean[i] + (x - mean[i]) / c
                if x < vmin[i]:
                    vmin[i] = x
                if x > vmax[i]:
                    vmax[i] = x

    numeric: Dict[str, Any] = {}
    for i, c in count.items():
        col = header[i]
        numeric[col] = {
            "count": c,
            "mean": mean.get(i),
            "min": vmin.get(i),
            "max": vmax.get(i),
        }
    return {"rows": rows, "cols": ncols, "numeric": numeric}


def _safe_float(s: Any) -> Optional[float]:
    try:
        if s is None:
            return None
        st = str(s).strip()
        if not st or st.lower() in {"nan", "none"}:
            return None
        return float(st)
    except Exception:
        return None


def _median(vals: List[float]) -> Optional[float]:
    if not vals:
        return None
    xs = sorted(vals)
    n = len(xs)
    mid = n // 2
    if n % 2 == 1:
        return xs[mid]
    return 0.5 * (xs[mid - 1] + xs[mid])


def _quantile(vals: List[float], q: float) -> Optional[float]:
    if not vals:
        return None
    xs = sorted(vals)
    if q <= 0:
        return xs[0]
    if q >= 1:
        return xs[-1]
    pos = (len(xs) - 1) * q
    lo = int(pos)
    hi = min(lo + 1, len(xs) - 1)
    frac = pos - lo
    return xs[lo] * (1 - frac) + xs[hi] * frac


def _pearson_corr(xs: List[float], ys: List[float]) -> Optional[float]:
    if len(xs) != len(ys) or len(xs) < 2:
        return None
    n = len(xs)
    mx = sum(xs) / n
    my = sum(ys) / n
    num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    denx = sum((x - mx) ** 2 for x in xs)
    deny = sum((y - my) ** 2 for y in ys)
    if denx <= 0 or deny <= 0:
        return None
    return num / (denx**0.5 * deny**0.5)


def _pill(label: str, level: str) -> str:
    # level: ok|warn|strong
    return f'<span class="pill {level}">{_e(label)}</span>'


def _impact_from_ratio(r: Optional[float]) -> Tuple[str, str]:
    if r is None:
        return ("", "")
    if r >= 3.0:
        return ("高", "strong")
    if r >= 1.5:
        return ("中", "warn")
    return ("低", "ok")


def _impact_from_eff(e: Optional[float]) -> Tuple[str, str]:
    if e is None:
        return ("", "")
    if e >= 0.80:
        return ("好", "ok")
    if e >= 0.60:
        return ("一般", "warn")
    return ("较差", "strong")


def _render_scaling_csv_ppt_summary(csv_path: Path) -> str:
    """PPT-ready one-page summary for *_scaling.csv (stdlib-only).

    Returns an HTML fragment.
    """
    name = csv_path.name
    text = _read_text(csv_path, max_bytes=8_000_000)
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    if not rows or not reader.fieldnames:
        return '<div class="muted">Empty CSV.</div>'

    cols = set(reader.fieldnames)

    def _norm_key(v: Any) -> str:
        fv = _safe_float(v)
        if fv is None:
            return str(v).strip()
        if abs(fv - int(fv)) < 1e-9:
            return str(int(fv))
        return f"{fv:g}"

    # Determine schema per CSV.
    if name == "token_len_scaling.csv":
        x = "token_len"
        group_cols = ["cpu_count", "kv_cap", "batch_size"]
        y = "tokens_per_sec" if "tokens_per_sec" in cols else "tps"
        corr_y = "tokens_per_sec" if "tokens_per_sec" in cols else None
        takeaway = "token_len 改变会显著影响单位时间可处理 token 数；通常 token_len 越大 tokens/sec 越低。"
    elif name == "batch_size_scaling.csv":
        x = "batch_size"
        group_cols = ["cpu_count", "kv_cap", "token_len"]
        y = "tps"
        corr_y = None
        takeaway = "batch_size 主要影响吞吐上限；总体影响常小于 token_len / kv_cap。"
    elif name == "kv_scaling.csv":
        x = "kv_cap"
        group_cols = ["cpu_count", "batch_size", "token_len"]
        y = "tps"
        corr_y = None
        takeaway = "KV cap 往往是关键瓶颈维度；kv_cap 不足会显著拉低 TPS。"
    elif name == "cpu_scaling.csv":
        # special: speedup/efficiency
        need = {"cpu_count", "tps", "kv_cap", "batch_size", "token_len"}
        if not need.issubset(cols):
            missing = ", ".join(sorted(need - cols))
            return f'<div class="muted">Missing required columns: {_e(missing)}</div>'
        groups: Dict[Tuple[str, str, str], List[Tuple[float, float]]] = {}
        for r in rows:
            cpu = _safe_float(r.get("cpu_count"))
            tps = _safe_float(r.get("tps"))
            if cpu is None or tps is None:
                continue
            key = (_norm_key(r.get("kv_cap")), _norm_key(r.get("batch_size")), _norm_key(r.get("token_len")))
            groups.setdefault(key, []).append((cpu, tps))

        speedups: List[float] = []
        effs: List[float] = []
        for pts in groups.values():
            pts2 = sorted(pts, key=lambda t: t[0])
            if len({p[0] for p in pts2}) < 2:
                continue
            min_cpu, tps_min = pts2[0]
            max_cpu, tps_max = pts2[-1]
            if min_cpu <= 0 or max_cpu <= 0 or tps_min <= 0:
                continue
            speedup = tps_max / tps_min
            ideal = max_cpu / min_cpu
            eff = speedup / ideal
            speedups.append(speedup)
            effs.append(eff)

        med_speed = _median(speedups)
        med_eff = _median(effs)
        lvl, cls = _impact_from_eff(med_eff)
        rec = ""
        if med_eff is not None:
            if med_eff >= 0.80:
                rec = "CPU 扩展性较好：加核能有效提吞吐；注意 NUMA/绑核一致性。"
            elif med_eff >= 0.60:
                rec = "CPU 扩展性一般：优先排查线程/NUMA/带宽瓶颈，再决定是否加核。"
            else:
                rec = "CPU 扩展性较差：单纯加核收益有限，优先解决瓶颈（kv_cap/内存/调度）。"

        items = [
            f"<li><b>一句话结论</b>：CPU 核数增加能提升吞吐，但总体呈次线性；需要关注并行效率。</li>",
        ]
        if lvl:
            items.append(f"<li><b>扩展性评价</b>：{_pill(lvl, cls)}（以效率衡量）</li>")
        items.append(f"<li><b>统计口径</b>：rows={len(rows)}，groups={len(groups)}（按 kv_cap,batch_size,token_len 分组）</li>")
        if med_speed is not None:
            items.append(f"<li><b>中位加速比</b>：min→max cpu_count 的 median speedup=<b>{med_speed:.2f}×</b></li>")
        if med_eff is not None:
            items.append(f"<li><b>中位并行效率</b>：median efficiency vs ideal linear=<b>{med_eff:.2f}</b></li>")
        if rec:
            items.append(f"<li><b>建议</b>：{_e(rec)}</li>")
        return f"<ul>{''.join(items)}</ul>"
    else:
        return '<div class="muted">Not a scaling CSV.</div>'

    # Generic (token/batch/kv) summary with grouping.
    need = {x, y, *group_cols}
    if not need.issubset(cols):
        missing = ", ".join(sorted(need - cols))
        return f'<div class="muted">Missing required columns: {_e(missing)}</div>'

    groups2: Dict[Tuple[str, ...], List[Tuple[float, float, Optional[float]]]] = {}
    for r in rows:
        xv = _safe_float(r.get(x))
        yv = _safe_float(r.get(y))
        if xv is None or yv is None:
            continue
        cy: Optional[float] = None
        if corr_y is not None:
            cy = _safe_float(r.get(corr_y))
        key = tuple(_norm_key(r.get(c)) for c in group_cols)
        groups2.setdefault(key, []).append((xv, yv, cy))

    best_xs: List[float] = []
    ratios: List[float] = []
    cors: List[float] = []
    for pts in groups2.values():
        # require >=2 distinct x
        if len({p[0] for p in pts}) < 2:
            continue
        # best-x by y
        px = max(pts, key=lambda t: t[1])[0]
        best_xs.append(px)
        ys = [p[1] for p in pts]
        ymin = min(ys)
        ymax = max(ys)
        if ymin > 0:
            ratios.append(ymax / ymin)
        if corr_y is not None:
            xs = [p[0] for p in pts if p[2] is not None]
            ys2 = [p[2] for p in pts if p[2] is not None]
            if len(xs) >= 2 and len(set(xs)) >= 2:
                cc = _pearson_corr(xs, ys2)
                if cc is not None:
                    cors.append(cc)

    med_best = _median(best_xs)
    med_ratio = _median(ratios)
    p90_ratio = _quantile(ratios, 0.90)
    med_corr = _median(cors)

    lvl, cls = _impact_from_ratio(med_ratio)
    rec = ""
    if name == "kv_scaling.csv" and med_ratio is not None and med_ratio >= 1.5:
        rec = "KV cap 影响较大：优先确认不会被 KV 上限卡住（必要时提升 kv_cap / SGLANG_MAX_TOTAL_TOKENS）。"
    elif name == "batch_size_scaling.csv" and med_ratio is not None and med_ratio < 1.2:
        rec = "batch_size 总体影响较小：可优先以延迟与内存压力为约束来选 batch_size。"
    elif name == "token_len_scaling.csv" and med_corr is not None and med_corr <= -0.3:
        rec = "token_len 增大时 tokens/sec 往往下降：建议按目标场景 token_len 分层对比，避免用单一 token_len 代表全部。"
    elif med_best is not None:
        rec = f"建议优先选择 {x}≈{int(round(med_best))}（按当前 sweep 的中位最优点）。"

    items2: List[str] = []
    items2.append(f"<li><b>一句话结论</b>：{_e(takeaway)}</li>")
    if lvl:
        items2.append(f"<li><b>总体影响等级</b>：{_pill(lvl, cls)}（以组内 max/min 衡量影响强度）</li>")
    items2.append(f"<li><b>统计口径</b>：rows={len(rows)}，groups={len(groups2)}（按 {', '.join(group_cols)} 分组）</li>")
    if med_best is not None:
        items2.append(f"<li><b>中位最优点</b>：{_e(x)}≈<b>{med_best:.0f}</b>（argmax {_e(y)}）</li>")
    if med_ratio is not None:
        p90_s = f"{p90_ratio:.2f}×" if p90_ratio is not None else ""
        items2.append(f"<li><b>影响幅度</b>：组内 {_e(y)} 的 median(max/min)=<b>{med_ratio:.2f}×</b>" + (f"（p90 {p90_s}）" if p90_s else "") + "</li>")
    if corr_y is not None and med_corr is not None:
        items2.append(f"<li><b>趋势</b>：median corr({_e(x)}, {_e(corr_y)})=<b>{med_corr:.2f}</b></li>")
    if rec:
        items2.append(f"<li><b>建议</b>：{_e(rec)}</li>")
    return f"<ul>{''.join(items2)}</ul>"


def _parse_cpu_expr(cpu_expr: str) -> Optional[int]:
    """Best-effort parse of CPU expr like '0-15' or '0,2,4' into count."""
    s = str(cpu_expr).strip()
    if not s:
        return None
    # Prefer simple N-M ranges.
    if "," not in s and "-" in s:
        try:
            a, b = s.split("-", 1)
            a_i = int(a.strip())
            b_i = int(b.strip())
            if b_i >= a_i:
                return b_i - a_i + 1
        except Exception:
            pass
    # Fallback: count comma separated.
    try:
        parts = [x.strip() for x in s.split(",") if x.strip()]
        if parts:
            return len(parts)
    except Exception:
        pass
    return None


def _html_page(title: str, body: str) -> str:
    css_href = "/static/style.css"
    return (
        "<!doctype html>\n"
        "<html>\n"
        "  <head>\n"
        '    <meta charset="utf-8" />\n'
        '    <meta name="viewport" content="width=device-width, initial-scale=1" />\n'
        f"    <title>{_e(title)}</title>\n"
        f'    <link rel="stylesheet" href="{css_href}" />\n'
        "  </head>\n"
        "  <body>\n"
        "    <header class=\"topbar\">\n"
        "      <div class=\"container\">\n"
        "        <div class=\"brand\"><a href=\"/\">Scale Test Results</a></div>\n"
        "        <div class=\"hint\">auto-discovered from scripts/scale-test/*/result/**/analysis</div>\n"
        "      </div>\n"
        "    </header>\n"
        "    <main class=\"container\">\n"
        f"{body}\n"
        "    </main>\n"
        "    <footer class=\"footer\">\n"
        "      <div class=\"container\">Generated by scripts/scale-test/web/server.py</div>\n"
        "    </footer>\n"
        "  </body>\n"
        "</html>\n"
    )


def _render_home(scale_test_root: Path, runs: List[RunInfo], q: Dict[str, List[str]]) -> str:
    task_filter = (q.get("task") or [""])[0].strip()
    suite_filter = (q.get("suite") or [""])[0].strip()
    run_filter = (q.get("run") or [""])[0].strip()
    model_filter = (q.get("model") or [""])[0].strip()
    cpu_filter = (q.get("cpu") or [""])[0].strip()
    sort_spec = (q.get("sort") or [""])[0].strip() or "task,suite,run,model,cpu,memory,mtime"
    limit_s = (q.get("limit") or ["50"])[0].strip()
    try:
        limit = max(1, min(500, int(limit_s)))
    except Exception:
        limit = 50

    # Apply sorting before filtering/limit (so pagination/limits make sense).
    runs2 = _apply_home_sort(runs, sort_spec=sort_spec)

    filtered = []
    for r in runs2:
        if task_filter and r.ref.task != task_filter:
            continue
        if suite_filter and r.ref.suite != suite_filter:
            continue
        if run_filter and run_filter.lower() not in r.ref.run_id.lower():
            continue
        if model_filter:
            meta = _extract_run_meta(r)
            m = (meta.get("model") or "").lower()
            if model_filter.lower() not in m:
                continue
        if cpu_filter:
            meta = _extract_run_meta(r)
            c = (meta.get("cpu") or "").lower()
            if cpu_filter.lower() not in c:
                continue
        filtered.append(r)

    # Task summary
    by_task: Dict[str, int] = {}
    for r in runs:
        by_task[r.ref.task] = by_task.get(r.ref.task, 0) + 1

    tasks_html = "".join(
        f'<a class="pill" href="/?task={_e(t)}">{_e(t)} <span class="pill-count">{n}</span></a>'
        for t, n in sorted(by_task.items())
    )

    rows = []
    for r in filtered[:limit]:
        meta = _extract_run_meta(r)
        model = meta.get("model") or "-"
        cpu = meta.get("cpu") or "-"
        memory = meta.get("memory") or "-"
        href = f"/run/{_e(r.ref.task)}/{_e(r.ref.suite)}/{_e(r.ref.run_id)}"
        run_key = f"{r.ref.task}/{r.ref.suite}/{r.ref.run_id}"
        rows.append(
            "<tr>"
            f"<td><input type=\"checkbox\" name=\"sel\" value=\"{_e(run_key)}\" /></td>"
            f"<td><a href=\"{href}\">{_e(r.ref.task)}</a></td>"
            f"<td>{_e(r.ref.suite)}</td>"
            f"<td><a href=\"{href}\">{_e(r.ref.run_id)}</a></td>"
            f"<td class=\"mono\">{_e(model)}</td>"
            f"<td class=\"mono\">{_e(cpu)}</td>"
            f"<td class=\"mono\">{_e(memory)}</td>"
            f"<td class=\"mono\">{_e(_fmt_dt(r.mtime))}</td>"
            "</tr>"
        )

    def _q_href(**updates: str) -> str:
        # Build a query string preserving current filters but overriding some keys.
        keep = {
            "task": task_filter,
            "suite": suite_filter,
            "run": run_filter,
            "model": model_filter,
            "cpu": cpu_filter,
            "limit": str(limit),
        }
        keep.update({k: v for k, v in updates.items()})
        parts = []
        for k, v in keep.items():
            if v is None:
                continue
            vs = str(v).strip()
            if not vs:
                continue
            parts.append(f"{quote(k)}={quote(vs)}")
        return "/" + ("?" + "&".join(parts) if parts else "")

    sort_bar = (
        '<div class="sub">'
        f"Sort: <span class=\"mono\">{_e(sort_spec)}</span> &nbsp;"
        + f"<a href=\"{_q_href(sort='task,suite,run,model,cpu,memory,mtime')}\">group</a> · "
        + f"<a href=\"{_q_href(sort='task')}\">task↑</a> · <a href=\"{_q_href(sort='-task')}\">task↓</a> · "
        + f"<a href=\"{_q_href(sort='suite')}\">suite↑</a> · <a href=\"{_q_href(sort='-suite')}\">suite↓</a> · "
        + f"<a href=\"{_q_href(sort='-mtime')}\">time↓</a> · <a href=\"{_q_href(sort='mtime')}\">time↑</a> · "
        + f"<a href=\"{_q_href(sort='model')}\">model↑</a> · <a href=\"{_q_href(sort='-model')}\">model↓</a> · "
        + f"<a href=\"{_q_href(sort='cpu')}\">cpu↑</a> · <a href=\"{_q_href(sort='-cpu')}\">cpu↓</a> · "
        + f"<a href=\"{_q_href(sort='memory')}\">mem↑</a> · <a href=\"{_q_href(sort='-memory')}\">mem↓</a>"
        + "</div>"
    )

    deleted_n = (q.get("deleted") or [""])[0].strip()
    failed_n = (q.get("delete_failed") or [""])[0].strip()
    delete_msg = ""
    if deleted_n.isdigit():
        msg = f"Deleted {int(deleted_n)} run(s)."
        if failed_n.isdigit() and int(failed_n) > 0:
            msg += f" Failed: {int(failed_n)}."
        delete_msg = f'<div class="sub warn">{_e(msg)}</div>'

    body = f"""
<section class="card">
  <h1>Scale Test Results</h1>
  <div class="sub">Root: <span class="mono">{_e(scale_test_root)}</span></div>
    {delete_msg}
  {sort_bar}

  <div class="toolbar">
    <form method="get" action="/" class="form-inline">
      <label>task <input name="task" value="{_e(task_filter)}" placeholder="embedding / vl / omni" /></label>
      <label>suite <input name="suite" value="{_e(suite_filter)}" placeholder="fix_token_len" /></label>
            <label>run <input name="run" value="{_e(run_filter)}" placeholder="20260210T..." /></label>
            <label>model <input name="model" value="{_e(model_filter)}" placeholder="qwen3-embedding-4b / Qwen3-Embedding-4B" /></label>
            <label>cpu <input name="cpu" value="{_e(cpu_filter)}" placeholder="Xeon / EPYC" /></label>
            <label>sort <input name="sort" value="{_e(sort_spec)}" placeholder="task,suite,run,model,cpu,memory,mtime" size="36" /></label>
      <label>limit <input name="limit" value="{_e(limit)}" size="4" /></label>
      <button type="submit">Filter</button>
      <a class="btn" href="/">Reset</a>
    </form>
  </div>

  <div class="pills">{tasks_html}</div>
</section>

<section class="card">
  <h2>Recent Runs</h2>
    <div class="sub">Select any two runs (checkbox left of task), then click <span class="mono">Compare selected</span>.</div>
    <form method="get" action="/compare">
        <table class="table">
            <thead>
                <tr>
                    <th></th>
                    <th>task</th>
                    <th>suite</th>
                    <th>run</th>
                    <th>model</th>
                    <th>cpu</th>
                    <th>memory</th>
                    <th>mtime</th>
                </tr>
            </thead>
            <tbody>
                {''.join(rows) if rows else '<tr><td colspan="8" class="muted">No runs found.</td></tr>'}
            </tbody>
        </table>
        <div style="margin-top: 10px; display: flex; gap: 10px; align-items: center; flex-wrap: wrap;">
            <button type="submit">Compare selected</button>
            <button
              type="submit"
              class="danger"
              formaction="/delete"
              formmethod="post"
              onclick="return confirm('Delete selected runs? This cannot be undone.');"
            >Delete selected</button>
        </div>
    </form>
</section>
"""
    return _html_page("Scale Test Results", body)


def _pick_lines(text: str, *, keep_prefixes: Iterable[str], max_lines: int = 32) -> str:
    prefixes = [str(p).lower() for p in keep_prefixes]
    out: List[str] = []
    for line in (text or "").splitlines():
        low = line.strip().lower()
        if not low:
            continue
        for p in prefixes:
            if low.startswith(p):
                out.append(line.rstrip())
                break
        if len(out) >= max_lines:
            break
    return "\n".join(out)


def _render_server_info_brief(run: RunInfo) -> str:
    hosts = _list_server_info_hosts(run)
    if not hosts:
        return '<div class="muted">No server_info found for this run.</div>'

    blocks: List[str] = []
    for it in hosts:
        tag = str(it.get("tag") or "")
        label = str(it.get("label") or tag or "host")
        base_rel = str(it.get("base_rel") or "")
        lscpu = (run.run_dir / base_rel / "lscpu.txt").resolve()
        if not (_is_within(lscpu, run.run_dir) and lscpu.exists() and lscpu.is_file()):
            blocks.append(f'<div class="muted">{_e(label)}: missing lscpu.txt</div>')
            continue

        raw = _read_text(lscpu, max_bytes=256_000)
        picked = _pick_lines(
            raw,
            keep_prefixes=(
                "Architecture:",
                "Model name:",
                "CPU(s):",
                "Thread(s) per core:",
                "Core(s) per socket:",
                "Socket(s):",
                "NUMA node(s):",
                "NUMA node0 CPU(s):",
                "NUMA node1 CPU(s):",
                "NUMA node2 CPU(s):",
                "NUMA node3 CPU(s):",
            ),
            max_lines=48,
        )
        if not picked:
            picked = "\n".join((raw or "").splitlines()[:24])

        host_hdr = "" if len(hosts) == 1 else f'<div class="sub"><span class="mono">{_e(label)}</span></div>'
        blocks.append(
            host_hdr
            + f'<pre class="mono" style="white-space: pre-wrap; margin: 8px 0 0 0;">{_e(picked)}</pre>'
        )

    return "".join(blocks)


def _render_cpu_info_brief(run: RunInfo) -> str:
    """Render a CPU-focused excerpt from server_info/lscpu.txt.

    For multi-host runs, shows one block per host.
    """

    hosts = _list_server_info_hosts(run)
    if not hosts:
        return '<div class="muted">No server_info found for this run.</div>'

    blocks: List[str] = []
    for it in hosts:
        tag = str(it.get("tag") or "")
        label = str(it.get("label") or tag or "host")
        base_rel = str(it.get("base_rel") or "")
        lscpu = (run.run_dir / base_rel / "lscpu.txt").resolve()
        if not (_is_within(lscpu, run.run_dir) and lscpu.exists() and lscpu.is_file()):
            blocks.append(f'<div class="muted">{_e(label)}: missing lscpu.txt</div>')
            continue

        raw = _read_text(lscpu, max_bytes=256_000)
        picked = _pick_lines(
            raw,
            keep_prefixes=(
                "Architecture:",
                "Model name:",
                "CPU(s):",
                "Thread(s) per core:",
                "Core(s) per socket:",
                "Socket(s):",
                "NUMA node(s):",
            ),
            max_lines=24,
        )
        if not picked:
            picked = "\n".join((raw or "").splitlines()[:16])

        host_hdr = "" if len(hosts) == 1 else f'<div class="sub"><span class="mono">{_e(label)}</span></div>'
        blocks.append(
            host_hdr
            + f'<pre class="mono" style="white-space: pre-wrap; margin: 8px 0 0 0;">{_e(picked)}</pre>'
        )

    return "".join(blocks)


def _render_run(scale_test_root: Path, run: RunInfo, q: Dict[str, List[str]]) -> str:
    _, csvs = _list_analysis_files(run.analysis_dir)

    meta = _extract_run_meta(run)
    model = meta.get("model") or "-"
    cpu = meta.get("cpu") or "-"

    run_key = f"{run.ref.task}/{run.ref.suite}/{run.ref.run_id}"
    compare_bar = (
        '<div class="sub">Compare: '
        f'<a class="btn" href="/compare?a={quote(run_key)}">Set as A</a> '
        f'<a class="btn" href="/compare?b={quote(run_key)}">Set as B</a>'
        "</div>"
    )

    # Auto-generate summary/plots if missing.
    summary_path = (run.analysis_dir / "run_summary.html").resolve()
    # Also require a representative scaling CSV for embedding on the run page.
    scaling_csv = (run.analysis_dir / "token_len_scaling.csv").resolve()
    ok, msg = _maybe_autogen_analysis(scale_test_root=scale_test_root, run=run, required=[summary_path, scaling_csv])

    # Run-level summary (pre-generated by analyze_run.py)
    summary_html = ""
    if _is_within(summary_path, run.analysis_dir) and summary_path.exists():
        # Embed as-is; content is generated by our analysis script.
        summary_html = _read_text(summary_path, max_bytes=2_000_000)
    else:
        hint = '<div class="muted">No run_summary.html found. Run: analyze_run.py</div>'
        if msg:
            hint += f'<div class="warn mono">autogen: {_e(msg)}</div>'
        summary_html = hint


    # Embed a scaling CSV table directly on the run page.
    scaling_embed_html = '<div class="muted">No scaling CSV preview available.</div>'
    if _is_within(scaling_csv, run.analysis_dir) and scaling_csv.exists() and scaling_csv.is_file():
        rel_scaling = scaling_csv.relative_to(run.run_dir).as_posix()
        scaling_raw = f"/raw/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(rel_scaling)}"
        scaling_preview = f"/csv/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(scaling_csv.name)}"
        scaling_embed_html = (
            '<div class="sub">Embedded table: <span class="mono">token_len_scaling.csv</span> '
            '(same row-set as <span class="mono">batch_size_scaling.csv</span>/<span class="mono">cpu_scaling.csv</span>/<span class="mono">kv_scaling.csv</span>; only sorted differently) '
            f'• <a href="{scaling_preview}">Full preview</a> '
            f'• <a href="{scaling_raw}" target="_blank">Download</a>'
            "</div>"
            + _render_csv_preview(scaling_csv, max_rows=200, run=run)
        )

    # CSV list
    csv_rows = []
    for c in csvs:
        rel = c.relative_to(run.run_dir).as_posix()
        raw = f"/raw/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(rel)}"
        preview = f"/csv/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(c.name)}"
        csv_rows.append(
            "<tr>"
            f"<td class=\"mono\">{_e(c.name)}</td>"
            f"<td><a href=\"{preview}\">Preview</a></td>"
            f"<td><a href=\"{raw}\" target=\"_blank\">Download</a></td>"
            "</tr>"
        )

    title = f"{run.ref.task} / {run.ref.suite} / {run.ref.run_id}"
    server_info_href = f"/server-info/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}"
    body = f"""
<section class="card">
  <div class="breadcrumbs">
    <a href="/">Home</a>
    <span class="sep">/</span>
    <span class="mono">{_e(run.ref.task)}</span>
    <span class="sep">/</span>
    <span class="mono">{_e(run.ref.suite)}</span>
    <span class="sep">/</span>
    <span class="mono">{_e(run.ref.run_id)}</span>
  </div>

  <h1>{_e(title)}</h1>
    <div class="sub">model: <span class="mono">{_e(model)}</span></div>
    <div class="sub">cpu: <span class="mono">{_e(cpu)}</span></div>
  <div class="sub">mtime: <span class="mono">{_e(_fmt_dt(run.mtime))}</span></div>
  <div class="sub">run_dir: <span class="mono">{_e(run.run_dir)}</span></div>
    <div class="sub">server info: <a href="{server_info_href}">View</a></div>
    {compare_bar}
</section>

<section class="card">
    <h2>Summary</h2>
    <div class="embed">{summary_html}</div>
</section>

<section class="card">
  <h2>CSVs</h2>
    {scaling_embed_html}
  <table class="table">
    <thead><tr><th>file</th><th>preview</th><th>download</th></tr></thead>
    <tbody>
      {''.join(csv_rows) if csv_rows else '<tr><td colspan="3" class="muted">No CSVs found.</td></tr>'}
    </tbody>
  </table>
</section>
"""
    return _html_page(title, body)


def _parse_run_key(s: str) -> Optional[RunRef]:
    raw = (s or "").strip()
    if not raw:
        return None
    # Accept either task/suite/run or task|suite|run
    if "|" in raw:
        parts = [p.strip() for p in raw.split("|") if p.strip()]
    else:
        parts = [p.strip() for p in raw.split("/") if p.strip()]
    if len(parts) != 3:
        return None
    return RunRef(task=parts[0], suite=parts[1], run_id=parts[2])


def _try_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _cmp_pct(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if a is None or b is None:
        return None
    if b == 0.0:
        if a == 0.0:
            return 0.0
        return float("inf")
    return (a / b - 1.0) * 100.0


def _fmt_pct(v: Optional[float]) -> str:
    if v is None:
        return "-"
    if v == float("inf"):
        return "inf"
    if v == float("-inf"):
        return "-inf"
    try:
        return f"{v:+.2f}%"
    except Exception:
        return str(v)


def _read_csv_dict_rows(path: Path, *, max_rows: int = 50_000) -> Tuple[List[str], List[Dict[str, str]]]:
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = [c for c in (reader.fieldnames or [])]
        rows: List[Dict[str, str]] = []
        for i, r in enumerate(reader):
            if i >= max_rows:
                break
            rows.append({k: ("" if v is None else str(v)) for k, v in r.items()})
    return fieldnames, rows


def _build_compare_csv_export_xlsx(
    *,
    a_path: Path,
    b_path: Path,
    a_raw: str,
    b_raw: str,
    csv_name: str,
    run_a: Optional[RunInfo] = None,
    run_b: Optional[RunInfo] = None,
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise RuntimeError(f"openpyxl is required for XLSX export: {e}")

    if not a_path.exists() or not b_path.exists():
        raise RuntimeError("Both A and B CSV files are required for export")

    a_cols, a_rows = _read_csv_dict_rows(a_path)
    b_cols, b_rows = _read_csv_dict_rows(b_path)
    cols = [c for c in a_cols if c in set(b_cols)]
    if not cols:
        cols = list(dict.fromkeys(a_cols + b_cols))

    def _median(vals: List[float]) -> Optional[float]:
        if not vals:
            return None
        s = sorted(vals)
        n = len(s)
        m = n // 2
        if n % 2 == 1:
            return s[m]
        return (s[m - 1] + s[m]) / 2.0

    def _safe_sheet_name(name: str) -> str:
        bad = set('[]:*?/\\')
        out = "".join(ch if ch not in bad else "_" for ch in (name or "sheet"))
        out = out.strip() or "sheet"
        return out[:31]

    def _write_rows(ws: Any, rows: List[List[Any]]) -> None:
        for row in rows:
            ws.append(row)

    wb = Workbook()
    wb.remove(wb.active)

    meta_a = _extract_run_meta(run_a) if run_a is not None else {}
    meta_b = _extract_run_meta(run_b) if run_b is not None else {}

    ws_info = wb.create_sheet("Info")
    _write_rows(
        ws_info,
        [
            ["field", "value"],
            ["export_time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["csv", csv_name],
            ["A", a_raw],
            ["B", b_raw],
            ["A_cpu", meta_a.get("cpu", "")],
            ["B_cpu", meta_b.get("cpu", "")],
            ["A_memory", meta_a.get("memory", "")],
            ["B_memory", meta_b.get("memory", "")],
            ["A_path", str(a_path)],
            ["B_path", str(b_path)],
            ["A_rows", len(a_rows)],
            ["B_rows", len(b_rows)],
            ["common_cols", len(cols)],
        ],
    )

    def _server_info_rows(run: Optional[RunInfo]) -> List[List[Any]]:
        rows: List[List[Any]] = [["host_tag", "host_label", "cpu_model", "memory", "lscpu_path", "meminfo_path"]]
        if run is None:
            rows.append(["", "", "", "", "", ""])
            return rows

        def _cpu_model_from_lscpu_text(s: str) -> str:
            try:
                for line in (s or "").splitlines():
                    if line.lower().startswith("model name"):
                        _, _, tail = line.partition(":")
                        v = tail.strip()
                        if v:
                            return v
            except Exception:
                return ""
            return ""

        def _mem_from_meminfo(s: str) -> str:
            try:
                for line in (s or "").splitlines():
                    if line.lower().startswith("memtotal:"):
                        m = re.search(r"MemTotal:\s*(\d+)\s*kB", line, re.IGNORECASE)
                        if not m:
                            continue
                        kb = int(m.group(1))
                        gib = kb / (1024.0 * 1024.0)
                        return f"{gib:.1f} GiB"
            except Exception:
                return ""
            return ""

        def _mem_from_lscpu(s: str) -> str:
            try:
                for line in (s or "").splitlines():
                    if line.lower().startswith("memory:"):
                        _, _, tail = line.partition(":")
                        return tail.strip()
            except Exception:
                return ""
            return ""

        hosts = _list_server_info_hosts(run)
        if not hosts:
            rows.append(["", "", "", "", "", ""])
            return rows

        for it in hosts:
            tag = str(it.get("tag") or "")
            label = str(it.get("label") or tag)
            base_rel = str(it.get("base_rel") or "")
            lscpu_p = (run.run_dir / base_rel / "lscpu.txt").resolve()
            meminfo_p = (run.run_dir / base_rel / "meminfo.txt").resolve()
            cpu_model = ""
            mem = ""
            if _is_within(lscpu_p, run.run_dir) and lscpu_p.exists() and lscpu_p.is_file():
                txt = _read_text(lscpu_p, max_bytes=256_000)
                cpu_model = _cpu_model_from_lscpu_text(txt)
                if not mem:
                    mem = _mem_from_lscpu(txt)
            if _is_within(meminfo_p, run.run_dir) and meminfo_p.exists() and meminfo_p.is_file():
                mtxt = _read_text(meminfo_p, max_bytes=128_000)
                m2 = _mem_from_meminfo(mtxt)
                if m2:
                    mem = m2
            rows.append([tag, label, cpu_model, mem, str(lscpu_p), str(meminfo_p)])
        return rows

    ws_srv_a = wb.create_sheet("ServerInfo_A")
    _write_rows(ws_srv_a, _server_info_rows(run_a))
    ws_srv_b = wb.create_sheet("ServerInfo_B")
    _write_rows(ws_srv_b, _server_info_rows(run_b))

    # Build the same compare table logic used by /compare-csv.
    dim_force = {
        "variant",
        "resource_cpu",
        "resource_cpu_count",
        "resource_mem_gb",
        "cpu_count",
        "batch_size",
        "token_len",
        "kv_cap",
        "sglang_max_total_tokens",
    }

    def is_numeric_col(name: str) -> bool:
        vs: List[str] = []
        for r in a_rows[:200]:
            v = r.get(name, "")
            if v:
                vs.append(v)
        for r in b_rows[:200]:
            v = r.get(name, "")
            if v:
                vs.append(v)
        if not vs:
            return False
        ok = sum(1 for v in vs if _try_float(v) is not None)
        return (ok / max(1, len(vs))) >= 0.85

    numeric_cols = {c for c in cols if is_numeric_col(c)}
    metrics = [c for c in cols if (c in numeric_cols and c not in dim_force)]
    candidate_key_cols = [c for c in cols if c not in set(metrics)]
    ab_only_cols = [c for c in ["resource_cpu", "resource_cpu_count", "cpu_count"] if c in cols]
    candidate_key_cols = [c for c in candidate_key_cols if c not in set(ab_only_cols)]

    volatile_name_re = re.compile(
        r"(path|file|dir|folder|artifact|output|stdout|stderr|log|xlsx|html|json|csv)$",
        re.IGNORECASE,
    )
    volatile_exact = {
        "emon_summary_xlsx",
        "auto_test_stdout_log",
        "auto_test_stderr_log",
        "run_dir",
        "run_path",
        "summary_xlsx",
    }

    def looks_like_path(v: str) -> bool:
        s = (v or "").strip()
        if not s:
            return False
        if "/" in s or "\\" in s or s.startswith("http://") or s.startswith("https://"):
            return True
        lower = s.lower()
        return any(lower.endswith(ext) for ext in (".xlsx", ".csv", ".json", ".html", ".log", ".txt", ".png"))

    def is_volatile_key_col(name: str) -> bool:
        n = (name or "").strip()
        if not n:
            return True
        if n in volatile_exact:
            return True
        if volatile_name_re.search(n):
            return True
        sample: List[str] = []
        for r in a_rows[:200]:
            v = (r.get(n) or "").strip()
            if v:
                sample.append(v)
        for r in b_rows[:200]:
            v = (r.get(n) or "").strip()
            if v:
                sample.append(v)
        if not sample:
            return False
        pathish = sum(1 for v in sample if looks_like_path(v))
        return (pathish / max(1, len(sample))) >= 0.60

    key_cols = [c for c in candidate_key_cols if not is_volatile_key_col(c)]
    if not key_cols:
        key_cols = candidate_key_cols

    def row_key(r: Dict[str, str], *, i: int) -> Tuple[str, ...]:
        if not key_cols:
            return (str(i),)
        return tuple((r.get(c, "") or "").strip() for c in key_cols)

    a_idx: Dict[Tuple[str, ...], Dict[str, str]] = {}
    for i, r in enumerate(a_rows):
        a_idx[row_key(r, i=i)] = r
    b_idx: Dict[Tuple[str, ...], Dict[str, str]] = {}
    for i, r in enumerate(b_rows):
        b_idx[row_key(r, i=i)] = r
    keys = sorted(set(a_idx.keys()) | set(b_idx.keys()))

    ws_cmp = wb.create_sheet("Comparison")
    cmp_head: List[str] = list(key_cols)
    for c in ab_only_cols:
        cmp_head.extend([f"{c}_A", f"{c}_B"])
    for c in metrics:
        cmp_head.extend([f"A_{c}", f"B_{c}", f"delta_{c}%"])
    ws_cmp.append(cmp_head)

    for k in keys[:50_000]:
        ar = a_idx.get(k) or {}
        br = b_idx.get(k) or {}
        row: List[Any] = []
        for j, _ in enumerate(key_cols):
            row.append(k[j] if j < len(k) else "")
        for c in ab_only_cols:
            row.append((ar.get(c, "") or "").strip())
            row.append((br.get(c, "") or "").strip())
        for c in metrics:
            av = _try_float((ar.get(c, "") or "").strip())
            bv = _try_float((br.get(c, "") or "").strip())
            row.append(av)
            row.append(bv)
            row.append(_cmp_pct(av, bv))
        ws_cmp.append(row)

    ws_raw_a = wb.create_sheet(_safe_sheet_name("Raw_A"))
    ws_raw_b = wb.create_sheet(_safe_sheet_name("Raw_B"))
    ws_raw_a.append(a_cols)
    for r in a_rows[:100_000]:
        ws_raw_a.append([r.get(c, "") for c in a_cols])
    ws_raw_b.append(b_cols)
    for r in b_rows[:100_000]:
        ws_raw_b.append([r.get(c, "") for c in b_cols])

    # Export auto-plot data and create line charts for scaling CSVs.
    stem = Path(csv_name).stem
    prefix = stem[:-8] if stem.endswith("_scaling") else stem
    x_candidates = [prefix]
    if "_" in prefix:
        x_candidates.append(prefix.split("_", 1)[0])
    x_col = next((c for c in x_candidates if c in cols), "")
    if not x_col and prefix == "cpu" and "cpu_count" in cols:
        x_col = "cpu_count"
    y_col = "tps" if "tps" in cols else ("tokens_per_sec" if "tokens_per_sec" in cols else "")

    if stem.endswith("_scaling") and x_col and y_col:
        split_pref: Dict[str, List[str]] = {
            "batch_size": ["token_len", "cpu_count", "kv_cap"],
            "token_len": ["batch_size", "cpu_count", "kv_cap"],
            "cpu_count": ["token_len", "batch_size", "kv_cap"],
            "kv_cap": ["token_len", "batch_size", "cpu_count"],
        }
        split_candidates = split_pref.get(x_col, ["token_len", "batch_size", "cpu_count", "kv_cap"])
        split_col = ""
        for c in split_candidates:
            if c == x_col or c not in cols:
                continue
            vals = set()
            for r in (a_rows[:800] + b_rows[:800]):
                v = (r.get(c, "") or "").strip()
                if v:
                    vals.add(v)
            if len(vals) > 1:
                split_col = c
                break

        split_values = [""]
        if split_col:
            vals = set((r.get(split_col, "") or "").strip() for r in (a_rows + b_rows))
            vals.discard("")

            def _sort_key(s: str) -> Tuple[int, float, str]:
                f = _try_float(s)
                if f is None:
                    return (1, 0.0, s)
                return (0, float(f), s)

            split_values = sorted(vals, key=_sort_key)[:6]

        ws_plot = wb.create_sheet("AutoPlotData")
        ws_plot.append(["split", x_col, "A_median", "B_median", "A_n", "B_n"])

        def _series(rows: List[Dict[str, str]], split_value: str = "") -> List[Tuple[float, float, int]]:
            groups: Dict[float, List[float]] = {}
            for r in rows:
                if split_col:
                    sv = (r.get(split_col, "") or "").strip()
                    if sv != split_value:
                        continue
                xv = _try_float((r.get(x_col, "") or "").strip())
                yv = _try_float((r.get(y_col, "") or "").strip())
                if xv is None or yv is None:
                    continue
                groups.setdefault(float(xv), []).append(float(yv))
            out: List[Tuple[float, float, int]] = []
            for xv in sorted(groups.keys()):
                vals = groups.get(xv, [])
                m = _median(vals)
                if m is not None:
                    out.append((xv, m, len(vals)))
            return out

        row0 = 2
        chart_meta: List[Tuple[str, int, int]] = []
        for sv in split_values:
            a_line = _series(a_rows, sv)
            b_line = _series(b_rows, sv)
            if not a_line and not b_line:
                continue
            xvals = sorted(set([x for x, _, _ in a_line] + [x for x, _, _ in b_line]))
            a_map = {x: (y, n) for x, y, n in a_line}
            b_map = {x: (y, n) for x, y, n in b_line}
            start = row0
            for x in xvals:
                ay, an = a_map.get(x, (None, 0))
                by, bn = b_map.get(x, (None, 0))
                ws_plot.append([sv if split_col else "all", x, ay, by, an, bn])
                row0 += 1
            end = row0 - 1
            chart_meta.append((sv if split_col else "all", start, end))
            row0 += 1

        ws_chart = wb.create_sheet("AutoPlotChart")
        ws_chart.append([f"x={x_col}", f"y={y_col} (median)", f"facet={split_col or 'none'}"])
        anchor_row = 3
        for sv, start, end in chart_meta:
            chart = LineChart()
            chart.title = f"{split_col}={sv}" if split_col else "all"
            chart.y_axis.title = y_col
            chart.x_axis.title = x_col
            cats = Reference(ws_plot, min_col=2, min_row=start, max_row=end)
            data_a = Reference(ws_plot, min_col=3, min_row=start - 1, max_row=end)
            data_b = Reference(ws_plot, min_col=4, min_row=start - 1, max_row=end)
            chart.add_data(data_a, titles_from_data=True)
            chart.add_data(data_b, titles_from_data=True)
            chart.set_categories(cats)
            ws_chart.add_chart(chart, f"A{anchor_row}")
            anchor_row += 16

    # Export batch_size x token_len TPS delta matrix and apply heatmap formatting.
    if {"batch_size", "token_len", "tps"}.issubset(set(cols)):
        def _agg(rows: List[Dict[str, str]]) -> Dict[Tuple[float, float], float]:
            g: Dict[Tuple[float, float], List[float]] = {}
            for r in rows:
                bs = _try_float((r.get("batch_size", "") or "").strip())
                tl = _try_float((r.get("token_len", "") or "").strip())
                tps = _try_float((r.get("tps", "") or "").strip())
                if bs is None or tl is None or tps is None:
                    continue
                g.setdefault((float(bs), float(tl)), []).append(float(tps))
            out: Dict[Tuple[float, float], float] = {}
            for k, vals in g.items():
                m = _median(vals)
                if m is not None:
                    out[k] = m
            return out

        amap = _agg(a_rows)
        bmap = _agg(b_rows)
        keys2 = sorted(set(amap.keys()) | set(bmap.keys()), key=lambda k: (k[1], k[0]))
        if keys2:
            batches = sorted(set(k[0] for k in keys2))
            tokens = sorted(set(k[1] for k in keys2))
            ws_hm = wb.create_sheet("BatchTokenDelta")
            ws_hm.append(["token_len\\batch_size"] + [f"{x:g}" for x in batches])
            for tl in tokens:
                row = [f"{tl:g}"]
                for bs in batches:
                    row.append(_cmp_pct(amap.get((bs, tl)), bmap.get((bs, tl))))
                ws_hm.append(row)
            if len(tokens) >= 1 and len(batches) >= 1:
                start = "B2"
                end_col = get_column_letter(1 + len(batches))
                end = f"{end_col}{1 + len(tokens)}"
                ws_hm.conditional_formatting.add(
                    f"{start}:{end}",
                    ColorScaleRule(start_type="num", start_value=-50, start_color="F4D03F", mid_type="num", mid_value=0, mid_color="1B2631", end_type="num", end_value=50, end_color="2ECC71"),
                )

    from io import BytesIO

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _render_csv_tps_scatter_compare(
    *,
    a_path: Path,
    b_path: Path,
    csv_name: str,
    max_points_each: int = 2000,
) -> str:
    if not a_path.exists() and not b_path.exists():
        return '<div class="muted">Scatter unavailable: missing in both A and B.</div>'
    if not a_path.exists() or not b_path.exists():
        return '<div class="muted">Scatter unavailable: only one side has this CSV.</div>'

    a_cols, a_rows = _read_csv_dict_rows(a_path)
    b_cols, b_rows = _read_csv_dict_rows(b_path)
    if not a_cols or not b_cols:
        return '<div class="muted">Scatter unavailable: empty CSV.</div>'

    cols = [c for c in a_cols if c in set(b_cols)]
    if not cols:
        cols = list(dict.fromkeys(a_cols + b_cols))

    stem = Path(csv_name).stem
    prefix = stem[:-8] if stem.endswith("_scaling") else stem

    def _median(vals: List[float]) -> Optional[float]:
        if not vals:
            return None
        s = sorted(vals)
        n = len(s)
        m = n // 2
        if n % 2 == 1:
            return s[m]
        return (s[m - 1] + s[m]) / 2.0

    # Infer x-axis from filename first token group, e.g. batch_size_scaling.csv -> batch_size.
    x_candidates = [prefix]
    if "_" in prefix:
        x_candidates.append(prefix.split("_", 1)[0])
    x_col = next((c for c in x_candidates if c in cols), "")
    if not x_col and prefix == "cpu" and "cpu_count" in cols:
        x_col = "cpu_count"

    if not x_col:
        expected = ", ".join([c for c in x_candidates if c])
        return (
            '<div class="muted">Scatter unavailable: cannot infer x-axis column from filename. '
            f'Expected one of <span class="mono">{_e(expected)}</span>.</div>'
        )

    y_col = "tps" if "tps" in cols else ("tokens_per_sec" if "tokens_per_sec" in cols else "")

    # summary_pivot.csv: compare each *_tps scenario using grouped bars.
    if stem == "summary_pivot":
        tps_cols = [c for c in cols if c.endswith("_tps") and not c.endswith("_tps_per_cpu")]
        if not tps_cols:
            return '<div class="muted">No *_tps columns found in summary_pivot.csv.</div>'

        def _summary_vals(rows: List[Dict[str, str]]) -> Dict[str, float]:
            out: Dict[str, float] = {}
            for c in tps_cols:
                vals: List[float] = []
                for r in rows:
                    v = _try_float((r.get(c, "") or "").strip())
                    if v is not None:
                        vals.append(v)
                m = _median(vals)
                if m is not None:
                    out[c] = m
            return out

        a_vals = _summary_vals(a_rows)
        b_vals = _summary_vals(b_rows)
        cats = [c for c in tps_cols if c in a_vals or c in b_vals]
        cats = cats[:18]
        if not cats:
            return '<div class="muted">No numeric values found for summary_pivot *_tps columns.</div>'

        y_values: List[float] = []
        for c in cats:
            if c in a_vals:
                y_values.append(a_vals[c])
            if c in b_vals:
                y_values.append(b_vals[c])

        y_min = 0.0
        y_max = max(y_values) if y_values else 1.0
        if y_max <= y_min:
            y_max = y_min + 1.0

        width = 1100
        height = 420
        ml, mr, mt, mb = 72, 20, 20, 90
        pw = max(1, width - ml - mr)
        ph = max(1, height - mt - mb)

        def sy(y: float) -> float:
            return mt + (y_max - y) * ph / (y_max - y_min)

        n = len(cats)
        band = pw / max(1, n)
        bw = max(3.0, min(16.0, band * 0.32))

        ticks = 5
        grid: List[str] = []
        for i in range(ticks + 1):
            yy = mt + ph * i / ticks
            yv = y_max - (y_max - y_min) * i / ticks
            grid.append(f'<line x1="{ml}" y1="{yy:.2f}" x2="{ml + pw}" y2="{yy:.2f}" stroke="var(--border)" stroke-width="1" opacity="0.65" />')
            grid.append(f'<text x="{ml - 8}" y="{yy + 4:.2f}" fill="var(--muted)" font-size="11" text-anchor="end">{_e(f"{yv:.4g}")}</text>')

        bars: List[str] = []
        labels: List[str] = []
        for i, c in enumerate(cats):
            cx = ml + (i + 0.5) * band
            av = a_vals.get(c)
            bv = b_vals.get(c)
            if av is not None:
                ay = sy(av)
                h = mt + ph - ay
                bars.append(
                    f'<rect x="{cx - bw - 1:.2f}" y="{ay:.2f}" width="{bw:.2f}" height="{h:.2f}" fill="var(--accent)" fill-opacity="0.88">'
                    f'<title>A | {c} | {av:.6g}</title></rect>'
                )
            if bv is not None:
                by = sy(bv)
                h = mt + ph - by
                bars.append(
                    f'<rect x="{cx + 1:.2f}" y="{by:.2f}" width="{bw:.2f}" height="{h:.2f}" fill="var(--accent2)" fill-opacity="0.88">'
                    f'<title>B | {c} | {bv:.6g}</title></rect>'
                )
            short = c.replace("_tps", "")
            labels.append(
                f'<text x="{cx:.2f}" y="{height - 12}" fill="var(--muted)" font-size="10" text-anchor="end" transform="rotate(-35,{cx:.2f},{height - 12})">{_e(short)}</text>'
            )

        axis = (
            f'<line x1="{ml}" y1="{mt + ph}" x2="{ml + pw}" y2="{mt + ph}" stroke="var(--text)" stroke-width="1.2" />'
            f'<line x1="{ml}" y1="{mt}" x2="{ml}" y2="{mt + ph}" stroke="var(--text)" stroke-width="1.2" />'
        )

        svg = (
            f'<svg viewBox="0 0 {width} {height}" width="100%" height="auto" role="img" aria-label="A/B summary pivot tps bars">'
            f'{"".join(grid)}{axis}{"".join(bars)}{"".join(labels)}'
            f'<text x="{ml + pw / 2:.2f}" y="{height - 36}" fill="var(--muted)" font-size="12" text-anchor="middle">scenario (from *_tps columns)</text>'
            f'<text x="16" y="{mt + ph / 2:.2f}" fill="var(--muted)" font-size="12" text-anchor="middle" transform="rotate(-90,16,{mt + ph / 2:.2f})">tps</text>'
            '</svg>'
        )
        legend = (
            '<div class="sub"><span style="color: var(--accent)">●</span> A median &nbsp; '
            '<span style="color: var(--accent2)">●</span> B median &nbsp; '
            f'· scenarios shown: <span class="mono">{len(cats)}</span></div>'
        )
        return f'<div class="plot"><div class="plot-title">Summary Pivot TPS Compare</div>{legend}{svg}</div>'

    if not y_col:
        return '<div class="muted">Scatter unavailable: missing y-axis column <span class="mono">tps</span>.</div>'

    # *_scaling.csv: use faceted line charts so secondary dimensions (e.g. token_len)
    # are not collapsed into a single median curve.
    if stem.endswith("_scaling") and x_col:
        split_pref: Dict[str, List[str]] = {
            "batch_size": ["token_len", "cpu_count", "kv_cap"],
            "token_len": ["batch_size", "cpu_count", "kv_cap"],
            "cpu_count": ["token_len", "batch_size", "kv_cap"],
            "kv_cap": ["token_len", "batch_size", "cpu_count"],
        }
        split_candidates = split_pref.get(x_col, ["token_len", "batch_size", "cpu_count", "kv_cap"])
        split_col = ""
        for c in split_candidates:
            if c == x_col or c not in cols:
                continue
            vals = set()
            for r in (a_rows[:800] + b_rows[:800]):
                v = (r.get(c, "") or "").strip()
                if v:
                    vals.add(v)
            if len(vals) > 1:
                split_col = c
                break

        def _series(rows: List[Dict[str, str]], split_value: str = "") -> List[Tuple[float, float, int]]:
            groups: Dict[float, List[float]] = {}
            for r in rows:
                if split_col:
                    sv = (r.get(split_col, "") or "").strip()
                    if sv != split_value:
                        continue
                xv = _try_float((r.get(x_col, "") or "").strip())
                yv = _try_float((r.get(y_col, "") or "").strip())
                if xv is None or yv is None:
                    continue
                groups.setdefault(float(xv), []).append(float(yv))
            out: List[Tuple[float, float, int]] = []
            for xv in sorted(groups.keys()):
                vals = groups.get(xv, [])
                m = _median(vals)
                if m is None:
                    continue
                out.append((xv, m, len(vals)))
            return out

        split_values = [""]
        if split_col:
            vals = set()
            for r in (a_rows + b_rows):
                v = (r.get(split_col, "") or "").strip()
                if v:
                    vals.add(v)

            def _sort_key(s: str) -> Tuple[int, float, str]:
                f = _try_float(s)
                if f is None:
                    return (1, 0.0, s)
                return (0, float(f), s)

            split_values = sorted(vals, key=_sort_key)[:6]

        facet_blocks: List[str] = []
        for sv in split_values:
            a_line = _series(a_rows, sv)
            b_line = _series(b_rows, sv)
            if not a_line and not b_line:
                continue

            all_x = [x for x, _, _ in a_line] + [x for x, _, _ in b_line]
            all_y = [y for _, y, _ in a_line] + [y for _, y, _ in b_line]
            x_min, x_max = min(all_x), max(all_x)
            y_min, y_max = min(all_y), max(all_y)
            if x_min == x_max:
                x_min -= 0.5
                x_max += 0.5
            if y_min == y_max:
                pad = max(1.0, abs(y_min) * 0.1)
                y_min -= pad
                y_max += pad

            width = 1000
            height = 340
            ml, mr, mt, mb = 64, 20, 20, 44
            pw = max(1, width - ml - mr)
            ph = max(1, height - mt - mb)

            def sx(x: float) -> float:
                return ml + (x - x_min) * pw / (x_max - x_min)

            def sy(y: float) -> float:
                return mt + (y_max - y) * ph / (y_max - y_min)

            grid: List[str] = []
            ticks = 5
            for i in range(ticks + 1):
                tx = ml + pw * i / ticks
                ty = mt + ph * i / ticks
                xv = x_min + (x_max - x_min) * i / ticks
                yv = y_max - (y_max - y_min) * i / ticks
                grid.append(f'<line x1="{tx:.2f}" y1="{mt}" x2="{tx:.2f}" y2="{mt + ph}" stroke="var(--border)" stroke-width="1" opacity="0.65" />')
                grid.append(f'<line x1="{ml}" y1="{ty:.2f}" x2="{ml + pw}" y2="{ty:.2f}" stroke="var(--border)" stroke-width="1" opacity="0.65" />')
                grid.append(f'<text x="{tx:.2f}" y="{height - 10}" fill="var(--muted)" font-size="11" text-anchor="middle">{_e(f"{xv:.4g}")}</text>')
                grid.append(f'<text x="{ml - 8}" y="{ty + 4:.2f}" fill="var(--muted)" font-size="11" text-anchor="end">{_e(f"{yv:.4g}")}</text>')

            def _poly(series: List[Tuple[float, float, int]], color: str, tag: str) -> str:
                if not series:
                    return ""
                pts = " ".join(f"{sx(x):.2f},{sy(y):.2f}" for x, y, _ in series)
                circles = []
                for x, y, n in series:
                    circles.append(
                        f'<circle cx="{sx(x):.2f}" cy="{sy(y):.2f}" r="3.4" fill="{color}" fill-opacity="0.9">'
                        f'<title>{tag} | {x_col}={x:.6g}, median {y_col}={y:.6g}, n={n}</title></circle>'
                    )
                return f'<polyline fill="none" stroke="{color}" stroke-width="2.2" points="{pts}" />' + "".join(circles)

            axis = (
                f'<line x1="{ml}" y1="{mt + ph}" x2="{ml + pw}" y2="{mt + ph}" stroke="var(--text)" stroke-width="1.2" />'
                f'<line x1="{ml}" y1="{mt}" x2="{ml}" y2="{mt + ph}" stroke="var(--text)" stroke-width="1.2" />'
            )
            svg = (
                f'<svg viewBox="0 0 {width} {height}" width="100%" height="auto" role="img" aria-label="A/B scaling line compare">'
                f'{"".join(grid)}{axis}{_poly(a_line, "var(--accent)", "A")}{_poly(b_line, "var(--accent2)", "B")}'
                f'<text x="{ml + pw / 2:.2f}" y="{height - 24}" fill="var(--muted)" font-size="12" text-anchor="middle">{_e(x_col)}</text>'
                f'<text x="16" y="{mt + ph / 2:.2f}" fill="var(--muted)" font-size="12" text-anchor="middle" transform="rotate(-90,16,{mt + ph / 2:.2f})">{_e(y_col)} (median)</text>'
                '</svg>'
            )
            facet_title = f'{split_col}={sv}' if split_col else 'all rows'
            facet_blocks.append(
                '<div class="plot">'
                f'<div class="plot-title">Scaling Line Compare · {_e(facet_title)}</div>'
                f'{svg}'
                '</div>'
            )

        if facet_blocks:
            legend = (
                f'<div class="sub">x-axis: <span class="mono">{_e(x_col)}</span> · y-axis: <span class="mono">{_e(y_col)}</span> (median by x)</div>'
                + (f'<div class="sub">facets by <span class="mono">{_e(split_col)}</span> (up to 6)</div>' if split_col else '')
                + '<div class="sub"><span style="color: var(--accent)">●</span> A &nbsp; <span style="color: var(--accent2)">●</span> B</div>'
            )
            return f'<div>{legend}<div class="plots">{"".join(facet_blocks)}</div></div>'

    def _collect_points(rows: List[Dict[str, str]], tag: str) -> List[Tuple[float, float, str]]:
        pts: List[Tuple[float, float, str]] = []
        for r in rows:
            xv = _try_float((r.get(x_col, "") or "").strip())
            yv = _try_float((r.get(y_col, "") or "").strip())
            if xv is None or yv is None:
                continue
            detail = []
            for k in cols:
                if k in {x_col, y_col}:
                    continue
                vv = (r.get(k, "") or "").strip()
                if vv:
                    detail.append(f"{k}={vv}")
            hover = f"{tag} | {x_col}={xv:.6g}, {y_col}={yv:.6g}"
            if detail:
                hover += " | " + ", ".join(detail[:6])
            pts.append((xv, yv, hover))
        if len(pts) <= max_points_each:
            return pts
        # Uniform downsample to keep page responsive.
        step = max(1, len(pts) // max_points_each)
        return pts[::step][:max_points_each]

    a_pts = _collect_points(a_rows, "A")
    b_pts = _collect_points(b_rows, "B")
    if not a_pts and not b_pts:
        return '<div class="muted">Scatter unavailable: no valid numeric points for inferred axes.</div>'

    all_pts = a_pts + b_pts
    xs = [p[0] for p in all_pts]
    ys = [p[1] for p in all_pts]
    x_min, x_max = min(xs), max(xs)
    y_min, y_max = min(ys), max(ys)

    if x_min == x_max:
        x_min -= 0.5
        x_max += 0.5
    if y_min == y_max:
        pad = max(1.0, abs(y_min) * 0.1)
        y_min -= pad
        y_max += pad

    width = 980
    height = 360
    ml, mr, mt, mb = 64, 20, 20, 44
    pw = max(1, width - ml - mr)
    ph = max(1, height - mt - mb)

    def sx(x: float) -> float:
        return ml + (x - x_min) * pw / (x_max - x_min)

    def sy(y: float) -> float:
        return mt + (y_max - y) * ph / (y_max - y_min)

    grid: List[str] = []
    ticks = 5
    for i in range(ticks + 1):
        tx = ml + pw * i / ticks
        ty = mt + ph * i / ticks
        xv = x_min + (x_max - x_min) * i / ticks
        yv = y_max - (y_max - y_min) * i / ticks
        grid.append(f'<line x1="{tx:.2f}" y1="{mt}" x2="{tx:.2f}" y2="{mt + ph}" stroke="var(--border)" stroke-width="1" opacity="0.65" />')
        grid.append(f'<line x1="{ml}" y1="{ty:.2f}" x2="{ml + pw}" y2="{ty:.2f}" stroke="var(--border)" stroke-width="1" opacity="0.65" />')
        grid.append(f'<text x="{tx:.2f}" y="{height - 10}" fill="var(--muted)" font-size="11" text-anchor="middle">{_e(f"{xv:.4g}")}</text>')
        grid.append(f'<text x="{ml - 8}" y="{ty + 4:.2f}" fill="var(--muted)" font-size="11" text-anchor="end">{_e(f"{yv:.4g}")}</text>')

    def dots(points: List[Tuple[float, float, str]], color: str) -> str:
        out: List[str] = []
        for x, y, hover in points:
            out.append(
                f'<circle cx="{sx(x):.2f}" cy="{sy(y):.2f}" r="3.3" fill="{color}" fill-opacity="0.88">'
                f'<title>{_e(hover)}</title></circle>'
            )
        return "".join(out)

    axis = (
        f'<line x1="{ml}" y1="{mt + ph}" x2="{ml + pw}" y2="{mt + ph}" stroke="var(--text)" stroke-width="1.2" />'
        f'<line x1="{ml}" y1="{mt}" x2="{ml}" y2="{mt + ph}" stroke="var(--text)" stroke-width="1.2" />'
    )

    svg = (
        f'<svg viewBox="0 0 {width} {height}" width="100%" height="auto" role="img" aria-label="A/B scatter for {_e(csv_name)}">'
        f'{"".join(grid)}'
        f'{axis}'
        f'{dots(a_pts, "var(--accent)")}'
        f'{dots(b_pts, "var(--accent2)")}'
        f'<text x="{ml + pw / 2:.2f}" y="{height - 24}" fill="var(--muted)" font-size="12" text-anchor="middle">{_e(x_col)}</text>'
        f'<text x="16" y="{mt + ph / 2:.2f}" fill="var(--muted)" font-size="12" text-anchor="middle" transform="rotate(-90,16,{mt + ph / 2:.2f})">{_e(y_col)}</text>'
        '</svg>'
    )

    legend = (
        f'<div class="sub">x-axis: <span class="mono">{_e(x_col)}</span> · y-axis: <span class="mono">{_e(y_col)}</span> '
        f'· points A/B: <span class="mono">{len(a_pts)}</span>/<span class="mono">{len(b_pts)}</span></div>'
        '<div class="sub"><span style="color: var(--accent)">●</span> A &nbsp; '
        '<span style="color: var(--accent2)">●</span> B &nbsp; (hover points for details)</div>'
    )

    return f'<div class="plot"><div class="plot-title">TPS Scatter Compare</div>{legend}{svg}</div>'


def _render_batch_token_tps_delta_heatmap(*, a_path: Path, b_path: Path) -> str:
    if not a_path.exists() or not b_path.exists():
        return '<div class="muted">Unavailable: this chart needs both A and B CSV files.</div>'

    a_cols, a_rows = _read_csv_dict_rows(a_path)
    b_cols, b_rows = _read_csv_dict_rows(b_path)
    cols = [c for c in a_cols if c in set(b_cols)]
    if not cols:
        cols = list(dict.fromkeys(a_cols + b_cols))

    need = {"batch_size", "token_len", "tps"}
    if not need.issubset(set(cols)):
        return '<div class="muted">Unavailable: requires columns <span class="mono">batch_size</span>, <span class="mono">token_len</span>, <span class="mono">tps</span>.</div>'

    def _median(vals: List[float]) -> Optional[float]:
        if not vals:
            return None
        s = sorted(vals)
        n = len(s)
        m = n // 2
        if n % 2 == 1:
            return s[m]
        return (s[m - 1] + s[m]) / 2.0

    def _agg(rows: List[Dict[str, str]]) -> Dict[Tuple[float, float], float]:
        g: Dict[Tuple[float, float], List[float]] = {}
        for r in rows:
            bs = _try_float((r.get("batch_size", "") or "").strip())
            tl = _try_float((r.get("token_len", "") or "").strip())
            tps = _try_float((r.get("tps", "") or "").strip())
            if bs is None or tl is None or tps is None:
                continue
            g.setdefault((float(bs), float(tl)), []).append(float(tps))
        out: Dict[Tuple[float, float], float] = {}
        for k, vals in g.items():
            m = _median(vals)
            if m is not None:
                out[k] = m
        return out

    amap = _agg(a_rows)
    bmap = _agg(b_rows)
    keys = sorted(set(amap.keys()) | set(bmap.keys()), key=lambda k: (k[1], k[0]))
    if not keys:
        return '<div class="muted">Unavailable: no valid numeric (batch_size, token_len, tps) rows.</div>'

    batch_vals = sorted(set(k[0] for k in keys))
    token_vals = sorted(set(k[1] for k in keys))
    if not batch_vals or not token_vals:
        return '<div class="muted">Unavailable: empty batch/token grid.</div>'

    deltas: Dict[Tuple[float, float], Optional[float]] = {}
    max_abs = 0.0
    for tl in token_vals:
        for bs in batch_vals:
            av = amap.get((bs, tl))
            bv = bmap.get((bs, tl))
            dv = _cmp_pct(av, bv)
            deltas[(bs, tl)] = dv
            if dv is not None and dv not in {float("inf"), float("-inf")}:
                max_abs = max(max_abs, abs(float(dv)))
    if max_abs <= 0:
        max_abs = 1.0

    cell_w = 92
    cell_h = 42
    ml, mt = 92, 36
    width = ml + len(batch_vals) * cell_w + 20
    height = mt + len(token_vals) * cell_h + 54

    cells: List[str] = []
    # Column labels (batch_size)
    for i, bs in enumerate(batch_vals):
        x = ml + i * cell_w
        cells.append(f'<text x="{x + cell_w / 2:.2f}" y="20" fill="var(--muted)" font-size="11" text-anchor="middle">{_e(f"{bs:g}")}</text>')
    # Row labels (token_len)
    for j, tl in enumerate(token_vals):
        y = mt + j * cell_h
        cells.append(f'<text x="{ml - 10}" y="{y + cell_h / 2 + 4:.2f}" fill="var(--muted)" font-size="11" text-anchor="end">{_e(f"{tl:g}")}</text>')

    for j, tl in enumerate(token_vals):
        for i, bs in enumerate(batch_vals):
            x = ml + i * cell_w
            y = mt + j * cell_h
            dv = deltas.get((bs, tl))
            av = amap.get((bs, tl))
            bv = bmap.get((bs, tl))

            if dv is None:
                fill = "transparent"
                op = 1.0
                txt = "-"
            elif dv == float("inf"):
                fill = "var(--accent2)"
                op = 0.85
                txt = "inf"
            elif dv == float("-inf"):
                fill = "var(--warn)"
                op = 0.85
                txt = "-inf"
            else:
                d = float(dv)
                fill = "var(--accent2)" if d >= 0 else "var(--warn)"
                op = min(0.88, 0.18 + 0.70 * abs(d) / max_abs)
                txt = f"{d:+.1f}%"

            title = f'batch_size={bs:g}, token_len={tl:g} | A={"-" if av is None else f"{av:.6g}"} B={"-" if bv is None else f"{bv:.6g}"} delta={txt}'
            cells.append(
                f'<rect x="{x + 1:.2f}" y="{y + 1:.2f}" width="{cell_w - 2:.2f}" height="{cell_h - 2:.2f}" '
                f'stroke="var(--border)" stroke-width="1" fill="{fill}" fill-opacity="{op:.3f}">'
                f'<title>{_e(title)}</title></rect>'
            )
            cells.append(
                f'<text x="{x + cell_w / 2:.2f}" y="{y + cell_h / 2 + 4:.2f}" fill="var(--text)" font-size="11" text-anchor="middle">{_e(txt)}</text>'
            )

    svg = (
        f'<svg viewBox="0 0 {width} {height}" width="100%" height="auto" role="img" aria-label="TPS delta heatmap by batch_size and token_len">'
        f'<text x="{ml + (len(batch_vals) * cell_w) / 2:.2f}" y="12" fill="var(--muted)" font-size="12" text-anchor="middle">batch_size</text>'
        f'<text x="16" y="{mt + (len(token_vals) * cell_h) / 2:.2f}" fill="var(--muted)" font-size="12" text-anchor="middle" transform="rotate(-90,16,{mt + (len(token_vals) * cell_h) / 2:.2f})">token_len</text>'
        f'{"".join(cells)}'
        '</svg>'
    )

    legend = (
        '<div class="sub">Cell value = <span class="mono">(A/B - 1) * 100%</span> on TPS (median if duplicated rows).</div>'
        '<div class="sub"><span style="color: var(--accent2)">■</span> A better than B &nbsp; '
        '<span style="color: var(--warn)">■</span> A worse than B &nbsp; (color depth = |delta|)</div>'
    )
    return f'<div class="plot"><div class="plot-title">Batch×Token TPS Delta Heatmap</div>{legend}{svg}</div>'


def _compare_csv_tables(
    *,
    a_path: Path,
    b_path: Path,
    max_out_rows: int = 200,
    cell_view: str = "pct",
) -> str:
    if not a_path.exists() and not b_path.exists():
        return '<div class="muted">Missing in both A and B.</div>'
    if not a_path.exists():
        return '<div class="warn">Missing in A.</div>'
    if not b_path.exists():
        return '<div class="warn">Missing in B.</div>'

    a_cols, a_rows = _read_csv_dict_rows(a_path)
    b_cols, b_rows = _read_csv_dict_rows(b_path)
    cols = [c for c in a_cols if c in set(b_cols)]
    if not cols:
        cols = list(dict.fromkeys(a_cols + b_cols))

    dim_force = {
        "variant",
        "resource_cpu",
        "resource_cpu_count",
        "resource_mem_gb",
        "cpu_count",
        "batch_size",
        "token_len",
        "kv_cap",
        "sglang_max_total_tokens",
    }

    def is_numeric_col(name: str) -> bool:
        vs: List[str] = []
        for r in a_rows[:200]:
            v = r.get(name, "")
            if v:
                vs.append(v)
        for r in b_rows[:200]:
            v = r.get(name, "")
            if v:
                vs.append(v)
        if not vs:
            return False
        ok = 0
        for v in vs:
            if _try_float(v) is not None:
                ok += 1
        return (ok / max(1, len(vs))) >= 0.85

    numeric_cols = {c for c in cols if is_numeric_col(c)}

    # Metrics: numeric columns that are not part of the known dimensions.
    # (We intentionally include e.g. *_per_cpu columns; the old regex-based filter
    # missed these and caused them to be treated as keys.)
    metrics = [c for c in cols if (c in numeric_cols and c not in dim_force)]

    # Key columns should be stable across runs; drop volatile columns that contain
    # file paths / logs / artifacts (these differ even when the measurement point is the same).
    candidate_key_cols = [c for c in cols if c not in set(metrics)]

    # Do not use CPU identity fields as join keys for compare-csv.
    # These can differ between A/B runs even when the rest of the point is comparable,
    # which causes row mismatch in the comparison output.
    ab_only_cols = [c for c in ["resource_cpu", "resource_cpu_count", "cpu_count"] if c in cols]
    candidate_key_cols = [c for c in candidate_key_cols if c not in set(ab_only_cols)]

    volatile_name_re = re.compile(
        r"(path|file|dir|folder|artifact|output|stdout|stderr|log|xlsx|html|json|csv)$",
        re.IGNORECASE,
    )
    volatile_exact = {
        "emon_summary_xlsx",
        "auto_test_stdout_log",
        "auto_test_stderr_log",
        "run_dir",
        "run_path",
        "summary_xlsx",
    }

    def looks_like_path(v: str) -> bool:
        s = (v or "").strip()
        if not s:
            return False
        # absolute/relative paths, urls, or obvious file extensions
        if "/" in s or "\\" in s or s.startswith("http://") or s.startswith("https://"):
            return True
        lower = s.lower()
        return any(lower.endswith(ext) for ext in (".xlsx", ".csv", ".json", ".html", ".log", ".txt", ".png"))

    def is_volatile_key_col(name: str) -> bool:
        n = (name or "").strip()
        if not n:
            return True
        if n in volatile_exact:
            return True
        if volatile_name_re.search(n):
            return True
        # If most values look like paths, treat as volatile.
        sample: List[str] = []
        for r in a_rows[:200]:
            v = (r.get(n) or "").strip()
            if v:
                sample.append(v)
        for r in b_rows[:200]:
            v = (r.get(n) or "").strip()
            if v:
                sample.append(v)
        if not sample:
            return False
        pathish = sum(1 for v in sample if looks_like_path(v))
        return (pathish / max(1, len(sample))) >= 0.60

    key_cols = [c for c in candidate_key_cols if not is_volatile_key_col(c)]
    if not key_cols:
        # Fallback: keep original candidate keys if everything was filtered.
        key_cols = candidate_key_cols

    def row_key(r: Dict[str, str], *, i: int) -> Tuple[str, ...]:
        if not key_cols:
            return (str(i),)
        return tuple((r.get(c, "") or "").strip() for c in key_cols)

    a_idx: Dict[Tuple[str, ...], Dict[str, str]] = {}
    for i, r in enumerate(a_rows):
        a_idx[row_key(r, i=i)] = r
    b_idx: Dict[Tuple[str, ...], Dict[str, str]] = {}
    for i, r in enumerate(b_rows):
        b_idx[row_key(r, i=i)] = r

    keys = sorted(set(a_idx.keys()) | set(b_idx.keys()))
    if not metrics:
        return '<div class="muted">No numeric metric columns detected; nothing to compute (A/B-1).</div>'

    view = (cell_view or "").strip().lower()
    if view not in {"pct", "full", "abdelta"}:
        view = "pct"

    def _fmt_raw_cell(v: str) -> str:
        s = (v or "").strip()
        if not s:
            return "-"
        fv = _try_float(s)
        if fv is None:
            return s if len(s) <= 64 else (s[:61] + "...")
        try:
            return f"{float(fv):.6g}"
        except Exception:
            return s

    if view == "pct":
        metric_heads = "".join(f"<th class=\"mono\">{_e(c)} %</th>" for c in metrics)
    elif view == "abdelta":
        metric_heads = "".join(
            f"<th class=\"mono\">A_{_e(c)}</th>"
            f"<th class=\"mono\">B_{_e(c)}</th>"
            f"<th class=\"mono\">delta_{_e(c)}%</th>"
            for c in metrics
        )
    else:
        metric_heads = "".join(f"<th class=\"mono\">{_e(c)}</th>" for c in metrics)
    ab_heads = "".join(
        f"<th>{_e(c)}_A</th><th>{_e(c)}_B</th>"
        for c in ab_only_cols
    )
    head = "<tr>" + "".join(f"<th>{_e(c)}</th>" for c in key_cols) + ab_heads + metric_heads + "</tr>"

    body_rows: List[str] = []

    def _fmt_delta_abdelta(a: Optional[float], b: Optional[float]) -> str:
        """Format delta for the abdelta view.

        - If A >= B and B > 0: show ratio as "{A/B}×" ("提升了多少倍").
        - If A < B: show percent drop as a negative percent (same as (A/B-1)*100%).
        - Otherwise fall back to percent formatting.
        """

        if a is None or b is None:
            return "-"
        if b == 0.0:
            if a == 0.0:
                return "1×"
            return "inf×"

        # Only use the "times" representation for positive baselines.
        if b > 0.0 and a >= b and a >= 0.0:
            try:
                return f"{(a / b):.6g}×"
            except Exception:
                pass
        return _fmt_pct(_cmp_pct(a, b))

    for k in keys[:max_out_rows]:
        ar = a_idx.get(k) or {}
        br = b_idx.get(k) or {}
        tds = []
        for j, c in enumerate(key_cols):
            v = k[j] if j < len(k) else ""
            tds.append(f"<td>{_e(v)}</td>")
        for c in ab_only_cols:
            tds.append(f"<td>{_e((ar.get(c, '') or '').strip())}</td>")
            tds.append(f"<td>{_e((br.get(c, '') or '').strip())}</td>")
        for c in metrics:
            a_raw = (ar.get(c, "") or "").strip()
            b_raw = (br.get(c, "") or "").strip()
            av = _try_float(a_raw)
            bv = _try_float(b_raw)

            pct_ab = _cmp_pct(av, bv)
            pct_ba = _cmp_pct(bv, av)

            if view == "abdelta":
                a_disp = _fmt_raw_cell(a_raw)
                b_disp = _fmt_raw_cell(b_raw)

                ratio_ab = None
                try:
                    if av is not None and bv is not None and bv != 0.0:
                        ratio_ab = av / bv
                except Exception:
                    ratio_ab = None

                ratio_str = "-"
                if ratio_ab == float("inf"):
                    ratio_str = "inf"
                elif ratio_ab == float("-inf"):
                    ratio_str = "-inf"
                elif ratio_ab is not None:
                    try:
                        ratio_str = f"{float(ratio_ab):.6g}"
                    except Exception:
                        ratio_str = str(ratio_ab)

                title = (
                    f"A={a_raw}  B={b_raw}  "
                    f"ratio(A/B)={ratio_str}  "
                    f"delta(A/B-1)={_fmt_pct(pct_ab)}  delta(B/A-1)={_fmt_pct(pct_ba)}"
                )
                tds.append(f"<td class=\"mono\" title=\"{_e(title)}\">{_e(a_disp)}</td>")
                tds.append(f"<td class=\"mono\" title=\"{_e(title)}\">{_e(b_disp)}</td>")
                tds.append(f"<td class=\"mono\" title=\"{_e(title)}\">{_e(_fmt_delta_abdelta(av, bv))}</td>")
            elif view == "full":
                a_disp = _fmt_raw_cell(a_raw)
                b_disp = _fmt_raw_cell(b_raw)
                title = (
                    f"A={a_raw}  B={b_raw}  "
                    f"(A/B-1)={_fmt_pct(pct_ab)}  (B/A-1)={_fmt_pct(pct_ba)}"
                )
                cell = (
                    f'<div class="mono">A={_e(a_disp)} '
                    f'<span class="muted">{_e(_fmt_pct(pct_ab))}</span></div>'
                    f'<div class="mono">B={_e(b_disp)} '
                    f'<span class="muted">{_e(_fmt_pct(pct_ba))}</span></div>'
                )
                tds.append(f"<td title=\"{_e(title)}\">{cell}</td>")
            else:
                title = f"A={a_raw}  B={b_raw}"
                tds.append(f"<td class=\"mono\" title=\"{_e(title)}\">{_e(_fmt_pct(pct_ab))}</td>")
        body_rows.append("<tr>" + "".join(tds) + "</tr>")

    more = ""
    if len(keys) > max_out_rows:
        more = f'<div class="muted">Showing first {max_out_rows} of {len(keys)} rows.</div>'

    return (
        '<div class="table-scroll csv-compare-scroll">'
        f'<table class="table small"><thead>{head}</thead><tbody>{"".join(body_rows)}</tbody></table>'
        "</div>"
        + more
    )


def _render_compare(scale_test_root: Path, runs: List[RunInfo], q: Dict[str, List[str]]) -> str:
    a_raw = (q.get("a") or [""])[0].strip()
    b_raw = (q.get("b") or [""])[0].strip()

    # Allow selecting from homepage checkboxes: /compare?sel=...&sel=...
    sels = [str(s).strip() for s in (q.get("sel") or []) if str(s).strip()]
    sel_warn = ""
    if sels and (not a_raw or not b_raw):
        if not a_raw and len(sels) >= 1:
            a_raw = sels[0]
        if not b_raw and len(sels) >= 2:
            b_raw = sels[1]
        if len(sels) > 2:
            sel_warn = '<div class="warn">More than 2 runs selected; using the first two.</div>'

    a_ref = _parse_run_key(a_raw)
    b_ref = _parse_run_key(b_raw)

    options: List[Tuple[str, str]] = []
    for r in runs:
        meta = _extract_run_meta(r)
        model = meta.get("model") or "-"
        cpu = meta.get("cpu") or "-"
        key = f"{r.ref.task}/{r.ref.suite}/{r.ref.run_id}"
        label = f"{key}  |  {model}  |  {cpu}  |  {_fmt_dt(r.mtime)}"
        options.append((key, label))

    def sel(name: str, chosen: str) -> str:
        opts = ['<option value="">(select)</option>']
        for key, label in options:
            s = " selected" if chosen and key == chosen else ""
            opts.append(f'<option value="{_e(key)}"{s}>{_e(label)}</option>')
        return f'<select name="{_e(name)}">' + "".join(opts) + "</select>"

    form = (
        '<section class="card">'
        '<h1>Compare Runs</h1>'
        '<div class="sub">Pick A/B → see server info → click a CSV to compare it.</div>'
        '<div class="sub">Per-cell percent: <span class="mono">(A/B - 1) * 100%</span>. Hover a cell to see raw A/B.</div>'
        '<form method="get" action="/compare" class="form-inline">'
        f'<label>A {sel("a", a_raw)}</label>'
        f'<label>B {sel("b", b_raw)}</label>'
        '<button type="submit">Go</button>'
        '<a class="btn" href="/compare">Reset</a>'
        '</form>'
        f'{sel_warn}'
        '</section>'
    )

    if not (a_ref and b_ref):
        return _html_page("Compare Runs", form)

    run_a = next((r for r in runs if r.ref == a_ref), None)
    run_b = next((r for r in runs if r.ref == b_ref), None)
    if run_a is None or run_b is None:
        warn = '<section class="card"><div class="warn">Invalid A/B selection (run not found or missing analysis/).</div></section>'
        return _html_page("Compare Runs", form + warn)

    _, a_csvs = _list_analysis_files(run_a.analysis_dir)
    _, b_csvs = _list_analysis_files(run_b.analysis_dir)
    a_map = {p.name: p for p in a_csvs}
    b_map = {p.name: p for p in b_csvs}
    names = sorted(set(a_map.keys()) | set(b_map.keys()))

    a_href = f"/run/{_e(run_a.ref.task)}/{_e(run_a.ref.suite)}/{_e(run_a.ref.run_id)}"
    b_href = f"/run/{_e(run_b.ref.task)}/{_e(run_b.ref.suite)}/{_e(run_b.ref.run_id)}"
    a_si = f"/server-info/{_e(run_a.ref.task)}/{_e(run_a.ref.suite)}/{_e(run_a.ref.run_id)}"
    b_si = f"/server-info/{_e(run_b.ref.task)}/{_e(run_b.ref.suite)}/{_e(run_b.ref.run_id)}"

    header = (
        '<section class="card">'
        '<h2>Selection</h2>'
        f'<div class="sub">A: <a class="mono" href="{a_href}">{_e(a_raw)}</a> · <a href="{a_si}">server info</a></div>'
        f'<div class="sub">B: <a class="mono" href="{b_href}">{_e(b_raw)}</a> · <a href="{b_si}">server info</a></div>'
        '</section>'
    )

    view = (q.get("view") or [""])[0].strip().lower()
    if view == "all":
        blocks: List[str] = []
        for name in names:
            a_p = a_map.get(name, Path("/dev/null"))
            b_p = b_map.get(name, Path("/dev/null"))
            table = _compare_csv_tables(a_path=a_p, b_path=b_p, max_out_rows=200, cell_view="pct")
            blocks.append(
                '<section class="card">'
                f'<details open><summary><span class="mono">{_e(name)}</span></summary>'
                f'{table}'
                '</details>'
                '</section>'
            )
        return _html_page("Compare Runs", form + header + "".join(blocks))

    server_infos = (
        '<section class="card">'
        '<h2>Server Info</h2>'
        '<div class="plots">'
        f'<div class="plot"><div class="plot-title">A</div>{_render_server_info_brief(run_a)}</div>'
        f'<div class="plot"><div class="plot-title">B</div>{_render_server_info_brief(run_b)}</div>'
        '</div>'
        '</section>'
    )

    csv_links: List[str] = []
    for name in names:
        have_a = name in a_map
        have_b = name in b_map
        missing = ""
        if not have_a:
            missing += ' <span class="warn">(missing A)</span>'
        if not have_b:
            missing += ' <span class="warn">(missing B)</span>'
        href = f"/compare-csv?a={quote(a_raw)}&b={quote(b_raw)}&csv={quote(name)}"
        csv_links.append(f'<div><a class="mono" href="{href}">{_e(name)}</a>{missing}</div>')

    csv_list = (
        '<section class="card">'
        '<h2>CSV Files</h2>'
        '<div class="sub">Click a CSV to open the CSV comparison page.</div>'
        '<div style="margin-top: 10px;">'
        + "".join(csv_links)
        + '</div>'
        '</section>'
    )

    all_href = f"/compare?a={quote(a_raw)}&b={quote(b_raw)}&view=all"
    hint = (
        '<section class="card">'
        f'<div class="sub">Need the old all-in-one view? <a href="{all_href}">View all comparisons</a></div>'
        '</section>'
    )

    return _html_page("Compare Runs", form + header + server_infos + csv_list + hint)


def _render_compare_csv(scale_test_root: Path, runs: List[RunInfo], q: Dict[str, List[str]]) -> str:
    a_raw = (q.get("a") or [""])[0].strip()
    b_raw = (q.get("b") or [""])[0].strip()
    csv_name = (q.get("csv") or [""])[0].strip()

    a_ref = _parse_run_key(a_raw)
    b_ref = _parse_run_key(b_raw)
    if not (a_ref and b_ref and csv_name):
        warn = '<section class="card"><div class="warn">Missing parameters. Need: a, b, csv.</div></section>'
        back = '<section class="card"><a class="btn" href="/compare">Back</a></section>'
        return _html_page("Compare CSV", warn + back)

    run_a = next((r for r in runs if r.ref == a_ref), None)
    run_b = next((r for r in runs if r.ref == b_ref), None)
    if run_a is None or run_b is None:
        warn = '<section class="card"><div class="warn">Invalid A/B selection (run not found or missing analysis/).</div></section>'
        back = f'<section class="card"><a class="btn" href="/compare?a={quote(a_raw)}&b={quote(b_raw)}">Back</a></section>'
        return _html_page("Compare CSV", warn + back)

    _, a_csvs = _list_analysis_files(run_a.analysis_dir)
    _, b_csvs = _list_analysis_files(run_b.analysis_dir)
    a_map = {p.name: p for p in a_csvs}
    b_map = {p.name: p for p in b_csvs}

    a_p = a_map.get(csv_name, Path("/dev/null"))
    b_p = b_map.get(csv_name, Path("/dev/null"))
    scatter = _render_csv_tps_scatter_compare(a_path=a_p, b_path=b_p, csv_name=csv_name)
    bt_heatmap = _render_batch_token_tps_delta_heatmap(a_path=a_p, b_path=b_p)
    table = _compare_csv_tables(a_path=a_p, b_path=b_p, max_out_rows=400, cell_view="abdelta")

    back_href = f"/compare?a={quote(a_raw)}&b={quote(b_raw)}"
    a_href = f"/run/{_e(run_a.ref.task)}/{_e(run_a.ref.suite)}/{_e(run_a.ref.run_id)}"
    b_href = f"/run/{_e(run_b.ref.task)}/{_e(run_b.ref.suite)}/{_e(run_b.ref.run_id)}"

    swap_href = f"/compare-csv?a={quote(b_raw)}&b={quote(a_raw)}&csv={quote(csv_name)}"
    export_href = f"/compare-csv-export.xlsx?a={quote(a_raw)}&b={quote(b_raw)}&csv={quote(csv_name)}"

    a_meta = _extract_run_meta(run_a)
    b_meta = _extract_run_meta(run_b)
    a_cpu = a_meta.get("cpu") or "-"
    b_cpu = b_meta.get("cpu") or "-"
    a_cores = a_meta.get("cpu_cores") or ""
    b_cores = b_meta.get("cpu_cores") or ""
    a_cores_txt = f" ({a_cores} cores)" if a_cores else ""
    b_cores_txt = f" ({b_cores} cores)" if b_cores else ""

    cpu_info = (
        '<section class="card">'
        '<h2>CPU Info</h2>'
        '<div class="plots">'
        f'<div class="plot"><div class="plot-title">A · {_e(a_cpu)}{_e(a_cores_txt)}</div>{_render_cpu_info_brief(run_a)}</div>'
        f'<div class="plot"><div class="plot-title">B · {_e(b_cpu)}{_e(b_cores_txt)}</div>{_render_cpu_info_brief(run_b)}</div>'
        '</div>'
        '</section>'
    )

    body = (
        '<section class="card">'
        '<div class="breadcrumbs">'
        f'<a href="/">Home</a><span class="sep">/</span>'
        f'<a href="{back_href}">Compare</a><span class="sep">/</span>'
        f'<span class="mono">{_e(csv_name)}</span>'
        '</div>'
        f'<h1><span class="mono">{_e(csv_name)}</span></h1>'
        f'<div style="margin-top: 10px; display: flex; gap: 10px; align-items: center; flex-wrap: wrap;">'
        f'<a class="btn" href="{back_href}">Back to CSV list</a>'
        f'<a class="btn" href="{swap_href}">Swap A/B</a>'
        f'<a class="btn" href="{export_href}">Download XLSX</a>'
        '</div>'
        f'<div class="sub">A cpu: <span class="mono">{_e(a_cpu)}{_e(a_cores_txt)}</span></div>'
        f'<div class="sub">B cpu: <span class="mono">{_e(b_cpu)}{_e(b_cores_txt)}</span></div>'
        f'<div class="sub">A: <a class="mono" href="{a_href}">{_e(a_raw)}</a></div>'
        f'<div class="sub">B: <a class="mono" href="{b_href}">{_e(b_raw)}</a></div>'
        '<div class="sub">Metrics are expanded into <span class="mono">A_*</span>, <span class="mono">B_*</span>, and <span class="mono">delta_*%</span>. For <span class="mono">delta_*%</span>: when <span class="mono">A ≥ B</span> (and <span class="mono">B &gt; 0</span>) it shows the ratio <span class="mono">A/B</span> as <span class="mono">×</span> ("how many times"); when <span class="mono">A &lt; B</span> it shows the percent drop <span class="mono">(A/B - 1) * 100%</span>.</div>'
        '</section>'
        f'{cpu_info}'
        '<section class="card">'
        '<h2>Auto Compare Plot</h2>'
        f'{scatter}'
        '</section>'
        '<section class="card">'
        '<h2>Batch/Token Delta</h2>'
        f'{bt_heatmap}'
        '</section>'
        '<section class="card">'
        f'{table}'
        '</section>'
    )
    return _html_page("Compare CSV", body)


def _list_server_info_hosts(run: RunInfo) -> List[Dict[str, str]]:
    """Return list of host entries with tags and label.

    Each entry: {"tag": "...", "label": "...", "base_rel": "..."}
    where base_rel is relative to run.run_dir.
    """

    out: List[Dict[str, str]] = []

    # Local single-host run: <run>/server_info/
    local_info = (run.run_dir / "server_info").resolve()
    if _is_within(local_info, run.run_dir) and local_info.exists() and local_info.is_dir():
        out.append({"tag": "local", "label": "local", "base_rel": "server_info"})

    # Multi-host: <run>/hosts/<host_tag>/server_info/
    hosts_dir = (run.run_dir / "hosts").resolve()
    if _is_within(hosts_dir, run.run_dir) and hosts_dir.exists() and hosts_dir.is_dir():
        try:
            for child in sorted(hosts_dir.iterdir(), key=lambda p: p.name):
                if not child.is_dir():
                    continue
                info_dir = (child / "server_info").resolve()
                if not _is_within(info_dir, run.run_dir) or not info_dir.exists() or not info_dir.is_dir():
                    continue
                label = child.name
                host_txt = (child / "server_host.txt").resolve()
                if _is_within(host_txt, run.run_dir) and host_txt.exists() and host_txt.is_file():
                    try:
                        label2 = _read_text(host_txt, max_bytes=4096).strip()
                        if label2:
                            label = label2
                    except Exception:
                        pass
                out.append({"tag": child.name, "label": label, "base_rel": f"hosts/{child.name}/server_info"})
        except Exception:
            pass

    # De-dup by tag
    seen: set[str] = set()
    uniq: List[Dict[str, str]] = []
    for it in out:
        t = str(it.get("tag") or "").strip()
        if not t or t in seen:
            continue
        seen.add(t)
        uniq.append(it)
    return uniq


def _render_server_info(scale_test_root: Path, run: RunInfo, q: Dict[str, List[str]], host_tag: str = "") -> str:
    hosts = _list_server_info_hosts(run)
    chosen = str(host_tag or "").strip()
    if not chosen:
        chosen = (q.get("host") or [""])[0].strip()

    # Default: first host entry
    if not chosen and hosts:
        chosen = str(hosts[0].get("tag") or "")

    chosen_entry: Optional[Dict[str, str]] = None
    for it in hosts:
        if str(it.get("tag") or "") == chosen:
            chosen_entry = it
            break

    # Breadcrumbs
    title = f"{run.ref.task} / {run.ref.suite} / {run.ref.run_id} / server info"
    run_href = f"/run/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}"

    # Host list table
    rows: List[str] = []
    for it in hosts:
        tag = str(it.get("tag") or "")
        label = str(it.get("label") or tag)
        base_rel = str(it.get("base_rel") or "")
        lscpu_rel = f"{base_rel}/lscpu.txt"
        lscpuj_rel = f"{base_rel}/lscpu.json"
        meminfo_rel = f"{base_rel}/meminfo.txt"

        view = f"/server-info/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(tag)}"
        dl_txt = f"/raw/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(lscpu_rel)}"
        dl_json = f"/raw/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(lscpuj_rel)}"
        dl_mem = f"/raw/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(meminfo_rel)}"
        rows.append(
            "<tr>"
            f"<td class=\"mono\">{_e(tag)}</td>"
            f"<td class=\"mono\">{_e(label)}</td>"
            f"<td><a href=\"{view}\">View</a></td>"
            f"<td><a href=\"{dl_txt}\" target=\"_blank\">lscpu.txt</a></td>"
            f"<td><a href=\"{dl_json}\" target=\"_blank\">lscpu.json</a></td>"
            f"<td><a href=\"{dl_mem}\" target=\"_blank\">meminfo.txt</a></td>"
            "</tr>"
        )

    host_table = (
        "<table class=\"table\">"
        "<thead><tr><th>tag</th><th>label</th><th>page</th><th>download</th><th>download</th><th>download</th></tr></thead>"
        f"<tbody>{''.join(rows) if rows else '<tr><td colspan=\"6\" class=\"muted\">No server_info found.</td></tr>'}</tbody>"
        "</table>"
    )

    # Content preview
    preview_html = "<div class=\"muted\">Select a host to view lscpu output.</div>"
    if chosen_entry is not None:
        base_rel = str(chosen_entry.get("base_rel") or "")
        lscpu_txt = (run.run_dir / base_rel / "lscpu.txt").resolve()
        if _is_within(lscpu_txt, run.run_dir) and lscpu_txt.exists() and lscpu_txt.is_file():
            txt = _read_text(lscpu_txt, max_bytes=2_000_000)
            preview_html = f"<pre class=\"mono\" style=\"white-space: pre-wrap\">{_e(txt)}</pre>"
        else:
            preview_html = f"<div class=\"warn mono\">Missing: {_e(str(lscpu_txt))}</div>"

    body = f"""
<section class="card">
  <div class="breadcrumbs">
    <a href="/">Home</a>
    <span class="sep">/</span>
    <a href="{run_href}">Run</a>
    <span class="sep">/</span>
    <span class="mono">server info</span>
  </div>
  <h1>{_e(title)}</h1>
  <div class="sub">run_dir: <span class="mono">{_e(run.run_dir)}</span></div>
</section>

<section class="card">
  <h2>Hosts</h2>
  {host_table}
</section>

<section class="card">
  <h2>lscpu</h2>
  {preview_html}
</section>
"""
    return _html_page(title, body)


def _render_csv_detail(scale_test_root: Path, run: RunInfo, csv_name: str, q: Dict[str, List[str]]) -> str:
    csv_path = (run.analysis_dir / csv_name).resolve()
    if not _is_within(csv_path, run.analysis_dir) or not csv_path.exists() or csv_path.suffix.lower() != ".csv":
        return _html_page("CSV not found", f"<section class=\"card\"><h1>CSV not found</h1><div class=\"mono\">{_e(csv_name)}</div></section>")

    # If plots/summaries are missing for this CSV, try auto-generate.
    required: List[Path] = [run.analysis_dir / "run_summary.html"]
    if csv_name == "token_len_scaling.csv":
        required += [run.analysis_dir / "plot_token_len_scaling.png"]
    elif csv_name == "batch_size_scaling.csv":
        required += [run.analysis_dir / "plot_batch_size_scaling.png"]
    elif csv_name == "cpu_scaling.csv":
        required += [run.analysis_dir / "plot_cpu_scaling.png"]
    elif csv_name == "kv_scaling.csv":
        required += [run.analysis_dir / "plot_kv_cap_scaling.png"]
    elif csv_name == "emon_socket_metrics.csv":
        # Per-job pies are driven by a manifest produced by analyze_run.py
        required += [run.analysis_dir / "emon_job_pies_manifest.json"]
    _maybe_autogen_analysis(scale_test_root=scale_test_root, run=run, required=required)

    # Special-case: emon_socket_metrics.csv page should only show preview.
    preview_only = csv_name == "emon_socket_metrics.csv"

    # Summary stats (skip for preview-only pages)
    num_rows = 0
    num_cols = 0
    num_table = ""
    if not preview_only:
        summ = _csv_numeric_summary(csv_path)
        numeric = summ.get("numeric") or {}
        num_rows = int(summ.get("rows") or 0)
        num_cols = int(summ.get("cols") or 0)

        # show up to N numeric columns
        num_items = list(numeric.items())
        num_items.sort(key=lambda kv: -(kv[1].get("count") or 0))
        num_items = num_items[:40]

    def fmtf(x: Any) -> str:
        try:
            return f"{float(x):.4g}"
        except Exception:
            return ""

        num_table_rows = []
        for col, st in num_items:
            num_table_rows.append(
                "<tr>"
                f"<td class=\"mono\">{_e(col)}</td>"
                f"<td class=\"mono\">{_e(st.get('count',''))}</td>"
                f"<td class=\"mono\">{_e(fmtf(st.get('mean')))}</td>"
                f"<td class=\"mono\">{_e(fmtf(st.get('min')))}</td>"
                f"<td class=\"mono\">{_e(fmtf(st.get('max')))}</td>"
                "</tr>"
            )

        num_table = (
            "<div class=\"table-scroll\">"
            "<table class=\"table small\">"
            "<thead><tr><th>numeric col</th><th>count</th><th>mean</th><th>min</th><th>max</th></tr></thead>"
            f"<tbody>{''.join(num_table_rows) if num_table_rows else '<tr><td colspan=\"5\" class=\"muted\">No numeric columns detected.</td></tr>'}</tbody>"
            "</table>"
            "</div>"
        )

    # Preview rows (default: 10 data rows)
    preview_rows_s = (q.get("preview_rows") or ["10"])[0].strip()
    try:
        preview_rows = max(1, min(500, int(preview_rows_s)))
    except Exception:
        preview_rows = 10

    # Preview
    preview_html = _render_csv_preview(csv_path, max_rows=preview_rows, run=run)

    # Preview controls (no JS)
    self_href = f"/csv/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(csv_name)}"
    preview_controls = (
        "<div class=\"sub\">Preview rows: "
        f"<a href=\"{self_href}?preview_rows=10\">10</a> "
        f"<a href=\"{self_href}?preview_rows=50\">50</a> "
        f"<a href=\"{self_href}?preview_rows=200\">200</a> "
        f"(current={_e(preview_rows)})"
        "</div>"
    )

    # Plots associated with this CSV
    plot_html = "<div class=\"muted\">No plots for this CSV.</div>"
    plot_paths: List[Path] = []
    if not preview_only:
        if csv_name == "token_len_scaling.csv":
            plot_paths = [run.analysis_dir / "plot_token_len_scaling.png"]
        elif csv_name == "batch_size_scaling.csv":
            plot_paths = [run.analysis_dir / "plot_batch_size_scaling.png"]
        elif csv_name == "cpu_scaling.csv":
            plot_paths = [run.analysis_dir / "plot_cpu_scaling.png"]
        elif csv_name == "kv_scaling.csv":
            plot_paths = [run.analysis_dir / "plot_kv_cap_scaling.png"]

    cards = []
    for p in plot_paths:
        if not p.exists():
            continue
        rel = p.relative_to(run.run_dir).as_posix()
        src = f"/raw/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(rel)}"
        cards.append(
            "<div class=\"plot\">"
            f"<div class=\"plot-title\">{_e(p.name)}</div>"
            f"<a href=\"{src}\" target=\"_blank\"><img src=\"{src}\" alt=\"{_e(p.name)}\" /></a>"
            "</div>"
        )
    if cards:
        plot_html = f"<div class=\"plots\">{''.join(cards)}</div>"

    # PPT-ready summary for scaling CSVs
    is_scaling = csv_name in {"token_len_scaling.csv", "batch_size_scaling.csv", "cpu_scaling.csv", "kv_scaling.csv"}
    ppt_summary_html = ""
    if is_scaling:
        ppt_summary_html = _render_scaling_csv_ppt_summary(csv_path)

    # Download link
    rel_csv = csv_path.relative_to(run.run_dir).as_posix()
    raw = f"/raw/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(rel_csv)}"
    back = f"/run/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}"


    title = f"{run.ref.task} / {run.ref.suite} / {run.ref.run_id} :: {csv_name}"
    if preview_only:
        body = f"""
<section class="card">
  <div class="breadcrumbs">
    <a href="/">Home</a>
    <span class="sep">/</span>
    <a href="{back}">{_e(run.ref.task)} / {_e(run.ref.suite)} / {_e(run.ref.run_id)}</a>
    <span class="sep">/</span>
    <span class="mono">{_e(csv_name)}</span>
  </div>
  <h1>CSV</h1>
  <div class="sub">file: <span class="mono">{_e(csv_name)}</span> • <a href="{raw}" target="_blank">Download</a></div>
  <div class="sub">This page intentionally shows preview only (no summary/plots) to keep the job navigation clean.</div>
</section>

<section class="card">
  <h2>Preview</h2>
    {preview_controls}
  {preview_html}
</section>
"""
    else:
        body = f"""
<section class="card">
  <div class="breadcrumbs">
    <a href="/">Home</a>
    <span class="sep">/</span>
    <a href="{back}">{_e(run.ref.task)} / {_e(run.ref.suite)} / {_e(run.ref.run_id)}</a>
    <span class="sep">/</span>
    <span class="mono">{_e(csv_name)}</span>
  </div>
  <h1>CSV</h1>
  <div class="sub">file: <span class="mono">{_e(csv_name)}</span> • rows~{_e(num_rows)} • cols~{_e(num_cols)} • <a href="{raw}" target="_blank">Download</a></div>
</section>

<section class="card">
    <h2>Summary</h2>
    {('<div class="sub">PPT 一页总结</div>' + ppt_summary_html + '<div class="sub">关键图</div>' + plot_html) if is_scaling else num_table}
</section>

{'' if is_scaling else f'''<section class="card">\n  <h2>Plots</h2>\n  {plot_html}\n</section>'''}

<section class="card">
  <h2>Preview</h2>
    {preview_controls}
  {preview_html}
</section>
"""
    return _html_page(title, body)


def _render_job_detail(scale_test_root: Path, run: RunInfo, job_name: str, q: Dict[str, List[str]], server_host: str = "") -> str:
    # Ensure analysis artifacts exist.
    required = [run.analysis_dir / "emon_socket_metrics.csv", run.analysis_dir / "emon_job_pies_manifest.json"]
    _maybe_autogen_analysis(scale_test_root=scale_test_root, run=run, required=required)

    csv_path = (run.analysis_dir / "emon_socket_metrics.csv").resolve()
    if not _is_within(csv_path, run.analysis_dir) or not csv_path.exists():
        return _html_page("Job not found", "<section class=\"card\"><h1>emon_socket_metrics.csv not found</h1></section>")

    # Load row for this job.
    text = _read_text(csv_path, max_bytes=20_000_000)
    reader = csv.DictReader(io.StringIO(text))
    row: Optional[Dict[str, str]] = None
    want_host = str(server_host or "").strip()
    for r in reader:
        if (r.get("job_name") or "").strip() != job_name:
            continue
        if want_host:
            if (r.get("server_host") or "").strip() != want_host:
                continue
        row = {k: ("" if v is None else str(v)) for k, v in r.items()}
        break

    back_run = f"/run/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}"
    back_csv = f"/csv/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/emon_socket_metrics.csv"
    if row is None:
        body_nf = f"""
<section class="card">
  <div class="breadcrumbs">
    <a href="/">Home</a>
    <span class="sep">/</span>
    <a href="{back_run}">{_e(run.ref.task)} / {_e(run.ref.suite)} / {_e(run.ref.run_id)}</a>
    <span class="sep">/</span>
    <a href="{back_csv}">emon_socket_metrics.csv</a>
    <span class="sep">/</span>
    <span class="mono">{_e(job_name)}</span>
  </div>
  <h1>Job</h1>
  <div class="warn mono">Job not found in emon_socket_metrics.csv</div>
</section>
"""
        return _html_page("Job not found", body_nf)

    # Sockets present.
    socket_keys = [k for k in row.keys() if k.startswith("socket_") and "__" in k]
    sockets = sorted({k.split("__", 1)[0] for k in socket_keys}, key=lambda s: (len(s), s))

    # Load pie manifest.
    manifest: Dict[str, Any] = {}
    man_path = (run.analysis_dir / "emon_job_pies_manifest.json").resolve()
    if _is_within(man_path, run.analysis_dir) and man_path.exists():
        try:
            manifest = json.loads(_read_text(man_path, max_bytes=8_000_000))
        except Exception:
            manifest = {}

    pies: Dict[str, str] = {}
    try:
        key = f"{want_host}::{job_name}" if want_host else job_name
        rec = manifest.get(key) or manifest.get(job_name) or {}
        pies = rec.get("pies") or {}
    except Exception:
        pies = {}

    # Summary.xlsx download link if available.
    xlsx_dl = ""
    emon_xlsx = (row.get("emon_summary_xlsx") or "").strip()
    if emon_xlsx:
        try:
            p = Path(emon_xlsx).resolve()
            if _is_within(p, run.run_dir) and p.exists() and p.is_file():
                rel = p.relative_to(run.run_dir).as_posix()
                dl = quote(f"{job_name}_summary.xlsx")
                href = f"/raw/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(rel)}?download_name={dl}"
                xlsx_dl = f'<a class="btn" href="{href}" target="_blank">Download summary.xlsx</a>'
        except Exception:
            pass

    # Metric list (as requested).
    metrics = [
        "metric_CPU operating frequency (in GHz)",
        "metric_uncore frequency GHz",
        "metric_DDR data rate (MT/sec)",
        "metric_UPI speed - GT/s",
        "metric_CPU utilization %",
        "metric_CPU utilization% in kernel mode",
        "metric_CPI",
        "metric_kernel_CPI",
        "metric_core IPC",
        "metric_package power (watts)",
        "metric_L1D MPI (includes data+rfo w/ prefetches)",
        "metric_L2 MPI (includes code+data+rfo w/ prefetches)",
        "metric_LLC MPI (includes code+data+rfo w/ prefetches)",
        "metric_NUMA %_Reads addressed to local DRAM",
        "metric_NUMA %_Reads addressed to remote DRAM",
        "metric_memory bandwidth read (MB/sec)",
        "metric_memory bandwidth write (MB/sec)",
        "metric_memory bandwidth total (MB/sec)",
        "metric_core c6 residency %",
        "metric_package c6 residency %",
        "metric_package c2 residency %",
        "metric_TMA_Frontend_Bound(%)",
        "metric_TMA_Bad_Speculation(%)",
        "metric_TMA_Retiring(%)",
        "metric_TMA_Backend_Bound(%)",
        "metric_TMA_..Memory_Bound(%)",
        "metric_TMA_....L1_Bound(%)",
        "metric_TMA_....L2_Bound(%)",
        "metric_TMA_....L3_Bound(%)",
        "metric_TMA_....DRAM_Bound(%)",
        "metric_TMA_....Store_Bound(%)",
        "metric_TMA_..Core_Bound(%)",
    ]

    def _fmt_metric(metric: str, v: str) -> str:
        st = (v or "").strip()
        if not st or st.lower() in {"nan", "none"}:
            return ""
        try:
            x = float(st)
        except Exception:
            return _e(st)
        m = metric.lower()
        is_pct = "(%)" in metric or "utilization" in m or "residency" in m or "tma_" in m
        if is_pct:
            return _e(f"{x:.2f}")
        if abs(x) >= 1000:
            return _e(f"{x:.0f}")
        if abs(x) >= 100:
            return _e(f"{x:.2f}")
        return _e(f"{x:.3f}")

    # Metric table.
    head = "<tr><th>metric</th>" + "".join(f"<th>{_e(s)}</th>" for s in sockets) + "</tr>"
    trs: List[str] = []
    for m in metrics:
        tds = [f"<td class=\"mono\">{_e(m)}</td>"]
        for s in sockets:
            key = f"{s}__{m}"
            tds.append(f"<td class=\"mono\">{_fmt_metric(m, row.get(key, ''))}</td>")
        trs.append("<tr>" + "".join(tds) + "</tr>")
    metric_table = (
        '<div class="table-scroll">'
        '<table class="table small preview-table">'
        f"<thead>{head}</thead><tbody>{''.join(trs)}</tbody>"
        "</table>"
        "</div>"
    )

    # Pie chart cards.
    pie_cards: List[str] = []
    for s in sockets:
        rel_png = str(pies.get(s) or "").strip()
        if not rel_png:
            continue
        try:
            p = (run.run_dir / rel_png).resolve()
            if not _is_within(p, run.run_dir) or not p.exists():
                continue
            rel = p.relative_to(run.run_dir).as_posix()
            src = f"/raw/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(rel)}"
            pie_cards.append(
                "<div class=\"plot\">"
                f"<div class=\"plot-title\">{_e(s)} TMA pie</div>"
                f"<a href=\"{src}\" target=\"_blank\"><img src=\"{src}\" alt=\"{_e(s)}\" /></a>"
                "</div>"
            )
        except Exception:
            continue
    pies_html = f"<div class=\"plots\">{''.join(pie_cards)}</div>" if pie_cards else '<div class="muted">No per-job TMA pie images found (try re-running analyze_run.py).</div>'

    # A few meta fields.
    meta_rows = []
    for k in ["variant", "resource_cpu", "resource_cpu_count", "sglang_max_total_tokens", "batch_size", "token_len"]:
        if k in row:
            meta_rows.append((k, row.get(k, "")))
    meta_trs = "".join(f"<tr><td class=\"mono\">{_e(k)}</td><td class=\"mono\">{_e(v)}</td></tr>" for k, v in meta_rows)
    meta_table = (
        '<div class="table-scroll">'
        '<table class="table small">'
        '<thead><tr><th>key</th><th>value</th></tr></thead>'
        f"<tbody>{meta_trs}</tbody></table></div>"
        if meta_rows
        else ""
    )

    title = f"{run.ref.task} / {run.ref.suite} / {run.ref.run_id} :: job {job_name}"
    host_line = f" • host: <span class=\"mono\">{_e(want_host)}</span>" if want_host else ""
    body = f"""
<section class="card">
  <div class="breadcrumbs">
    <a href="/">Home</a>
    <span class="sep">/</span>
    <a href="{back_run}">{_e(run.ref.task)} / {_e(run.ref.suite)} / {_e(run.ref.run_id)}</a>
    <span class="sep">/</span>
    <a href="{back_csv}">emon_socket_metrics.csv</a>
    <span class="sep">/</span>
    <span class="mono">{_e(job_name)}</span>
  </div>
  <h1>Job</h1>
    <div class="sub">job_name: <span class="mono">{_e(job_name)}</span>{host_line}</div>
  <div class="toolbar">{xlsx_dl}</div>
  {meta_table}
</section>

<section class="card">
  <h2>Socket TMA pies</h2>
  {pies_html}
</section>

<section class="card">
  <h2>Socket metrics</h2>
  {metric_table}
</section>
"""
    return _html_page(title, body)


def _render_csv_preview(path: Path, max_rows: int = 200, *, run: Optional[RunInfo] = None) -> str:
    # Special-case: make job_name clickable to download summary.xlsx.
    if path.name == "emon_socket_metrics.csv" and run is not None:
        text = _read_text(path, max_bytes=2_000_000)
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return '<div class="muted">Empty CSV.</div>'

        fieldnames = list(reader.fieldnames)
        rows: List[Dict[str, str]] = []
        for i, r in enumerate(reader):
            rows.append({k: ("" if v is None else str(v)) for k, v in r.items()})
            if i + 1 >= max_rows:
                break

        def _td_html(cell_html: str, *, is_head: bool = False, cls: str = "") -> str:
            tag = "th" if is_head else "td"
            cls_attr = f' class="{cls}"' if cls else ""
            return f"<{tag}{cls_attr}>{cell_html}</{tag}>"

        def _td_text(cell: str, *, is_head: bool = False) -> str:
            tag = "th" if is_head else "td"
            cls = "mono" if (len(cell) < 64 and any(ch.isdigit() for ch in cell)) else ""
            return f"<{tag} class=\"{cls}\">{_e(cell)}</{tag}>"

        def _summary_href(emon_path_s: str, job_name: str) -> Optional[str]:
            s = (emon_path_s or "").strip()
            if not s:
                return None
            try:
                p = Path(s).resolve()
                if not _is_within(p, run.run_dir) or not p.exists() or not p.is_file():
                    return None
                rel = p.relative_to(run.run_dir).as_posix()
                dl = quote(f"{job_name}_summary.xlsx")
                return f"/raw/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{_e(rel)}?download_name={dl}"
            except Exception:
                return None

        def _job_href(job_name: str, server_host: str) -> str:
            sh = (server_host or "").strip()
            if sh:
                return f"/job/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{quote(sh)}/{quote(job_name)}"
            return f"/job/{_e(run.ref.task)}/{_e(run.ref.suite)}/{_e(run.ref.run_id)}/{quote(job_name)}"

        thead = "<tr>" + "".join(_td_text(c, is_head=True) for c in fieldnames) + "</tr>"
        body_trs: List[str] = []
        for r in rows:
            tds: List[str] = []
            job = r.get("job_name", "")
            sh = r.get("server_host", "")
            emon_xlsx = r.get("emon_summary_xlsx", "")
            href = _summary_href(emon_xlsx, job)
            for c in fieldnames:
                v = r.get(c, "")
                if c == "job_name" and job:
                    tds.append(_td_html(f'<a href="{_job_href(job, sh)}">{_e(job)}</a>', cls="mono"))
                elif c == "emon_summary_xlsx" and href:
                    tds.append(_td_html(f'<a href="{href}" target="_blank">summary.xlsx</a>', cls="mono"))
                else:
                    tds.append(_td_text(v, is_head=False))
            body_trs.append("<tr>" + "".join(tds) + "</tr>")

        tbody = "".join(body_trs)
        return (
            f'<div class="sub">Preview: <span class="mono">{_e(path.name)}</span> (first {len(rows)} rows) — click <span class="mono">job_name</span> to open the job page</div>'
            '<div class="table-scroll preview-scroll">'
            f'<table class="table small preview-table"><thead>{thead}</thead><tbody>{tbody}</tbody></table>'
            "</div>"
        )

    # Default: stream the CSV and escape for HTML.
    buf = io.StringIO(_read_text(path, max_bytes=2_000_000))
    reader2 = csv.reader(buf)
    header = next(reader2, None)
    if not header:
        return '<div class="muted">Empty CSV.</div>'
    body_rows: List[List[str]] = []
    for i, row in enumerate(reader2):
        body_rows.append([str(c) for c in row])
        if i + 1 >= max_rows:
            break

    def td(cell: str, is_head: bool = False) -> str:
        tag = "th" if is_head else "td"
        cls = "mono" if (len(cell) < 64 and any(ch.isdigit() for ch in cell)) else ""
        return f"<{tag} class=\"{cls}\">{_e(cell)}</{tag}>"

    thead2 = "<tr>" + "".join(td(c, is_head=True) for c in header) + "</tr>"
    tbody2 = "".join("<tr>" + "".join(td(c, is_head=False) for c in row) + "</tr>" for row in body_rows)
    return (
        f'<div class="sub">Preview: <span class="mono">{_e(path.name)}</span> (first {len(body_rows)} rows)</div>'
        '<div class="table-scroll preview-scroll">'
        f'<table class="table small preview-table"><thead>{thead2}</thead><tbody>{tbody2}</tbody></table>'
        "</div>"
    )


def _is_within(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except Exception:
        return False


class Handler(BaseHTTPRequestHandler):
    server_version = "scale-test-web/0.1"

    def do_POST(self) -> None:  # noqa: N802
        try:
            self._do_POST()
        except Exception as e:
            self._send(500, f"Internal error: {e}\n", content_type="text/plain")

    def do_GET(self) -> None:  # noqa: N802
        try:
            self._do_GET()
        except Exception as e:
            self._send(500, f"Internal error: {e}\n", content_type="text/plain")

    def _do_GET(self) -> None:
        srv: "WebServer" = self.server  # type: ignore[assignment]
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        q = parse_qs(parsed.query)

        if path == "/" or path == "":
            runs = discover_runs(srv.scale_test_root)
            html_text = _render_home(srv.scale_test_root, runs, q)
            self._send(200, html_text, content_type="text/html; charset=utf-8")
            return

        if path == "/compare":
            runs = discover_runs(srv.scale_test_root)
            html_text = _render_compare(srv.scale_test_root, runs, q)
            self._send(200, html_text, content_type="text/html; charset=utf-8")
            return

        if path == "/compare-csv":
            runs = discover_runs(srv.scale_test_root)
            html_text = _render_compare_csv(srv.scale_test_root, runs, q)
            self._send(200, html_text, content_type="text/html; charset=utf-8")
            return

        if path == "/compare-csv-export.xlsx":
            a_raw = (q.get("a") or [""])[0].strip()
            b_raw = (q.get("b") or [""])[0].strip()
            csv_name = (q.get("csv") or [""])[0].strip()
            a_ref = _parse_run_key(a_raw)
            b_ref = _parse_run_key(b_raw)
            if not (a_ref and b_ref and csv_name):
                self._send(400, "Missing parameters. Need: a, b, csv.\n", content_type="text/plain")
                return
            run_a = srv.get_run(a_ref.task, a_ref.suite, a_ref.run_id)
            run_b = srv.get_run(b_ref.task, b_ref.suite, b_ref.run_id)
            if run_a is None or run_b is None:
                self._send(404, "Run not found\n", content_type="text/plain")
                return
            a_p = (run_a.analysis_dir / csv_name).resolve()
            b_p = (run_b.analysis_dir / csv_name).resolve()
            if not (_is_within(a_p, run_a.run_dir) and _is_within(b_p, run_b.run_dir)):
                self._send(403, "Forbidden\n", content_type="text/plain")
                return
            try:
                payload = _build_compare_csv_export_xlsx(
                    a_path=a_p,
                    b_path=b_p,
                    a_raw=a_raw,
                    b_raw=b_raw,
                    csv_name=csv_name,
                    run_a=run_a,
                    run_b=run_b,
                )
            except Exception as e:
                self._send(500, f"Export failed: {e}\n", content_type="text/plain")
                return
            dl = _sanitize_download_name(f"compare_{Path(csv_name).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", default_ext=".xlsx")
            self._send_bytes(
                200,
                payload,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                download_name=dl,
            )
            return

        if path.startswith("/static/"):
            rel = path[len("/static/") :]
            file_path = (WEB_DIR / "static" / rel).resolve()
            if not _is_within(file_path, WEB_DIR / "static") or not file_path.exists():
                self._send(404, "Not Found\n", content_type="text/plain")
                return
            self._send_file(file_path)
            return

        if path.startswith("/run/"):
            parts = [p for p in path.split("/") if p]
            if len(parts) != 4:
                self._send(404, "Not Found\n", content_type="text/plain")
                return
            _, task, suite, run_id = parts
            run = srv.get_run(task, suite, run_id)
            if run is None:
                self._send(404, "Run not found\n", content_type="text/plain")
                return
            html_text = _render_run(srv.scale_test_root, run, q)
            self._send(200, html_text, content_type="text/html; charset=utf-8")
            return

        if path.startswith("/server-info/"):
            parts = [p for p in path.split("/") if p]
            if len(parts) not in {4, 5}:
                self._send(404, "Not Found\n", content_type="text/plain")
                return
            host_tag = ""
            if len(parts) == 4:
                _, task, suite, run_id = parts
            else:
                _, task, suite, run_id, host_tag = parts
            run = srv.get_run(task, suite, run_id)
            if run is None:
                self._send(404, "Run not found\n", content_type="text/plain")
                return
            html_text = _render_server_info(srv.scale_test_root, run, q, host_tag=host_tag)
            self._send(200, html_text, content_type="text/html; charset=utf-8")
            return

        if path.startswith("/csv/"):
            parts = [p for p in path.split("/") if p]
            if len(parts) != 5:
                self._send(404, "Not Found\n", content_type="text/plain")
                return
            _, task, suite, run_id, csv_name = parts
            run = srv.get_run(task, suite, run_id)
            if run is None:
                self._send(404, "Run not found\n", content_type="text/plain")
                return
            html_text = _render_csv_detail(srv.scale_test_root, run, csv_name, q)
            self._send(200, html_text, content_type="text/html; charset=utf-8")
            return

        if path.startswith("/job/"):
            parts = [p for p in path.split("/") if p]
            if len(parts) not in {5, 6}:
                self._send(404, "Not Found\n", content_type="text/plain")
                return
            server_host = ""
            if len(parts) == 5:
                _, task, suite, run_id, job_name = parts
            else:
                _, task, suite, run_id, server_host, job_name = parts
            run = srv.get_run(task, suite, run_id)
            if run is None:
                self._send(404, "Run not found\n", content_type="text/plain")
                return
            html_text = _render_job_detail(srv.scale_test_root, run, job_name, q, server_host=server_host)
            self._send(200, html_text, content_type="text/html; charset=utf-8")
            return

        if path.startswith("/raw/"):
            parts = [p for p in path.split("/") if p]
            if len(parts) < 5:
                self._send(404, "Not Found\n", content_type="text/plain")
                return
            _, task, suite, run_id, *rest = parts
            run = srv.get_run(task, suite, run_id)
            if run is None:
                self._send(404, "Run not found\n", content_type="text/plain")
                return
            rel = "/".join(rest)
            file_path = (run.run_dir / rel).resolve()
            if not _is_within(file_path, run.run_dir) or not file_path.exists() or not file_path.is_file():
                self._send(404, "Not Found\n", content_type="text/plain")
                return
            dl_name = (q.get("download_name") or q.get("filename") or [""])[0].strip()
            if dl_name:
                # If the client asks for a custom name, always treat as attachment.
                # Keep extension consistent with served content when possible.
                ext = file_path.suffix if file_path.suffix else None
                safe_name = _sanitize_download_name(dl_name, default_ext=ext)
                self._send_file(file_path, download_name=safe_name)
            else:
                self._send_file(file_path)
            return

        self._send(404, "Not Found\n", content_type="text/plain")

    def _do_POST(self) -> None:
        srv: "WebServer" = self.server  # type: ignore[assignment]
        parsed = urlparse(self.path)
        path = unquote(parsed.path)

        length = int(self.headers.get("Content-Length") or "0")
        raw = self.rfile.read(length) if length > 0 else b""
        try:
            form = parse_qs(raw.decode("utf-8", errors="replace"))
        except Exception:
            form = {}

        if path == "/delete":
            sels = [s for s in (form.get("sel") or []) if str(s).strip()]

            deleted = 0
            failed = 0
            for run_key in sels:
                parts = [p for p in str(run_key).split("/") if p]
                if len(parts) != 3:
                    failed += 1
                    continue
                task, suite, run_id = parts

                # Match WebServer.get_run() layout.
                run_dir = (srv.scale_test_root / task / "result" / suite / run_id).resolve()
                analysis_dir = run_dir / "analysis"
                if not _is_within(run_dir, srv.scale_test_root):
                    failed += 1
                    continue
                if not run_dir.exists() or not run_dir.is_dir():
                    failed += 1
                    continue
                # Only delete runs that look like runs (analysis dir exists).
                if not analysis_dir.exists():
                    failed += 1
                    continue
                try:
                    shutil.rmtree(run_dir)
                    deleted += 1
                except Exception:
                    failed += 1

            # Redirect back to home.
            self.send_response(303)
            self.send_header("Location", f"/?deleted={deleted}&delete_failed={failed}")
            self.end_headers()
            return

        self._send(404, "Not Found\n", content_type="text/plain")

    def _send(self, code: int, body: str, *, content_type: str) -> None:
        data = body.encode("utf-8", errors="replace")
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _send_bytes(self, code: int, data: bytes, *, content_type: str, download_name: Optional[str] = None) -> None:
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        if download_name:
            safe = _sanitize_download_name(download_name, default_ext=".xlsx")
            self.send_header("Content-Disposition", f'attachment; filename="{safe}"')
        self.end_headers()
        self.wfile.write(data)

    def _send_file(self, path: Path, *, download_name: Optional[str] = None) -> None:
        ctype, _ = mimetypes.guess_type(str(path))
        if not ctype:
            ctype = "application/octet-stream"
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        if download_name:
            safe = _sanitize_download_name(download_name, default_ext=path.suffix)
            self.send_header("Content-Disposition", f'attachment; filename="{safe}"')
        # allow browser caching for static-ish artifacts
        self.send_header("Cache-Control", "public, max-age=60")
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, fmt: str, *args: Any) -> None:
        # Keep logs concise.
        syslog = os.environ.get("SCALE_TEST_WEB_LOG", "1")
        if syslog != "0":
            super().log_message(fmt, *args)


class WebServer(ThreadingHTTPServer):
    def __init__(self, server_address: Tuple[str, int], handler_cls: Any, *, scale_test_root: Path):
        super().__init__(server_address, handler_cls)
        self.scale_test_root = scale_test_root

    def get_run(self, task: str, suite: str, run_id: str) -> Optional[RunInfo]:
        run_dir = (self.scale_test_root / task / "result" / suite / run_id).resolve()
        analysis_dir = run_dir / "analysis"
        if not _is_within(run_dir, self.scale_test_root) or not analysis_dir.exists():
            return None
        try:
            mtime = run_dir.stat().st_mtime
        except Exception:
            mtime = 0.0
        return RunInfo(ref=RunRef(task=task, suite=suite, run_id=run_id), run_dir=run_dir, analysis_dir=analysis_dir, mtime=mtime)


def main() -> int:
    ap = argparse.ArgumentParser(description="Scale-test results web UI")
    ap.add_argument("--host", default="0.0.0.0")
    ap.add_argument("--port", type=int, default=8080)
    ap.add_argument(
        "--scale-test-root",
        default=str(SCALE_TEST_ROOT_DEFAULT),
        help="Path to scripts/scale-test (default: inferred from this file)",
    )
    args = ap.parse_args()

    scale_test_root = Path(args.scale_test_root).resolve()
    if not scale_test_root.exists():
        raise SystemExit(f"scale-test root not found: {scale_test_root}")

    # Warm up mimetypes.
    mimetypes.init()

    httpd = WebServer((args.host, int(args.port)), Handler, scale_test_root=scale_test_root)
    print(f"[ok] Serving: http://{args.host}:{args.port}/")
    print(f"[ok] Root: {scale_test_root}")
    try:
        httpd.serve_forever(poll_interval=0.5)
    except KeyboardInterrupt:
        pass
    finally:
        httpd.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
