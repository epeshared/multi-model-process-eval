#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import ipaddress
import json
import os
import socket
import shlex
import shutil
import subprocess
import sys
import tarfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from urllib.parse import urlparse


REPO_ROOT = Path(__file__).resolve().parents[3]


_LOCAL_ADDR_CACHE: Optional[set[str]] = None


def _local_addrs() -> set[str]:
    """Best-effort set of IP strings that should be treated as local."""

    global _LOCAL_ADDR_CACHE
    if _LOCAL_ADDR_CACHE is not None:
        return set(_LOCAL_ADDR_CACHE)

    addrs: set[str] = set()
    # Canonical loopback aliases.
    addrs.update({"127.0.0.1", "::1"})

    def _add_from_getaddrinfo(name: str) -> None:
        try:
            for fam, _, _, _, sockaddr in socket.getaddrinfo(name, None):
                if fam == socket.AF_INET:
                    addrs.add(str(sockaddr[0]))
                elif fam == socket.AF_INET6:
                    addrs.add(str(sockaddr[0]))
        except Exception:
            return

    # Hostname/fqdn/localhost often cover the common on-box addresses.
    _add_from_getaddrinfo("localhost")
    try:
        _add_from_getaddrinfo(socket.gethostname())
    except Exception:
        pass
    try:
        _add_from_getaddrinfo(socket.getfqdn())
    except Exception:
        pass

    # Best-effort: hostname -I (Linux) provides all interface addresses.
    try:
        p = subprocess.run(["hostname", "-I"], stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, text=True, check=False)
        if p.returncode == 0 and p.stdout:
            for tok in (p.stdout or "").split():
                t = tok.strip()
                if t:
                    addrs.add(t)
    except Exception:
        pass

    _LOCAL_ADDR_CACHE = set(addrs)
    return set(addrs)


def _is_local_host(host: str) -> bool:
    """Return True if host refers to the current machine."""

    h = str(host or "").strip()
    if not h:
        return False
    if h.lower() in {"localhost", "ip6-localhost", "ip6-loopback"}:
        return True
    if h in {"127.0.0.1", "::1", "0.0.0.0", "::"}:
        return True

    # Direct IP literal.
    try:
        ip = ipaddress.ip_address(h)
        if ip.is_loopback or ip.is_unspecified:
            return True
        return h in _local_addrs()
    except Exception:
        pass

    # Resolve hostname -> IPs and compare.
    try:
        local = _local_addrs()
        for fam, _, _, _, sockaddr in socket.getaddrinfo(h, None):
            if fam == socket.AF_INET:
                if str(sockaddr[0]) in local:
                    return True
            elif fam == socket.AF_INET6:
                if str(sockaddr[0]) in local:
                    return True
    except Exception:
        pass
    return False


def _sync_tree(src_dir: Path, dst_dir: Path) -> None:
    """Rsync-like copy from src_dir into dst_dir (overwrite files, create dirs)."""

    src = Path(src_dir)
    dst = Path(dst_dir)
    if not src.exists() or not src.is_dir():
        return
    dst.mkdir(parents=True, exist_ok=True)
    for root, dirs, files in os.walk(str(src)):
        root_p = Path(root)
        rel = root_p.relative_to(src)
        out_root = dst / rel
        out_root.mkdir(parents=True, exist_ok=True)
        for d in dirs:
            (out_root / d).mkdir(parents=True, exist_ok=True)
        for fn in files:
            s = root_p / fn
            t = out_root / fn
            try:
                t.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(str(s), str(t))
            except Exception:
                # Best-effort: ignore individual copy failures.
                pass


def _utc_compact() -> str:
    return dt.datetime.now(tz=dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _dump_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _which(bin_name: str) -> Optional[str]:
    try:
        return shutil.which(bin_name)
    except Exception:
        return None


def _parse_cpus(expr: str) -> List[int]:
    s = (expr or "").strip()
    if not s:
        return []
    out: List[int] = []
    for part in s.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            try:
                lo = int(a.strip())
                hi = int(b.strip())
            except Exception:
                continue
            if hi < lo:
                lo, hi = hi, lo
            for x in range(lo, hi + 1):
                out.append(x)
        else:
            try:
                out.append(int(part))
            except Exception:
                continue
    return sorted(set(out))


def _cpu_count(expr: str) -> int:
    """Return number of logical CPUs described by a linux CPU list expression."""

    try:
        return len(_parse_cpus(expr))
    except Exception:
        return 0


def _cpu_to_numa_node(cpu_id: int) -> Optional[int]:
    """Best-effort map CPU id -> NUMA node id using sysfs."""

    cpu_dir = Path(f"/sys/devices/system/cpu/cpu{int(cpu_id)}")
    if not cpu_dir.exists():
        return None
    try:
        for child in cpu_dir.iterdir():
            name = child.name
            if name.startswith("node") and name[4:].isdigit():
                return int(name[4:])
    except Exception:
        return None
    return None


def _infer_single_numa_node_from_cpu_expr(cpu_expr: str) -> Optional[int]:
    """Infer a single NUMA node id if all requested CPUs map to the same node."""

    cpus = _parse_cpus(cpu_expr)
    if not cpus:
        return None

    nodes: List[int] = []
    for cpu in cpus[:256]:
        n = _cpu_to_numa_node(cpu)
        if n is None:
            return None
        nodes.append(n)

    uniq = sorted(set(nodes))
    if len(uniq) == 1:
        return uniq[0]
    return None


def _bytes_from_gb(gb: Optional[float]) -> Optional[int]:
    if gb is None:
        return None
    try:
        return int(float(gb) * (1024**3))
    except Exception:
        return None


def _try_systemd_scope_cmd(
    *,
    cmd: List[str],
    cpu_expr: str,
    mem_bytes: Optional[int],
) -> Optional[List[str]]:
    systemd_run = _which("systemd-run")
    if not systemd_run:
        return None

    # Some environments accept systemd-run properties but do not actually
    # enforce cpuset constraints (especially with --user scopes). As an extra
    # guardrail, apply an explicit affinity mask via taskset *inside* the scope.
    cmd2 = list(cmd)
    taskset_bin = _which("taskset")
    if (cpu_expr or "").strip() and taskset_bin:
        cmd2 = [taskset_bin, "-c", cpu_expr.strip()] + cmd2

    props: List[str] = []
    if (cpu_expr or "").strip():
        # For transient scope units, systemd commonly supports AllowedCPUs= (cpuset).
        # CPUAffinity= is a service property and may be rejected for scopes.
        props += ["-p", f"AllowedCPUs={cpu_expr.strip()}"]
    if mem_bytes and mem_bytes > 0:
        props += ["-p", f"MemoryMax={mem_bytes}"]

    # Try user scope first (works in most non-root environments).
    try:
        test = subprocess.run(
            [systemd_run, "--user", "--scope", "--quiet", "--", "true"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
        if test.returncode == 0:
            return [systemd_run, "--user", "--scope"] + props + ["--"] + cmd2
    except Exception:
        pass

    # Fallback: system scope (may require privileges but sometimes works).
    try:
        test = subprocess.run(
            [systemd_run, "--scope", "--quiet", "--", "true"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=False,
        )
        if test.returncode == 0:
            return [systemd_run, "--scope"] + props + ["--"] + cmd2
    except Exception:
        pass

    return None


def _wrap_cmd_taskset_prlimit(
    *,
    cmd: List[str],
    cpu_expr: str,
    mem_bytes: Optional[int],
) -> List[str]:
    out = list(cmd)

    prlimit = _which("prlimit")
    if mem_bytes and mem_bytes > 0 and prlimit:
        out = [prlimit, f"--as={mem_bytes}", "--"] + out

    taskset = _which("taskset")
    if (cpu_expr or "").strip() and taskset:
        out = [taskset, "-c", cpu_expr.strip()] + out

    return out


def _constrained_cmd(*, cmd: List[str], cpu_expr: str, mem_bytes: Optional[int]) -> Tuple[List[str], str]:
    # If no constraints are requested, return the command unchanged.
    if not (cpu_expr or "").strip() and not (mem_bytes and mem_bytes > 0):
        return list(cmd), "none"

    sysd = _try_systemd_scope_cmd(cmd=cmd, cpu_expr=cpu_expr, mem_bytes=mem_bytes)
    if sysd is not None:
        return sysd, "systemd-run"

    wrapped = _wrap_cmd_taskset_prlimit(cmd=cmd, cpu_expr=cpu_expr, mem_bytes=mem_bytes)
    method = "taskset/prlimit" if wrapped != cmd else "none"
    return wrapped, method


def _safe_name(s: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "-_." else "_" for ch in (s or "").strip())


def _parse_bool(x: Any) -> bool:
    if isinstance(x, bool):
        return x
    s = str(x or "").strip().lower()
    if s in ("1", "true", "t", "yes", "y", "on"):
        return True
    if s in ("0", "false", "f", "no", "n", "off", ""):
        return False
    raise SystemExit(f"invalid boolean value: {x}")


def _validate_remote_repo_dir_for_delete(remote_repo_dir: str) -> str:
    """Defensive validation for destructive operations on remote paths."""
    p = str(remote_repo_dir or "").strip()
    if not p:
        raise SystemExit("refusing to delete empty remote_repo_dir")
    if not p.startswith("/"):
        raise SystemExit(f"refusing to delete non-absolute remote_repo_dir: {p}")
    norm = os.path.normpath(p)
    # Refuse obvious dangerous paths.
    forbidden = {
        "/",
        "/home",
        "/root",
        "/tmp",
        "/var",
        "/usr",
        "/opt",
        "/mnt",
        "/media",
    }
    if norm in forbidden:
        raise SystemExit(f"refusing to delete dangerous remote_repo_dir: {norm}")
    # Require at least /a/b/c (three components) to reduce foot-guns.
    parts = [x for x in norm.split("/") if x]
    if len(parts) < 3:
        raise SystemExit(f"refusing to delete shallow remote_repo_dir: {norm}")
    return norm


def _as_str_list(x: Any) -> List[str]:
    if x is None:
        return []
    if isinstance(x, list):
        return [str(v).strip() for v in x if str(v).strip()]
    s = str(x).strip()
    return [s] if s else []


@dataclass(frozen=True)
class PreRequirement:
    file: str
    stage: str  # before_conda | after_conda
    mode: str  # bash | source
    source_bashrc_after: bool


@dataclass(frozen=True)
class ServerConfig:
    host: str
    user: str
    password: str
    port: Optional[int]
    identity_file: str
    ssh_options: List[str]
    remote_repo_dir: str
    remote_result_root: Optional[str]
    remote_python: str
    conda_env: str
    install_requirements: bool
    requirements_profile: str
    requirements_files: List[str]
    pip_extra_args: List[str]
    pre_setup_cmds: List[str]
    pre_requirements: List[PreRequirement]


@dataclass(frozen=True)
class SSHDispatchConfig:
    servers: List[ServerConfig]


@dataclass(frozen=True)
class ScaleConfig:
    template_auto_test_config: Path
    result_root: Path
    token_lens: List[int]
    batch_sizes: List[int]
    repeats: int
    repeat_threshold_sec: Optional[float]
    repeat_max_repeats: int
    warmup_runs: int
    continue_on_error: bool
    cpu_exprs: List[str]
    mem_gb: Optional[float]
    sglang_max_total_tokens: List[str]
    job_template: Dict[str, Any]
    emon_process_after_run: bool
    emon_process_cmd: List[str]
    emon_expected_output: str
    dispatch: SSHDispatchConfig


def _parse_scale_config(cfg_path: Path, *, resolve_ssh_passwords: bool = True) -> ScaleConfig:
    raw = _load_json(cfg_path)
    template = Path(str(raw.get("template_auto_test_config") or "scripts/auto-test/embedding/config_fix_token_len.json"))
    if not template.is_absolute():
        template = (REPO_ROOT / template).resolve()

    result_root = Path(str(raw.get("result_root") or "scripts/scale-test/embedding/result/fix_token_len"))
    if not result_root.is_absolute():
        result_root = (REPO_ROOT / result_root).resolve()

    run = raw.get("run") or {}
    token_lens = [int(x) for x in (run.get("token_lens") or [])]
    if not token_lens:
        raise SystemExit("config.run.token_lens is empty")

    batch_sizes_raw = run.get("batch_sizes")
    batch_sizes: List[int] = []
    if batch_sizes_raw is not None:
        batch_sizes = [int(x) for x in (batch_sizes_raw or [])]

    repeats = int(run.get("repeats") or 1)

    # Optional: time-threshold-based repeats.
    # Compatible aliases (including the user-typed misspelling):
    # - repeat_threshold_sec
    # - repeat-threash-hold
    repeat_threshold_raw = (
        run.get("repeat_threshold_sec")
        or run.get("repeat_threshold")
        or run.get("repeat-threshold")
        or run.get("repeat_threash_hold")
        or run.get("repeat-threash-hold")
        or run.get("repeat_threash_hold_sec")
        or run.get("repeat-threash-hold-sec")
    )
    repeat_threshold_sec: Optional[float] = None
    if repeat_threshold_raw is not None and str(repeat_threshold_raw).strip() != "":
        try:
            v = float(repeat_threshold_raw)
        except Exception:
            raise SystemExit("config.run.repeat_threshold_sec must be a number > 0")
        if v > 0:
            repeat_threshold_sec = float(v)

    # Safety default: if threshold mode is not enabled and user didn't explicitly
    # set repeats, treat repeats as 1.
    if repeat_threshold_sec is None and ("repeats" not in run):
        repeats = 1

    repeat_max_raw = run.get("repeat_max_repeats") or run.get("repeat_max") or run.get("max_repeats")
    repeat_max_repeats = 100
    if repeat_max_raw is not None and str(repeat_max_raw).strip() != "":
        try:
            repeat_max_repeats = max(1, int(float(repeat_max_raw)))
        except Exception:
            raise SystemExit("config.run.repeat_max_repeats must be an integer >= 1")
    warmup_runs = int(run.get("warmup_runs") or 0)

    continue_on_error = bool(run.get("continue_on_error") or run.get("continue_on_failure") or False)

    cpu_raw = (run.get("cpu") or {}).get("cpus")
    cpu_exprs = _as_str_list(cpu_raw)
    if not cpu_exprs:
        cpu_exprs = [""]

    mem_gb = (run.get("memory") or {}).get("max_gb")
    mem_gb_f: Optional[float] = None
    if mem_gb is not None and str(mem_gb).strip() != "":
        mem_gb_f = float(mem_gb)

    sglang_mtt = _as_str_list(run.get("sglang_max_total_tokens"))

    job_template = raw.get("job_template") or {}
    if not isinstance(job_template, dict):
        raise SystemExit("config.job_template must be an object")

    # Backward compatible: if run.sglang_max_total_tokens is not provided,
    # we still allow a single value in job_template.extra_env.
    if not sglang_mtt:
        extra_env = job_template.get("extra_env") or {}
        if not isinstance(extra_env, dict):
            extra_env = {}
        v = str(extra_env.get("SGLANG_MAX_TOTAL_TOKENS") or "").strip()
        sglang_mtt = [v] if v else [""]

    if not batch_sizes:
        # Default to job_template.batch_size if not specified.
        try:
            batch_sizes = [int((job_template.get("batch_size") or 0))]
        except Exception:
            batch_sizes = []
    if not batch_sizes or any(bs <= 0 for bs in batch_sizes):
        raise SystemExit("config.run.batch_sizes is empty/invalid (or job_template.batch_size invalid)")

    emon = raw.get("emon") or {}
    emon_process_after_run = bool(emon.get("process_after_run", True))
    process_cmd = emon.get("process_cmd") or [
        "emon",
        "-process-pyedp",
        "/opt/intel/sep/config/edp/pyedp_config.txt",
    ]
    if not isinstance(process_cmd, list) or not all(isinstance(x, str) for x in process_cmd):
        raise SystemExit("config.emon.process_cmd must be a string list")
    expected_output = str(emon.get("expected_output") or "summary.xlsx")

    # Optional: multi-host dispatch via SSH (run the sweep on each server and copy back).
    ssh = run.get("ssh") or {}
    if not isinstance(ssh, dict):
        ssh = {}
    ssh_user_default = str(ssh.get("user") or ssh.get("username") or "").strip()

    ssh_port_default: Optional[int] = None
    ssh_port_raw = ssh.get("port")
    if ssh_port_raw is not None and str(ssh_port_raw).strip() != "":
        try:
            ssh_port_default = int(ssh_port_raw)
        except Exception:
            ssh_port_default = None

    ssh_identity_file_default = str(ssh.get("identity_file") or "").strip()
    ssh_options_default = _as_str_list(ssh.get("options"))

    password_file_default_raw = str(ssh.get("password_file") or run.get("password_file") or "").strip()
    password_file_default: Optional[Path] = None
    if password_file_default_raw:
        password_file_default = Path(password_file_default_raw)
        if not password_file_default.is_absolute():
            password_file_default = (REPO_ROOT / password_file_default).resolve()
    else:
        password_file_default = (REPO_ROOT / "scripts/scale-test/embedding/passwords.json").resolve()

    _password_file_cache: Dict[str, Dict[str, Any]] = {}

    def _load_password_file(path: Path) -> Dict[str, Any]:
        key = str(path)
        if key in _password_file_cache:
            return _password_file_cache[key]
        if not path.exists():
            _password_file_cache[key] = {}
            return _password_file_cache[key]
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            raise SystemExit(f"failed to read password_file: {path} ({exc})")
        if not isinstance(data, dict):
            raise SystemExit(f"password_file must be a JSON object: {path}")
        _password_file_cache[key] = data
        return data

    def _resolve_password_from_db(db: Dict[str, Any], *, host: str, user: str) -> Optional[str]:
        h = str(host or "").strip()
        u = str(user or "").strip()

        hosts = db.get("hosts")
        if isinstance(hosts, dict) and h:
            hv = hosts.get(h)
            if isinstance(hv, dict):
                if u and u in hv:
                    v = hv.get(u)
                    return str(v) if v is not None else None
            elif hv is not None:
                return str(hv)

        users = db.get("users")
        if isinstance(users, dict) and u in users:
            v = users.get(u)
            return str(v) if v is not None else None

        if u and u in db:
            v = db.get(u)
            return str(v) if v is not None else None
        if h and h in db:
            hv = db.get(h)
            if isinstance(hv, dict):
                if u and u in hv:
                    v = hv.get(u)
                    return str(v) if v is not None else None
            elif hv is not None:
                return str(hv)
        return None

    def _password_from_file(host: str, user: str, password_file_raw: Optional[str]) -> str:
        pf = password_file_default
        if password_file_raw:
            pf = Path(str(password_file_raw))
            if not pf.is_absolute():
                pf = (REPO_ROOT / pf).resolve()
        if pf is None:
            raise SystemExit("password is null but no password_file is configured")
        db = _load_password_file(pf)
        if not db:
            raise SystemExit(f"password is null but password_file is missing/empty: {pf}")
        pwd = _resolve_password_from_db(db, host=host, user=user)
        if not pwd:
            raise SystemExit(f"password is null but no entry for user='{user}' host='{host}' in password_file: {pf}")
        return pwd

    remote_repo_dir_default = str(run.get("remote_repo_dir") or str(REPO_ROOT)).strip()
    remote_result_root_default = str(run.get("remote_result_root") or "").strip() or None

    install_requirements_default = bool(run.get("install_requirements") or False)
    requirements_profile_default = str(run.get("requirements_profile") or run.get("profile") or "").strip().lower()
    requirements_files_default = _as_str_list(run.get("requirements_files"))
    pip_extra_args_default = _as_str_list(run.get("pip_extra_args"))
    pre_requirements_file_default = str(run.get("pre_requirements_file") or "").strip()

    def _parse_pre_requirements_from_obj(obj: Any) -> List[PreRequirement]:
        out: List[PreRequirement] = []
        if obj is None:
            return out
        if isinstance(obj, list):
            for it in obj:
                if isinstance(it, str):
                    s = it.strip()
                    if not s:
                        continue
                    out.append(PreRequirement(file=s, stage="before_conda", mode="bash", source_bashrc_after=False))
                elif isinstance(it, dict):
                    f = str(it.get("file") or it.get("path") or "").strip()
                    if not f:
                        continue
                    stage = str(it.get("stage") or "before_conda").strip().lower()
                    if stage not in {"before_conda", "after_conda"}:
                        stage = "before_conda"
                    mode = str(it.get("mode") or "bash").strip().lower()
                    if mode not in {"bash", "source"}:
                        mode = "bash"
                    sb = bool(it.get("source_bashrc_after") or False)
                    out.append(PreRequirement(file=f, stage=stage, mode=mode, source_bashrc_after=sb))
        return out

    def _parse_pre_requirements_legacy(item: Dict[str, Any]) -> List[PreRequirement]:
        # Legacy:
        # - pre_requirements_file
        # - pre_requirements_use_conda_env
        f = str(item.get("pre_requirements_file") or item.get("pre_requirements") or "").strip()
        if not f:
            return []
        use_conda = bool(item.get("pre_requirements_use_conda_env") or False)
        if use_conda:
            return [PreRequirement(file=f, stage="after_conda", mode="bash", source_bashrc_after=False)]
        # Preserve historical behavior (source) for before-conda.
        return [PreRequirement(file=f, stage="before_conda", mode="source", source_bashrc_after=False)]

    def _remote_python_cmd(remote_python_raw: Any, conda_env: str) -> str:
        if isinstance(remote_python_raw, list):
            parts = [str(x).strip() for x in remote_python_raw if str(x).strip()]
            return " ".join(shlex.quote(p) for p in parts) if parts else "python3"
        s = str(remote_python_raw or "").strip()
        if s:
            return s
        if str(conda_env or "").strip():
            parts = ["conda", "run", "-n", str(conda_env).strip(), "python"]
            return " ".join(shlex.quote(p) for p in parts)
        return "python3"

    remote_python_default = _remote_python_cmd(run.get("remote_python"), conda_env="")

    servers_raw = run.get("servers")
    server_specs: List[ServerConfig] = []
    if isinstance(servers_raw, list):
        _PASSWORD_MISSING = object()
        for item in servers_raw:
            if isinstance(item, str):
                host = str(item).strip()
                if not host:
                    continue
                server_specs.append(
                    ServerConfig(
                        host=host,
                        user=ssh_user_default,
                        password="",
                        port=ssh_port_default,
                        identity_file=ssh_identity_file_default,
                        ssh_options=list(ssh_options_default),
                        remote_repo_dir=remote_repo_dir_default,
                        remote_result_root=remote_result_root_default,
                        remote_python=remote_python_default,
                        conda_env="",
                        install_requirements=install_requirements_default,
                        requirements_profile=requirements_profile_default,
                        requirements_files=list(requirements_files_default),
                        pip_extra_args=list(pip_extra_args_default),
                        pre_setup_cmds=[],
                        pre_requirements=(
                            [PreRequirement(file=pre_requirements_file_default, stage="before_conda", mode="source", source_bashrc_after=False)]
                            if pre_requirements_file_default
                            else []
                        ),
                    )
                )
            elif isinstance(item, dict):
                host = str(item.get("ip") or item.get("host") or "").strip()
                if not host:
                    continue
                is_local = _is_local_host(host)
                user = str(item.get("username") or item.get("user") or ssh_user_default or "").strip()
                raw_password = item.get("password", _PASSWORD_MISSING)
                if raw_password is None and "password" in item:
                    # Only resolve passwords when SSH dispatch is enabled.
                    # When running locally (e.g. on remote workers with --no-ssh-dispatch),
                    # configs may still contain password=null for dispatch, but the
                    # password file is intentionally not present.
                    if resolve_ssh_passwords and (not is_local):
                        password = _password_from_file(host, user, item.get("password_file"))
                    else:
                        password = ""
                elif raw_password is _PASSWORD_MISSING:
                    password = ""
                else:
                    password = str(raw_password or "").strip()

                port: Optional[int] = ssh_port_default
                port_raw = item.get("port")
                if port_raw is not None and str(port_raw).strip() != "":
                    try:
                        port = int(port_raw)
                    except Exception:
                        port = ssh_port_default

                identity_file = str(item.get("identity_file") or ssh_identity_file_default or "").strip()
                ssh_options = _as_str_list(item.get("ssh_options"))
                if not ssh_options:
                    ssh_options = list(ssh_options_default)

                remote_repo_dir = str(item.get("remote_repo_dir") or remote_repo_dir_default or str(REPO_ROOT)).strip()
                remote_result_root = str(item.get("remote_result_root") or "").strip() or remote_result_root_default

                conda_env = str(item.get("conda_env") or item.get("conda") or "").strip()
                remote_python = _remote_python_cmd(item.get("remote_python", run.get("remote_python")), conda_env=conda_env)

                install_requirements = bool(
                    item.get("install_requirements")
                    if ("install_requirements" in item)
                    else install_requirements_default
                )
                requirements_profile = str(
                    item.get("requirements_profile")
                    or item.get("profile")
                    or item.get("kind")
                    or requirements_profile_default
                    or ""
                ).strip().lower()

                requirements_files = _as_str_list(item.get("requirements_files"))
                if not requirements_files:
                    requirements_files = list(requirements_files_default)

                pip_extra_args = _as_str_list(item.get("pip_extra_args"))
                if not pip_extra_args:
                    pip_extra_args = list(pip_extra_args_default)

                pre_setup_cmds = _as_str_list(item.get("pre_setup_cmds") or item.get("setup_cmds"))

                pre_reqs: List[PreRequirement] = []
                # New format: explicit list
                pre_reqs += _parse_pre_requirements_from_obj(item.get("pre_requirements"))
                # New format: split lists
                for f in _as_str_list(item.get("pre_requirements_before_conda")):
                    pre_reqs.append(PreRequirement(file=f, stage="before_conda", mode="bash", source_bashrc_after=False))
                for f in _as_str_list(item.get("pre_requirements_after_conda")):
                    pre_reqs.append(PreRequirement(file=f, stage="after_conda", mode="bash", source_bashrc_after=False))
                # Legacy fallback
                if not pre_reqs:
                    pre_reqs += _parse_pre_requirements_legacy(item)
                # Global default legacy fallback
                if not pre_reqs and pre_requirements_file_default:
                    pre_reqs.append(
                        PreRequirement(file=pre_requirements_file_default, stage="before_conda", mode="source", source_bashrc_after=False)
                    )

                server_specs.append(
                    ServerConfig(
                        host=host,
                        user=user,
                        password=password,
                        port=port,
                        identity_file=identity_file,
                        ssh_options=list(ssh_options),
                        remote_repo_dir=remote_repo_dir,
                        remote_result_root=remote_result_root,
                        remote_python=remote_python,
                        conda_env=conda_env,
                        install_requirements=install_requirements,
                        requirements_profile=requirements_profile,
                        requirements_files=list(requirements_files),
                        pip_extra_args=list(pip_extra_args),
                        pre_setup_cmds=list(pre_setup_cmds),
                        pre_requirements=list(pre_reqs),
                    )
                )

    dispatch = SSHDispatchConfig(servers=server_specs)

    return ScaleConfig(
        template_auto_test_config=template,
        result_root=result_root,
        token_lens=token_lens,
        batch_sizes=batch_sizes,
        repeats=repeats,
        repeat_threshold_sec=repeat_threshold_sec,
        repeat_max_repeats=repeat_max_repeats,
        warmup_runs=warmup_runs,
        continue_on_error=continue_on_error,
        cpu_exprs=cpu_exprs,
        mem_gb=mem_gb_f,
        sglang_max_total_tokens=sglang_mtt,
        job_template=job_template,
        emon_process_after_run=emon_process_after_run,
        emon_process_cmd=[str(x) for x in process_cmd],
        emon_expected_output=expected_output,
        dispatch=dispatch,
    )


def _ssh_host_target(host: str, user: str) -> str:
    h = str(host or "").strip()
    if not h:
        return h
    if "@" in h or not user:
        return h
    return f"{user}@{h}"


def _ssh_base_args(server: ServerConfig) -> List[str]:
    # Allow interactive prompting when password is not provided.
    # This also allows prompting for key passphrases when needed.
    args: List[str] = ["-o", "BatchMode=no", "-o", "StrictHostKeyChecking=accept-new"]
    if server.port:
        args += ["-p", str(int(server.port))]
    if server.identity_file:
        args += ["-i", server.identity_file]
    for opt in server.ssh_options or []:
        s = str(opt).strip()
        if not s:
            continue
        if s.startswith("-"):
            args += s.split()
        else:
            args += ["-o", s]
    return args


def _scp_base_args(server: ServerConfig) -> List[str]:
    # scp uses -P for port (ssh uses -p). It also supports -o options.
    args: List[str] = ["-o", "BatchMode=no", "-o", "StrictHostKeyChecking=accept-new"]
    if server.port:
        args += ["-P", str(int(server.port))]
    if server.identity_file:
        args += ["-i", server.identity_file]
    for opt in server.ssh_options or []:
        s = str(opt).strip()
        if not s:
            continue
        if s.startswith("-"):
            # Best-effort: only keep '-o ...' style flags.
            parts = s.split()
            if parts and parts[0] == "-o" and len(parts) >= 2:
                args += ["-o", " ".join(parts[1:])]
        else:
            args += ["-o", s]
    return args


def _rsync_ssh_command(server: ServerConfig) -> str:
    parts: List[str] = []
    if server.password:
        parts += ["sshpass", "-p", server.password]
    parts += ["ssh"] + _ssh_base_args(server)
    return " ".join(shlex.quote(p) for p in parts)


def _log_meta_header(*, cmd_for_log: str, rc: int) -> str:
    ts = dt.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    return f"[meta] utc={ts} rc={int(rc)}\n[meta] cmd={cmd_for_log}\n"


def _maybe_sshpass_prefix(server: ServerConfig) -> List[str]:
    if not server.password:
        return []
    return ["sshpass", "-p", server.password]


def _ssh_bash_lc_remote(cmd: str) -> str:
    """Build a remote command string for ssh that safely runs a bash -lc command.

    Note: ssh does not preserve argv boundaries for the remote side; it concatenates
    the provided arguments into a single command string. Therefore, we must pass
    a *single* remote command string that properly quotes the inner command.
    """

    return f"bash -lc {shlex.quote(str(cmd))}"


def _dispatch_subprocess_env() -> Dict[str, str]:
    """Return an env suitable for calling system ssh/scp/rsync.

    Users often run this tool from inside a conda env that sets LD_LIBRARY_PATH
    to conda-provided OpenSSL libs. That can break system binaries like ssh/scp
    with errors like "OpenSSL version mismatch".

    We only sanitize env for the dispatch subprocesses (ssh/scp/rsync). The
    actual benchmark/python execution is remote and unaffected.
    """

    env = dict(os.environ)
    # Common offenders.
    for k in [
        "LD_LIBRARY_PATH",
        "DYLD_LIBRARY_PATH",
        "CONDA_PREFIX",
        "CONDA_DEFAULT_ENV",
        "CONDA_SHLVL",
        "CONDA_PROMPT_MODIFIER",
        "_CONDA_EXE",
        "_CE_CONDA",
        "_CE_M",
    ]:
        env.pop(k, None)
    return env


def _maybe_preflight_socks_proxy(
    *,
    server: ServerConfig,
    local_host_dir: Path,
    dispatch_env: Dict[str, str],
) -> bool:
    """If ProxyCommand uses nc + SOCKS5, test reachability early.

    When the SOCKS proxy cannot route to the target, ssh often surfaces this as
    a confusing "banner exchange" timeout and an UNKNOWN port (e.g. 65535).
    We run a cheap nc preflight so failures are obvious and logged.
    """

    proxy_cmd: Optional[str] = None
    for opt in server.ssh_options or []:
        if str(opt).startswith("ProxyCommand="):
            proxy_cmd = str(opt).split("=", 1)[1].strip()
            break
    if not proxy_cmd:
        return True

    # Only handle the known pattern we generate in configs:
    #   ProxyCommand=nc -x host:port -X 5 %h %p
    try:
        argv = shlex.split(proxy_cmd)
    except Exception:
        argv = []
    if not argv:
        return True
    if os.path.basename(argv[0]) != "nc":
        return True

    proxy_addr: Optional[str] = None
    proxy_proto: Optional[str] = None
    for i, a in enumerate(argv):
        if a == "-x" and i + 1 < len(argv):
            proxy_addr = argv[i + 1]
        if a == "-X" and i + 1 < len(argv):
            proxy_proto = argv[i + 1]
    if not proxy_addr:
        return True
    if proxy_proto not in {"5", "socks5"}:
        return True

    nc_bin = _which("nc")
    if not nc_bin:
        return True

    out_path = local_host_dir / "socks_preflight.log"
    cmd = [
        str(nc_bin),
        "-vz",
        "-w",
        "8",
        "-x",
        str(proxy_addr),
        "-X",
        "5",
        str(server.host),
        str(int(server.port)),
    ]
    p = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        check=False,
        env=dispatch_env,
    )
    out_path.write_text(
        _log_meta_header(cmd_for_log=" ".join(shlex.quote(x) for x in cmd), rc=int(p.returncode)) + (p.stdout or ""),
        encoding="utf-8",
    )
    return p.returncode == 0


def _requirements_profile_default_file(profile: str) -> str:
    p = str(profile or "").strip().lower()
    if p in {"cpu", "host", "x86"}:
        return "requirements-cpu.txt"
    if p in {"cuda", "gpu"}:
        return "requirements-cuda.txt"
    return ""


def _port_from_base_url(job_template: Dict[str, Any]) -> Optional[int]:
    """Extract port from job_template base_url/openai_base_url, best-effort."""

    if not isinstance(job_template, dict):
        return None
    raw = (
        job_template.get("base_url")
        or job_template.get("openai_base_url")
        or job_template.get("api_base")
        or ""
    )
    s = str(raw or "").strip()
    if not s:
        return None
    try:
        u = urlparse(s)
        if u.port:
            return int(u.port)
    except Exception:
        pass
    # Fallback: handle forms like "127.0.0.1:30000".
    try:
        if ":" in s and "//" not in s:
            host, port_s = s.rsplit(":", 1)
            if host and port_s.isdigit():
                return int(port_s)
    except Exception:
        pass
    return None


def _sha256_hex(data: bytes) -> str:
    h = hashlib.sha256()
    h.update(data)
    return h.hexdigest()


def _resolve_repo_path(p: str) -> Path:
    pp = Path(str(p)).expanduser()
    if pp.is_absolute():
        return pp
    return (REPO_ROOT / pp).resolve()


def _compute_requirements_marker_key(*, server: ServerConfig, reqs: List[str]) -> str:
    """Stable key for caching pip installs across remote runs."""

    parts: List[str] = []
    parts.append(f"remote_python={server.remote_python}")
    parts.append(f"conda_env={server.conda_env}")
    if server.pip_extra_args:
        parts.append("pip_extra_args=" + " ".join(server.pip_extra_args))
    for rf in reqs:
        rfs = str(rf).strip()
        if not rfs:
            continue
        digest = ""
        try:
            local_p = _resolve_repo_path(rfs)
            if local_p.exists() and local_p.is_file():
                digest = _sha256_hex(local_p.read_bytes())
        except Exception:
            digest = ""
        parts.append(f"req={rfs}#{digest}")
    return _sha256_hex("\n".join(parts).encode("utf-8", errors="replace"))


def _rewrite_remote_path_to_local(*, val: str, remote_run_dir: str, local_host_dir: Path) -> str:
    s = str(val or "").strip()
    if not s:
        return ""
    rr = str(remote_run_dir).rstrip("/")
    if s.startswith(rr + "/"):
        rel = s[len(rr) + 1 :]
        return str((local_host_dir / rel).resolve())
    return s


def _read_csv_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        r = csv.DictReader(f)
        out: List[Dict[str, str]] = []
        for row in r:
            out.append({k: ("" if v is None else str(v)) for k, v in row.items()})
        return out


def _capture_local_lscpu(*, out_dir: Path) -> None:
    info_dir = (out_dir / "server_info").resolve()
    info_dir.mkdir(parents=True, exist_ok=True)
    try:
        p = subprocess.run(
            ["lscpu"],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            check=False,
            timeout=30,
        )
        (info_dir / "lscpu.txt").write_text(p.stdout or "", encoding="utf-8")
        (info_dir / "lscpu.rc").write_text(str(int(p.returncode)) + "\n", encoding="utf-8")
    except Exception as e:
        (info_dir / "lscpu.error.txt").write_text(str(e) + "\n", encoding="utf-8")

    # Capture host memory capacity (Linux): /proc/meminfo contains MemTotal.
    try:
        meminfo = Path("/proc/meminfo")
        if meminfo.exists() and meminfo.is_file():
            (info_dir / "meminfo.txt").write_text(meminfo.read_text(encoding="utf-8", errors="replace"), encoding="utf-8")
    except Exception as e:
        (info_dir / "meminfo.error.txt").write_text(str(e) + "\n", encoding="utf-8")


def _capture_remote_lscpu(*, local_host_dir: Path, server: ServerConfig, dispatch_env: Optional[Dict[str, str]] = None) -> None:
    info_dir = (local_host_dir / "server_info").resolve()
    info_dir.mkdir(parents=True, exist_ok=True)
    host = str(server.host or "")
    try:
        (local_host_dir / "server_host.txt").write_text(host + "\n", encoding="utf-8")
    except Exception:
        pass

    target = _ssh_host_target(server.host, server.user)
    ssh_args = _ssh_base_args(server)
    prefix = _maybe_sshpass_prefix(server)

    def run_capture(cmd: str, out_name: str) -> None:
        try:
            p = subprocess.run(
                prefix + ["ssh"] + ssh_args + [target, _ssh_bash_lc_remote(cmd)],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                check=False,
                timeout=30,
                env=dispatch_env,
            )
            (info_dir / out_name).write_text(p.stdout or "", encoding="utf-8")
            (info_dir / (out_name + ".rc")).write_text(str(int(p.returncode)) + "\n", encoding="utf-8")
        except Exception as e:
            (info_dir / (out_name + ".error.txt")).write_text(str(e) + "\n", encoding="utf-8")

    run_capture("lscpu", "lscpu.txt")
    # JSON output is useful when present; ignore failures on older util-linux.
    run_capture("lscpu -J", "lscpu.json")
    # Memory capacity (Linux): this includes MemTotal in kB.
    run_capture("cat /proc/meminfo", "meminfo.txt")


def _build_repo_bundle(*, out_path: Path) -> None:
    """Create a gzipped tarball of the current repo for remote bootstrap."""

    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.exists() and out_path.is_file():
        try:
            if out_path.stat().st_size > 0:
                return
        except Exception:
            pass

    tar_bin = _which("tar")
    if not tar_bin:
        raise SystemExit("tar is required to bundle the repo for remote bootstrap")

    # Prefer respecting .gitignore (and other exclude rules) via git.
    git_bin = _which("git")
    if git_bin and (REPO_ROOT / ".git").exists():
        try:
            ls = subprocess.run(
                [
                    git_bin,
                    "-C",
                    str(REPO_ROOT),
                    "ls-files",
                    "-z",
                    "--cached",
                    "--others",
                    "--exclude-standard",
                ],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                check=False,
            )
            if ls.returncode == 0 and ls.stdout:
                # tar reads NUL-separated names from stdin.
                cmd: List[str] = [
                    tar_bin,
                    "-C",
                    str(REPO_ROOT),
                    "-czf",
                    str(out_path),
                    "--null",
                    "--files-from=-",
                ]
                p = subprocess.run(cmd, input=ls.stdout, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, check=False)
                if p.returncode != 0:
                    msg = (p.stdout or b"").decode("utf-8", errors="replace")
                    raise SystemExit(f"Failed to create repo bundle (tar rc={p.returncode}): {msg}")
                return
        except Exception:
            # Fall back to tar excludes below.
            pass

    # Fallback: keep the bundle reasonably small with common excludes.
    excludes = [
        ".git",
        ".venv",
        "**/.venv",
        "__pycache__",
        "**/__pycache__",
        "*.pyc",
        ".mypy_cache",
        ".pytest_cache",
        ".ruff_cache",
        "scripts/scale-test/**/result",
        "scripts/auto-test/**/result",
        "scripts/**/results",
    ]

    cmd2: List[str] = [tar_bin, "-czf", str(out_path)]
    for ex in excludes:
        cmd2.append(f"--exclude={ex}")
    cmd2 += ["-C", str(REPO_ROOT), "."]

    p2 = subprocess.run(cmd2, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, check=False)
    if p2.returncode != 0:
        raise SystemExit(f"Failed to create repo bundle (rc={p2.returncode}): {p2.stdout}")


def _ensure_remote_repo_dir(
    *,
    server: ServerConfig,
    target: str,
    local_host_dir: Path,
    remote_cfg_dir: str,
    remote_repo_dir: str,
    scp_bin: str,
    dispatch_env: Dict[str, str],
    continue_on_error: bool,
    repo_bundle_path: Path,
) -> bool:
    """Ensure remote_repo_dir exists; if missing, upload+extract current repo."""

    ssh_args = _ssh_base_args(server)
    scp_args = _scp_base_args(server)
    prefix = _maybe_sshpass_prefix(server)

    check_cmd = f"test -d {shlex.quote(remote_repo_dir)}"
    p = subprocess.run(
        prefix + ["ssh"] + ssh_args + [target, _ssh_bash_lc_remote(check_cmd)],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        check=False,
        env=dispatch_env,
    )
    (local_host_dir / "remote_repo_check.log").write_text(p.stdout or "", encoding="utf-8")
    if p.returncode == 0:
        return True

    print(f"[warn] remote_repo_dir missing; bootstrapping repo (host={server.host})")
    try:
        _build_repo_bundle(out_path=repo_bundle_path)
    except Exception as e:
        (local_host_dir / "repo_bundle_build.error.txt").write_text(str(e) + "\n", encoding="utf-8")
        return False if continue_on_error else (_ for _ in ()).throw(e)

    remote_bundle = f"{remote_cfg_dir.rstrip('/')}/{repo_bundle_path.name}"
    scp_cmd = prefix + [str(scp_bin)] + scp_args + [str(repo_bundle_path), f"{target}:{remote_bundle}"]
    p2 = subprocess.run(
        scp_cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        check=False,
        env=dispatch_env,
    )
    (local_host_dir / "scp_repo_bundle.log").write_text(p2.stdout or "", encoding="utf-8")
    if p2.returncode != 0:
        if continue_on_error:
            return False
        raise SystemExit(f"scp repo bundle failed (host={server.host}, rc={p2.returncode})")

    extract_cmd = (
        f"mkdir -p {shlex.quote(remote_repo_dir)} && "
        f"tar -xzf {shlex.quote(remote_bundle)} -C {shlex.quote(remote_repo_dir)}"
    )
    p3 = subprocess.run(
        prefix + ["ssh"] + ssh_args + [target, _ssh_bash_lc_remote(extract_cmd)],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        check=False,
        env=dispatch_env,
    )
    (local_host_dir / "remote_repo_extract.log").write_text(p3.stdout or "", encoding="utf-8")
    if p3.returncode != 0:
        if continue_on_error:
            return False
        raise SystemExit(f"remote repo extract failed (host={server.host}, rc={p3.returncode})")

    return True


def _dispatch_multi_host(
    *,
    cfg_path: Path,
    scale: ScaleConfig,
    result_root: Path,
    scale_id: str,
    tee: bool,
    dry_run: bool,
    remote_clean_repo: bool,
    resume: bool,
) -> int:
    servers = [s for s in (scale.dispatch.servers or []) if str(getattr(s, "host", "")).strip()]
    if not servers:
        raise SystemExit("config.run.servers is empty")

    # Local hosts (127.0.0.1 / localhost / on-box IPs) can run without SSH/rsync/scp.
    remote_servers = [s for s in servers if not _is_local_host(str(getattr(s, "host", "")))]
    has_remote = bool(remote_servers)

    out_dir = (result_root / scale_id).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    hosts_dir = out_dir / "hosts"
    hosts_dir.mkdir(parents=True, exist_ok=True)

    # Save a copy of the config used for dispatch.
    try:
        shutil.copy2(str(cfg_path), str(out_dir / "scale_config.used.json"))
    except Exception:
        pass

    rsync_bin = _which("rsync")
    scp_bin = _which("scp")
    sshpass_bin = _which("sshpass")
    if has_remote:
        if not scp_bin and not dry_run:
            raise SystemExit("scp is required for multi-host dispatch")
        if not rsync_bin and not dry_run:
            raise SystemExit("rsync is required for multi-host dispatch (install rsync)")

    all_rows: List[Dict[str, Any]] = []
    dispatch_env = _dispatch_subprocess_env()
    repo_bundle_path = (out_dir / "repo_bundle.tar.gz").resolve()

    def _effective_dispatch_rc(*, ssh_rc: int, remote_run_log: Path) -> Tuple[int, Optional[int]]:
        """Reconcile SSH transport rc with remote sweep rc.

        The remote dispatch script prints a line like:
          [dispatch] sweep_rc=0

        The SSH command may still return non-zero due to post-sweep debug
        collection steps (or other shell/transport quirks). In that case, the
        sweep rc is the authoritative signal of success/failure.
        """

        sweep_rc = _parse_dispatch_sweep_rc(remote_run_log)
        if sweep_rc is None:
            return int(ssh_rc), None
        return int(sweep_rc), int(sweep_rc)

    def _run_stream_to_log(cmd: List[str], *, log_path: Path, echo: bool) -> int:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("w", encoding="utf-8", errors="replace") as f:
            p = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1,
                env=dispatch_env,
            )
            assert p.stdout is not None
            try:
                for line in p.stdout:
                    f.write(line)
                    f.flush()
                    if echo:
                        # Stream as-is (already contains newline)
                        sys.stdout.write(line)
                        sys.stdout.flush()
            finally:
                try:
                    p.stdout.close()
                except Exception:
                    pass
            return int(p.wait())

    prestage_cache_dir = (REPO_ROOT / "scripts/scale-test/embedding/.cache/prestage").resolve()

    def _ensure_prestaged_sglang_tarball(*, log_dir: Path, echo: bool = False) -> Optional[Path]:
        """Best-effort local clone of sglang and tarball creation for remote prestage."""

        prestage_cache_dir.mkdir(parents=True, exist_ok=True)
        clone_dir = prestage_cache_dir / "sglang"
        tar_path = prestage_cache_dir / "sglang.tar.gz"
        git_bin = _which("git")
        log_lines: List[str] = []

        if not git_bin:
            (log_dir / "prestage_sglang.log").write_text("[error] local git not found\n", encoding="utf-8")
            return None

        def _run_git(args: List[str], *, cwd: Optional[Path] = None) -> bool:
            cmd = [str(git_bin), *args]
            log_lines.append("$ " + " ".join(shlex.quote(a) for a in cmd))
            try:
                p = subprocess.run(
                    cmd,
                    cwd=str(cwd or prestage_cache_dir),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    check=False,
                    env=dispatch_env,
                )
            except Exception as exc:
                log_lines.append(f"[error] git invocation failed: {type(exc).__name__}: {exc}")
                return False
            if p.stdout:
                log_lines.append(p.stdout)
                if echo:
                    sys.stdout.write(p.stdout)
                    sys.stdout.flush()
            return p.returncode == 0

        # Ensure a usable clone.
        if (clone_dir / ".git").is_dir():
            # Best-effort refresh; do not fail prestage if fetch/reset fails.
            _run_git(["-C", str(clone_dir), "fetch", "--all", "--prune"])
            _run_git(["-C", str(clone_dir), "reset", "--hard", "origin/main"])
        else:
            if clone_dir.exists():
                try:
                    shutil.rmtree(clone_dir)
                except Exception:
                    pass
            ok = _run_git(["clone", "--depth", "1", "https://github.com/sgl-project/sglang.git", str(clone_dir)])
            if not ok:
                (log_dir / "prestage_sglang.log").write_text("\n".join(log_lines) + "\n", encoding="utf-8")
                return None

        # Build tarball (include .git so remote script sees a git repo).
        try:
            tmp_tar = tar_path.with_suffix(tar_path.suffix + ".tmp")
            if tmp_tar.exists():
                try:
                    tmp_tar.unlink()
                except Exception:
                    pass
            with tarfile.open(tmp_tar, "w:gz") as tf:
                tf.add(str(clone_dir), arcname="sglang")
            tmp_tar.replace(tar_path)
        except Exception as exc:
            log_lines.append(f"[error] failed to create tarball: {type(exc).__name__}: {exc}")
            (log_dir / "prestage_sglang.log").write_text("\n".join(log_lines) + "\n", encoding="utf-8")
            return None

        (log_dir / "prestage_sglang.log").write_text("\n".join(log_lines) + "\n", encoding="utf-8")
        return tar_path

    def _remote_log_needs_sglang_prestage(log_text: str) -> bool:
        t = (log_text or "").lower()
        return (
            "failed to clone sglang after retries" in t
            or ("git clone failed" in t and "cloning into 'sglang'" in t)
            or ("sglang" in t and "git" in t and "clone" in t and "failed" in t)
        )

    for server in servers:
        host = str(server.host)
        is_local = _is_local_host(host)
        target = _ssh_host_target(server.host, server.user)
        host_tag = _safe_name(host.replace(":", "_"))
        local_host_dir = (hosts_dir / host_tag).resolve()
        local_host_dir.mkdir(parents=True, exist_ok=True)

        ssh_args = _ssh_base_args(server)
        scp_args = _scp_base_args(server)
        prefix = _maybe_sshpass_prefix(server)
        if (not is_local) and server.password and not sshpass_bin and not dry_run:
            raise SystemExit("sshpass is required when config.run.servers[*].password is set (install sshpass or use SSH keys)")

        # Pre-flight connectivity check for clearer errors in restricted networks.
        # If a proxy is configured (ProxyCommand/ProxyJump), a direct TCP connect
        # to host:port may be expected to fail, so skip the preflight.
        uses_proxy = False
        for opt in (server.ssh_options or []):
            s = str(opt or "").strip()
            if not s:
                continue
            if "proxycommand" in s.lower() or "proxyjump" in s.lower() or s.startswith("-J"):
                uses_proxy = True
                break

        if (not is_local) and (not uses_proxy):
            connect_port = int(server.port or 22)
            try:
                sock = socket.create_connection((host, connect_port), timeout=6.0)
                try:
                    sock.close()
                except Exception:
                    pass
            except Exception as exc:
                (local_host_dir / "tcp_connect.log").write_text(
                    f"host={host} port={connect_port} error={type(exc).__name__}: {exc}\n",
                    encoding="utf-8",
                )
                raise SystemExit(
                    f"cannot reach SSH target (host={host}, port={connect_port}). "
                    "Check security group/firewall/VPN routing, or set run.ssh.options/servers[*].ssh_options (e.g. ProxyJump/ProxyCommand)."
                )

        remote_cfg_dir = f"/tmp/scale_test_dispatch/{scale_id}"
        remote_cfg_path = f"{remote_cfg_dir}/{cfg_path.name}"
        remote_repo_dir = server.remote_repo_dir
        remote_result_root = server.remote_result_root or str(result_root)
        if is_local:
            # Avoid pathological cases where the run directory contains the host dir (recursion).
            # Also keep local execution isolated per-host.
            remote_result_root = str((result_root / "_local_remote_result_root" / host_tag).resolve())
        remote_run_dir = f"{remote_result_root.rstrip('/')}/{scale_id}"

        # Resume shortcut: if we already have a successful local copy for this host,
        # skip remote execution and reuse the existing host aggregate.
        if resume and not dry_run:
            local_host_agg = local_host_dir / "aggregate.csv"
            sweep_rc = _parse_dispatch_sweep_rc(local_host_dir / "remote_run.log")
            if local_host_agg.exists() and sweep_rc == 0:
                if tee:
                    print(f"[resume] host already complete; skipping remote run (host={host})")
                host_rows = _read_csv_rows(local_host_agg)
                for r in host_rows:
                    r2: Dict[str, Any] = dict(r)
                    r2["server_host"] = host
                    for k in ["log_path", "metrics_path", "emon_output_path", "emon_summary_xlsx", "emon_process_log"]:
                        if k in r2:
                            r2[k] = _rewrite_remote_path_to_local(
                                val=str(r2.get(k, "")),
                                remote_run_dir=remote_run_dir,
                                local_host_dir=local_host_dir,
                            )
                    all_rows.append(r2)
                continue

        pre_req_items: List[Tuple[PreRequirement, Path, str]] = []
        for pr in (server.pre_requirements or []):
            f = str(pr.file or "").strip()
            if not f:
                continue
            p = Path(f).expanduser()
            if not p.is_absolute():
                p = (REPO_ROOT / p).resolve()
            remote_p = f"{remote_cfg_dir}/{p.name}"
            pre_req_items.append((pr, p, remote_p))

        before_cmds: List[str] = []
        after_cmds: List[str] = []
        for pr, _, remote_p in pre_req_items:
            cmd = f"bash {shlex.quote(remote_p)}" if pr.mode == "bash" else f". {shlex.quote(remote_p)}"
            if pr.source_bashrc_after:
                cmd = cmd + "; . ~/.bashrc >/dev/null 2>&1 || true"
            if pr.stage == "after_conda":
                after_cmds.append(cmd)
            else:
                before_cmds.append(cmd)

        pre_req_prefix = ""
        # Export job_template model hints so pre-requirement scripts (e.g. download-models.sh)
        # can download only what the current sweep actually uses.
        jt_model_dir = str(scale.job_template.get("model_dir") or "").strip()
        jt_model = str(scale.job_template.get("model") or "").strip()
        jt_model_id = str(
            scale.job_template.get("model_id")
            or scale.job_template.get("served_model_name")
            or scale.job_template.get("served_model")
            or ""
        ).strip()
        env_exports: List[str] = []
        if jt_model_dir:
            env_exports.append(f"export MMPE_JOB_TEMPLATE_MODEL_DIR={shlex.quote(jt_model_dir)}")
        if jt_model:
            env_exports.append(f"export MMPE_JOB_TEMPLATE_MODEL={shlex.quote(jt_model)}")
        if jt_model_id:
            env_exports.append(f"export MMPE_JOB_TEMPLATE_MODEL_ID={shlex.quote(jt_model_id)}")
        if env_exports:
            pre_req_prefix += " ".join(x + ";" for x in env_exports) + " "
        if before_cmds:
            pre_req_prefix += " ".join(c + ";" for c in before_cmds) + " "
        if after_cmds:
            if not str(server.conda_env or "").strip():
                raise SystemExit(f"pre_requirements has after_conda stage but conda_env is empty (host={host})")
            # Run in the same shell so exports can affect later steps.
            pre_req_prefix += (
                "eval \"$(conda shell.bash hook)\"; "
                + f"conda activate {shlex.quote(server.conda_env.strip())}; "
                + " ".join(c + ";" for c in after_cmds)
                + " "
            )

        setup_cmds: List[str] = []

        # Preflight: kill any stale listener on the configured base_url port *before*
        # running setup/pip installs. This avoids confusing "CPU is pegged" symptoms
        # from an old server process that is unrelated to the current run.
        base_port = _port_from_base_url(scale.job_template)
        if base_port:
            setup_cmds.append(
                "PORT="
                + shlex.quote(str(int(base_port)))
                + "; "
                + "if command -v lsof >/dev/null 2>&1; then "
                + "pids=$(lsof -t -iTCP:${PORT} -sTCP:LISTEN 2>/dev/null || true); "
                + "if [ -n \"$pids\" ]; then "
                + "echo \"[setup] killing stale listeners on :${PORT}: $pids\"; "
                + "kill -TERM $pids >/dev/null 2>&1 || true; "
                + "sleep 1; "
                + "kill -KILL $pids >/dev/null 2>&1 || true; "
                + "fi; "
                + "fi"
            )

        for s in (server.pre_setup_cmds or []):
            ss = str(s).strip()
            if ss:
                setup_cmds.append(ss)

        resume_pip_marker_file = ""
        if bool(server.install_requirements):
            reqs: List[str] = []
            base_req = _requirements_profile_default_file(server.requirements_profile)
            if base_req:
                reqs.append(base_req)
            for rf in (server.requirements_files or []):
                rfs = str(rf).strip()
                if rfs:
                    reqs.append(rfs)
            # De-dup while preserving order.
            seen: set[str] = set()
            reqs2: List[str] = []
            for rf in reqs:
                if rf in seen:
                    continue
                seen.add(rf)
                reqs2.append(rf)
            reqs = reqs2

            pip_args = " ".join(shlex.quote(a) for a in (server.pip_extra_args or []))
            marker_key = _compute_requirements_marker_key(server=server, reqs=reqs)
            # Do NOT use a quoted '~' path here: tilde expansion does not occur
            # inside quotes, which breaks the cache marker. Use $HOME instead.
            # (No spaces expected; keep unquoted for shell expansion.)
            marker_dir = "$HOME/.cache/multi-model-process-eval/pip_markers"
            marker_file = f"{marker_dir}/{marker_key}.ok"
            resume_pip_marker_file = marker_file

            # Install all requirements in one shot (less output), and cache via marker.
            # Marker is only written on success.
            req_flags = " ".join(f"-r {shlex.quote(rf)}" for rf in reqs)
            pip_base = (
                f"{server.remote_python} -m pip install --progress-bar off --disable-pip-version-check --no-input --root-user-action=ignore {pip_args}".strip()
            )
            setup_cmds.append(
                f"mkdir -p {marker_dir}; "
                + f"if [ -f {marker_file} ]; then "
                + f"echo \"[setup] pip requirements cached: {marker_key}\"; "
                + "else "
                + f"echo \"[setup] installing pip requirements (cache miss): {marker_key}\"; "
                + f"{pip_base} {req_flags}; "
                + f"date -u +%Y-%m-%dT%H:%M:%SZ > {marker_file}; "
                + "fi"
            )

        setup = " ".join(f"{cmd};" for cmd in (setup_cmds or [])).strip()

        # Remote lock + early cleanup:
        # - Avoid overlapping dispatches killing each other mid-run.
        # - Honor the user's "if something is already running, kill it first".
        # We do this BEFORE pre_requirements so a slow pre_requirements phase
        # in one dispatch won't kill an already-running newer dispatch.
        lock_dir = "/tmp/mmpe_scale_test_lock_fix_token_len"
        remote_repo_dir_norm = remote_repo_dir.rstrip("/")
        kill_auto_test_pat = f"{remote_repo_dir_norm}/scripts/auto-test/embedding/run_auto_test.py"
        kill_scale_pat = f"{remote_repo_dir_norm}/scripts/scale-test/embedding/run_scale_fix_token_len.py"
        # Use a bracket trick in pgrep patterns so the pgrep process won't match
        # itself. We still must exclude $$/$PPID because the remote `bash -lc`
        # command line contains the script paths.
        kill_auto_test_pat_pgrep = kill_auto_test_pat.replace("/run_auto_test.py", "/[r]un_auto_test.py")
        kill_scale_pat_pgrep = kill_scale_pat.replace("/run_scale_fix_token_len.py", "/[r]un_scale_fix_token_len.py")
        lock_and_cleanup = (
            f"LOCK_DIR={shlex.quote(lock_dir)}; "
            + "echo \"[dispatch] acquiring lock: $LOCK_DIR\"; "
            + "if mkdir \"$LOCK_DIR\" 2>/dev/null; then "
            + "echo $$ > \"$LOCK_DIR/pid\"; "
            + "else "
            + "echo '[dispatch] lock exists; killing stale processes and restarting'; "
            + "if command -v pgrep >/dev/null 2>&1; then "
            + f"pids=$(pgrep -f {shlex.quote(kill_auto_test_pat_pgrep)} 2>/dev/null || true); "
            + "for pid in $pids; do if [ \"$pid\" != \"$$\" ] && [ \"$pid\" != \"$PPID\" ]; then kill -TERM $pid >/dev/null 2>&1 || true; fi; done; "
            + f"pids=$(pgrep -f {shlex.quote(kill_scale_pat_pgrep)} 2>/dev/null || true); "
            + "for pid in $pids; do if [ \"$pid\" != \"$$\" ] && [ \"$pid\" != \"$PPID\" ]; then kill -TERM $pid >/dev/null 2>&1 || true; fi; done; "
            + "fi; "
            + "sleep 2; "
            + "rm -rf \"$LOCK_DIR\" >/dev/null 2>&1 || true; "
            + "mkdir \"$LOCK_DIR\" 2>/dev/null || { echo '[dispatch] failed to create lock dir'; exit 1; }; "
            + "echo $$ > \"$LOCK_DIR/pid\" || { echo '[dispatch] failed to write lock pid'; exit 1; }; "
            + "fi; "
            + "trap \"rm -rf \\\"$LOCK_DIR\\\" >/dev/null 2>&1 || true\" EXIT; "
            + "if command -v pgrep >/dev/null 2>&1; then "
            + f"pids=$(pgrep -f {shlex.quote(kill_auto_test_pat_pgrep)} 2>/dev/null || true); "
            + "for pid in $pids; do if [ \"$pid\" != \"$$\" ] && [ \"$pid\" != \"$PPID\" ]; then kill -TERM $pid >/dev/null 2>&1 || true; fi; done; "
            + f"pids=$(pgrep -f {shlex.quote(kill_scale_pat_pgrep)} 2>/dev/null || true); "
            + "for pid in $pids; do if [ \"$pid\" != \"$$\" ] && [ \"$pid\" != \"$PPID\" ]; then kill -TERM $pid >/dev/null 2>&1 || true; fi; done; "
            + "sleep 1; "
            + f"pids=$(pgrep -f {shlex.quote(kill_auto_test_pat_pgrep)} 2>/dev/null || true); "
            + "for pid in $pids; do if [ \"$pid\" != \"$$\" ] && [ \"$pid\" != \"$PPID\" ]; then kill -KILL $pid >/dev/null 2>&1 || true; fi; done; "
            + f"pids=$(pgrep -f {shlex.quote(kill_scale_pat_pgrep)} 2>/dev/null || true); "
            + "for pid in $pids; do if [ \"$pid\" != \"$$\" ] && [ \"$pid\" != \"$PPID\" ]; then kill -KILL $pid >/dev/null 2>&1 || true; fi; done; "
            + "fi; "
        )

        # Also kill a stale listener early (if any), so we don't later confuse
        # "server won't bind" errors with installation/setup steps.
        if base_port:
            lock_and_cleanup += (
                "PORT="
                + shlex.quote(str(int(base_port)))
                + "; "
                + "if command -v lsof >/dev/null 2>&1; then "
                + "pids=$(lsof -t -iTCP:${PORT} -sTCP:LISTEN 2>/dev/null || true); "
                + "if [ -n \"$pids\" ]; then "
                + "echo \"[dispatch] killing stale listeners on :${PORT}: $pids\"; "
                + "kill -TERM $pids >/dev/null 2>&1 || true; "
                + "sleep 1; "
                + "kill -KILL $pids >/dev/null 2>&1 || true; "
                + "fi; "
                + "fi; "
            )
        # Always create remote_run_dir early and collect debug artifacts (server logs, port/process state)
        # even when the remote sweep fails. This makes post-mortem analysis possible.
        remote_run_dir_quoted = shlex.quote(remote_run_dir)
        remote_repo_dir_quoted = shlex.quote(remote_repo_dir)
        debug_dir = f"{remote_run_dir.rstrip('/')}/dispatch_debug"
        debug_dir_q = shlex.quote(debug_dir)
        base_port_q = shlex.quote(str(int(base_port))) if base_port else ""

        # Resume behavior: in multi-host dispatch, users typically expect --resume
        # to continue the remote test quickly without re-running pre-requirements
        # and dependency installation if the environment is already prepared.
        #
        # Heuristic:
        # - If the pip requirements marker exists on the remote host, treat setup
        #   as already satisfied.
        # - Otherwise, if the remote run directory already exists and is non-empty
        #   (from a previous attempt), also skip setup.
        #
        # We also write a per-run setup marker under the run directory for clarity.
        remote_setup_marker = f"{remote_run_dir.rstrip('/')}/dispatch_debug/dispatch_setup.ok"
        remote_setup_marker_q = shlex.quote(remote_setup_marker)
        need_conda = bool(str(server.conda_env or "").strip()) or ("conda" in str(server.remote_python or "").lower())

        # This bootstrap is cheap and safe to run even when skipping setup.
        # It helps when conda exists but isn't on PATH in a non-interactive shell.
        conda_bootstrap = (
            "if ! command -v conda >/dev/null 2>&1; then "
            "  if [ -x \"$HOME/miniforge3/bin/conda\" ]; then export PATH=\"$HOME/miniforge3/bin:$PATH\"; fi; "
            "  if [ -f \"$HOME/miniforge3/etc/profile.d/conda.sh\" ]; then . \"$HOME/miniforge3/etc/profile.d/conda.sh\" >/dev/null 2>&1 || true; fi; "
            "fi; "
        )

        # Guard to decide whether to skip remote pre_requirements/setup.
        # Always define variables so non-resume runs behave correctly.
        setup_skip_guard = f"SETUP_MARKER={remote_setup_marker_q}; SKIP_SETUP=0; "
        if resume:
            if resume_pip_marker_file:
                setup_skip_guard += f"if [ -f {resume_pip_marker_file} ]; then SKIP_SETUP=1; fi; "
            setup_skip_guard += (
                "if [ \"$SKIP_SETUP\" = \"0\" ] && [ -d \"$REMOTE_RUN_DIR\" ]; then "
                "  if [ -n \"$(ls -A \"$REMOTE_RUN_DIR\" 2>/dev/null || true)\" ]; then SKIP_SETUP=1; fi; "
                "fi; "
            )
            if need_conda:
                setup_skip_guard += (
                    # If conda is truly missing, we cannot run the sweep; force setup.
                    conda_bootstrap
                    + "if [ \"$SKIP_SETUP\" = \"1\" ] && ! command -v conda >/dev/null 2>&1; then "
                    + "  echo '[dispatch] resume: conda not found; will run pre_requirements/setup'; SKIP_SETUP=0; "
                    + "fi; "
                )
            setup_skip_guard += (
                "if [ \"$SKIP_SETUP\" = \"1\" ]; then echo '[dispatch] resume: skipping pre_requirements/setup'; "
                "else echo '[dispatch] resume: running pre_requirements/setup'; fi; "
            )

        remote_cmd = (
            "set -e; "
            + lock_and_cleanup
            + f"REMOTE_RUN_DIR={remote_run_dir_quoted}; "
            + f"REMOTE_REPO_DIR={remote_repo_dir_quoted}; "
            + f"DEBUG_DIR={debug_dir_q}; "
            # Detect whether we can skip remote setup before creating the run dir.
            + setup_skip_guard
            + "mkdir -p \"$REMOTE_RUN_DIR\" \"$DEBUG_DIR\"; "
            + "echo \"[dispatch] remote_run_dir=$REMOTE_RUN_DIR\"; "
            + ("if [ \"$SKIP_SETUP\" = \"1\" ]; then echo '[dispatch] phase=pre_requirements (skipped)'; else echo '[dispatch] phase=pre_requirements'; " + pre_req_prefix + "fi; ")
            + ("if [ \"$SKIP_SETUP\" = \"1\" ]; then echo '[dispatch] phase=setup (skipped)'; else echo '[dispatch] phase=setup'; " + f"cd {remote_repo_dir_quoted}; " + ("echo '[dispatch] setup_cmds=begin'; " if setup else "") + (setup + " " if setup else "") + ("echo '[dispatch] setup_cmds=end'; " if setup else "") + "date -u +%Y-%m-%dT%H:%M:%SZ > \"$SETUP_MARKER\"; fi; ")
            # Always cd into repo before running.
            + f"cd {remote_repo_dir_quoted}; "
            + "echo '[dispatch] phase=run'; "
            # Best-effort: ensure conda is discoverable for remote_python.
            + conda_bootstrap
            # Run the remote sweep but DO NOT abort the whole SSH session on failure;
            # we still want to collect server logs and snapshots.
            + "set +e; "
            # Force python to flush promptly so SSH logs show progress.
            + f"PYTHONUNBUFFERED=1 {server.remote_python} scripts/scale-test/embedding/run_scale_fix_token_len.py "
            + f"--config {shlex.quote(remote_cfg_path)} "
            + "--no-ssh-dispatch "
            + ("--resume " if resume else "")
            + f"--scale-id {shlex.quote(scale_id)} "
            + f"--result-root {shlex.quote(remote_result_root)} "
            + ("--tee " if tee else "")
            + ("--dry-run " if dry_run else "")
            # Background the sweep and emit periodic heartbeats so we can see remote progress even
            # if the sweep is silent (e.g. stuck starting a server / waiting on a request).
            + "& SWEEP_PID=$!; "
            + "echo \"[dispatch] sweep_pid=$SWEEP_PID\"; "
            + "HB=0; AT_LOG=\"\"; "
            + "while kill -0 \"$SWEEP_PID\" 2>/dev/null; do "
            + "HB=$((HB+1)); "
            + "echo \"[dispatch] heartbeat=$HB ts=$(date -u +%Y-%m-%dT%H:%M:%SZ)\"; "
            + "ps -o pid,etime,pcpu,pmem,cmd -p \"$SWEEP_PID\" 2>/dev/null | sed -E 's/^/[dispatch] /' || true; "
            + "if [ -z \"$AT_LOG\" ]; then AT_LOG=$(find \"$REMOTE_RUN_DIR\" -maxdepth 3 -type f -name auto_test_stdout.log 2>/dev/null | head -n 1 || true); fi; "
            + "if [ -n \"$AT_LOG\" ] && [ -f \"$AT_LOG\" ]; then "
            + "  echo \"[dispatch] tail_auto_test_stdout=$(basename \"$AT_LOG\")\"; "
            + "  tail -n 5 \"$AT_LOG\" 2>/dev/null | sed -E 's/^/[dispatch] /' || true; "
            + "fi; "
            + "sleep 60; "
            + "done; "
            + "wait \"$SWEEP_PID\"; RC=$?; set -e; "
            + "echo \"[dispatch] sweep_rc=$RC\"; "
            # Snapshot: port/process state (best-effort)
            + (f"(echo '=== ss -ltnp | grep :{int(base_port)} ==='; ss -ltnp 2>/dev/null | grep -F ':{int(base_port)}' || true) > \"$DEBUG_DIR/port_{int(base_port)}.txt\" 2>&1; " if base_port else "")
            + "(echo '=== ps -ef | egrep (sglang|uvicorn|vllm) ==='; ps -ef | egrep -i 'sglang|uvicorn|vllm' | head -n 200 || true) > \"$DEBUG_DIR/ps_servers.txt\" 2>&1; "
            + "(echo '=== df -h ==='; df -h || true) > \"$DEBUG_DIR/df.txt\" 2>&1; "
            # Collect common server log directories from the repo checkout.
            + "mkdir -p \"$DEBUG_DIR/server_logs\"; "
            + "for d in scripts/embedding/sglang/sglang_logs scripts/embedding/vllm/vllm_logs; do "
            + "  if [ -d \"$REMOTE_REPO_DIR/$d\" ]; then "
            + "    bn=$(basename \"$d\"); "
            + "    cp -a \"$REMOTE_REPO_DIR/$d\" \"$DEBUG_DIR/server_logs/$bn\" 2>/dev/null || true; "
            + "  fi; "
            + "done; "
            # Print a small tail of the newest logs to the SSH stdout so it lands in remote_run.log.
            + "if [ -d \"$DEBUG_DIR/server_logs\" ]; then "
            + "  echo '[dispatch] server_logs_tail=begin'; "
            + "  find \"$DEBUG_DIR/server_logs\" \\( -name '*.log' -o -name '*.txt' \\) -type f 2>/dev/null | head -n 20 | while read -r f; do "
            + "    echo \"--- $f ---\"; tail -n 80 \"$f\" 2>/dev/null || true; "
            + "  done; "
            + "  echo '[dispatch] server_logs_tail=end'; "
            + "fi; "
            + "exit $RC"
        ).strip()

        print(f"[info] dispatch host={host}{' (local)' if is_local else ''}")
        print(f"[info] remote_repo_dir={remote_repo_dir}")
        print(f"[info] remote_run_dir={remote_run_dir}")

        if not tee:
            # Without --tee, the SSH output is still captured to this log, but nothing
            # is printed to the console until the remote command finishes.
            remote_run_log_hint = (local_host_dir / "remote_run.log").resolve()
            print(f"[info] (no --tee) streaming disabled; tail this log for progress: {remote_run_log_hint}")

        if dry_run:
            # Avoid printing secrets. (sshpass is only used when password is provided.)
            if (not is_local) and server.password:
                print(f"[dry-run] ssh (password auth) {target} bash -lc <remote_cmd>")
            else:
                if is_local:
                    print(f"[dry-run] local bash -lc <remote_cmd> (host={host})")
                else:
                    print(f"[dry-run] ssh {target} bash -lc {remote_cmd}")
            continue

        # Local fast-path: no SSH/scp/rsync; execute the same dispatch script locally.
        if is_local:
            # Ensure the repo dir exists locally; do not copy the repo.
            repo_dir_p = Path(str(remote_repo_dir)).expanduser()
            if not repo_dir_p.is_absolute():
                repo_dir_p = (REPO_ROOT / repo_dir_p).resolve()
            else:
                repo_dir_p = repo_dir_p.resolve()
            if not repo_dir_p.exists():
                raise SystemExit(
                    f"local dispatch requires remote_repo_dir to exist on this machine (host={host}): {repo_dir_p}. "
                    f"Tip: set remote_repo_dir={REPO_ROOT} for local runs."
                )

            # Prepare local dispatch cfg dir and copy config + pre-req scripts.
            local_cfg_dir = Path(remote_cfg_dir)
            local_cfg_dir.mkdir(parents=True, exist_ok=True)
            try:
                shutil.copy2(str(cfg_path), str(local_cfg_dir / cfg_path.name))
            except Exception as e:
                raise SystemExit(f"failed to stage config into {local_cfg_dir}: {e}")
            for _, local_p, remote_p in pre_req_items:
                try:
                    Path(remote_p).parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(str(local_p), str(remote_p))
                except Exception:
                    pass

            # Run the dispatch command locally (same bash logic as remote).
            remote_run_log = local_host_dir / "remote_run.log"
            rc_local = _run_stream_to_log(["bash", "-lc", remote_cmd], log_path=remote_run_log, echo=bool(tee))

            eff_rc_local, sweep_rc_local = _effective_dispatch_rc(ssh_rc=int(rc_local), remote_run_log=remote_run_log)
            if rc_local != eff_rc_local and sweep_rc_local is not None:
                print(
                    f"[warn] local dispatch rc mismatch (host={host}): "
                    f"bash_rc={rc_local} but sweep_rc={sweep_rc_local}. "
                    f"Using sweep_rc as authoritative. See {remote_run_log}"
                )
            rc_local = int(eff_rc_local)

            if rc_local != 0:
                print(f"[error] local run failed (host={host}, rc={rc_local}). See {remote_run_log}")
                if not scale.continue_on_error:
                    return int(rc_local)

            # Capture server info locally (best-effort).
            try:
                info_dir = local_host_dir / "server_info"
                info_dir.mkdir(parents=True, exist_ok=True)
                subprocess.run(["lscpu"], stdout=(info_dir / "lscpu.txt").open("w", encoding="utf-8"), stderr=subprocess.STDOUT, text=True, check=False)
                subprocess.run(["bash", "-lc", "lscpu -J"], stdout=(info_dir / "lscpu.json").open("w", encoding="utf-8"), stderr=subprocess.STDOUT, text=True, check=False)
                subprocess.run(["bash", "-lc", "cat /proc/meminfo"], stdout=(info_dir / "meminfo.txt").open("w", encoding="utf-8"), stderr=subprocess.STDOUT, text=True, check=False)
            except Exception:
                pass

            # Copy back the run dir into local_host_dir (rsync-back equivalent).
            _sync_tree(Path(remote_run_dir), local_host_dir)

            # Load host aggregate and rewrite paths to local copies.
            host_agg = local_host_dir / "aggregate.csv"
            host_rows = _read_csv_rows(host_agg)
            for r in host_rows:
                r2: Dict[str, Any] = dict(r)
                r2["server_host"] = host
                for k in ["log_path", "metrics_path", "emon_output_path", "emon_summary_xlsx", "emon_process_log"]:
                    if k in r2:
                        r2[k] = _rewrite_remote_path_to_local(
                            val=str(r2.get(k, "")),
                            remote_run_dir=remote_run_dir,
                            local_host_dir=local_host_dir,
                        )
                all_rows.append(r2)
            continue

        if not _maybe_preflight_socks_proxy(server=server, local_host_dir=local_host_dir, dispatch_env=dispatch_env):
            print(
                f"[error] SOCKS proxy cannot reach {host}:{server.port}. "
                f"Check proxy ACL/routing. See {local_host_dir / 'socks_preflight.log'}"
            )
            return 255

        # 1) mkdir remote cfg dir
        mkdir_cmd = prefix + ["ssh"] + ssh_args + [target, _ssh_bash_lc_remote(f"mkdir -p {shlex.quote(remote_cfg_dir)}")]
        p = subprocess.run(
            mkdir_cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            check=False,
            env=dispatch_env,
        )
        (local_host_dir / "ssh_mkdir.log").write_text(p.stdout or "", encoding="utf-8")
        if p.returncode != 0:
            print(f"[error] ssh mkdir failed (host={host}, rc={p.returncode})")
            return int(p.returncode)

        # Optional: blow away remote_repo_dir, then recreate it before rsync.
        if remote_clean_repo:
            safe_repo_dir = _validate_remote_repo_dir_for_delete(remote_repo_dir)
            clean_cmd = f"rm -rf {shlex.quote(safe_repo_dir)} && mkdir -p {shlex.quote(safe_repo_dir)}"
            if tee:
                print(f"[dispatch] cleaning remote_repo_dir on {host}: {safe_repo_dir}")
            if dry_run:
                (local_host_dir / "remote_repo_clean.log").write_text(
                    _log_meta_header(cmd_for_log=clean_cmd, rc=0) + "[dry-run]\n",
                    encoding="utf-8",
                )
            else:
                p_clean = subprocess.run(
                    prefix + ["ssh"] + ssh_args + [target, _ssh_bash_lc_remote(clean_cmd)],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    check=False,
                    env=dispatch_env,
                )
                (local_host_dir / "remote_repo_clean.log").write_text(
                    _log_meta_header(cmd_for_log=clean_cmd, rc=int(p_clean.returncode)) + (p_clean.stdout or ""),
                    encoding="utf-8",
                )
                if p_clean.returncode != 0:
                    msg = f"remote repo clean failed (host={host}, rc={p_clean.returncode})"
                    if scale.continue_on_error:
                        print(f"[warn] {msg}; continuing. See {local_host_dir / 'remote_repo_clean.log'}")
                    else:
                        raise SystemExit(msg)

        # 1b) If remote_repo_dir is missing, upload+extract the local repo into place.
        ok_repo = True
        if not remote_clean_repo:
            ok_repo = _ensure_remote_repo_dir(
                server=server,
                target=target,
                local_host_dir=local_host_dir,
                remote_cfg_dir=remote_cfg_dir,
                remote_repo_dir=remote_repo_dir,
                scp_bin=str(scp_bin),
                dispatch_env=dispatch_env,
                continue_on_error=scale.continue_on_error,
                repo_bundle_path=repo_bundle_path,
            )
        if not ok_repo:
            print(f"[error] remote repo bootstrap failed (host={host}). See logs under {local_host_dir}")
            if not scale.continue_on_error:
                return 1
            continue

        # 1c) Sync local repo to the remote repo dir so the remote host always
        # runs the latest runner code. Honor .gitignore (rsync filter) and
        # exclude .git. This avoids confusing cases where remote runs stale code
        # and misses pinning debug output.
        if rsync_bin and not dry_run:
            rsync_ssh = _rsync_ssh_command(server)
            rsync_ssh_log = rsync_ssh
            if server.password:
                rsync_ssh_log = rsync_ssh_log.replace(str(server.password), "<redacted>")
            rsync_repo_cmd = [
                str(rsync_bin),
                "-az",
                "--delete",
                "--force",
                "--prune-empty-dirs",
                "--filter=:- .gitignore",
                "--exclude=.git/",
                "-e",
                rsync_ssh,
                str(REPO_ROOT).rstrip("/") + "/",
                f"{target}:{remote_repo_dir.rstrip('/')}/",
            ]
            if tee:
                print(f"[dispatch] syncing repo to {host}:{remote_repo_dir} (honor .gitignore)")
            p1c = subprocess.run(
                rsync_repo_cmd,
                cwd=str(REPO_ROOT),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                check=False,
                env=dispatch_env,
            )
            rsync_dest = f"{target}:{remote_repo_dir.rstrip('/')}/"
            rsync_cmd_for_log = (
                f"{shlex.quote(str(rsync_bin))} -az --delete --force --prune-empty-dirs --filter=:- .gitignore --exclude=.git/ "
                + f"-e {shlex.quote(rsync_ssh_log)} {shlex.quote(str(REPO_ROOT).rstrip('/') + '/')} {shlex.quote(rsync_dest)}"
            )
            (local_host_dir / "rsync_repo_push.log").write_text(
                _log_meta_header(cmd_for_log=rsync_cmd_for_log, rc=int(p1c.returncode)) + (p1c.stdout or ""),
                encoding="utf-8",
            )
            if p1c.returncode != 0:
                print(
                    f"[warn] rsync repo push failed (host={host}, rc={p1c.returncode}). "
                    f"Remote may run stale code. See {local_host_dir / 'rsync_repo_push.log'}"
                )
                if not scale.continue_on_error:
                    return int(p1c.returncode)

        # 2) scp config
        scp_cmd = prefix + [str(scp_bin)] + scp_args + [str(cfg_path), f"{target}:{remote_cfg_path}"]
        p2 = subprocess.run(
            scp_cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            check=False,
            env=dispatch_env,
        )
        (local_host_dir / "scp_config.log").write_text(p2.stdout or "", encoding="utf-8")
        if p2.returncode != 0:
            print(f"[error] scp config failed (host={host}, rc={p2.returncode})")
            if not scale.continue_on_error:
                return int(p2.returncode)
            continue

        # 2b) scp pre-requirements scripts (optional)
        skip_pre_req_scp = False
        if resume and pre_req_items and not dry_run:
            # If resuming and the remote host already has a prepared environment,
            # avoid SCPing pre-requirements scripts (can be flaky behind proxies).
            conda_ok = "true"
            if need_conda:
                conda_ok = "(command -v conda >/dev/null 2>&1 || [ -x \"$HOME/miniforge3/bin/conda\" ])"
            if resume_pip_marker_file:
                preflight = (
                    f"if {conda_ok}; then "
                    + f"  if [ -f {resume_pip_marker_file} ]; then echo SKIP_PIP; "
                    + f"  elif [ -d {shlex.quote(remote_run_dir)} ] && [ -n \"$(ls -A {shlex.quote(remote_run_dir)} 2>/dev/null || true)\" ]; then echo SKIP_RUNDIR; "
                    + "  else echo NEED; fi; "
                    + "else echo NEED_CONDA; fi"
                )
            else:
                preflight = (
                    f"if {conda_ok}; then "
                    + f"  if [ -d {shlex.quote(remote_run_dir)} ] && [ -n \"$(ls -A {shlex.quote(remote_run_dir)} 2>/dev/null || true)\" ]; then echo SKIP_RUNDIR; "
                    + "  else echo NEED; fi; "
                    + "else echo NEED_CONDA; fi"
                )
            chk_cmd = prefix + ["ssh"] + ssh_args + [target, _ssh_bash_lc_remote(preflight)]
            ppf = subprocess.run(
                chk_cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                check=False,
                env=dispatch_env,
            )
            (local_host_dir / "resume_pre_req_preflight.log").write_text(ppf.stdout or "", encoding="utf-8")
            if "SKIP_PIP" in (ppf.stdout or "") or "SKIP_RUNDIR" in (ppf.stdout or ""):
                skip_pre_req_scp = True

        if pre_req_items and not skip_pre_req_scp:
            log_lines: List[str] = []
            for _, local_p, remote_p in pre_req_items:
                if not local_p.exists():
                    raise SystemExit(f"pre_requirements file not found: {local_p}")
                scp_cmd2 = prefix + [str(scp_bin)] + scp_args + [str(local_p), f"{target}:{remote_p}"]
                p2b = subprocess.run(
                    scp_cmd2,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    check=False,
                    env=dispatch_env,
                )
                log_lines.append(p2b.stdout or "")
                if p2b.returncode != 0:
                    (local_host_dir / "scp_pre_requirements.log").write_text("\n".join(log_lines), encoding="utf-8")
                    lvl = "warn" if resume else "error"
                    print(f"[{lvl}] scp pre-requirements failed (host={host}, rc={p2b.returncode})")
                    if not scale.continue_on_error:
                        return int(p2b.returncode)
                    continue
            (local_host_dir / "scp_pre_requirements.log").write_text("\n".join(log_lines), encoding="utf-8")
        elif skip_pre_req_scp:
            (local_host_dir / "scp_pre_requirements.log").write_text(
                "[resume] skipped scp of pre-requirements scripts (remote env appears ready)\n",
                encoding="utf-8",
            )

        # 3) run remote sweep
        run_cmd = prefix + ["ssh"] + ssh_args + [target, _ssh_bash_lc_remote(remote_cmd)]
        remote_run_log = local_host_dir / "remote_run.log"
        rc3 = _run_stream_to_log(run_cmd, log_path=remote_run_log, echo=bool(tee))
        if rc3 != 0:
            did_retry = False
            if (not dry_run) and (not is_local):
                try:
                    log_text = remote_run_log.read_text(encoding="utf-8", errors="replace")
                except Exception:
                    log_text = ""

                if _remote_log_needs_sglang_prestage(log_text):
                    if tee:
                        print(f"[dispatch] detected sglang clone failure; prestaging locally and retrying (host={host})")
                    tar_path = _ensure_prestaged_sglang_tarball(log_dir=local_host_dir, echo=False)
                    if tar_path:
                        remote_assets_dir = f"{remote_cfg_dir}/assets"
                        remote_tar = f"{remote_assets_dir}/sglang.tar.gz"

                        # Ensure remote assets dir exists.
                        subprocess.run(
                            prefix
                            + ["ssh"]
                            + ssh_args
                            + [target, _ssh_bash_lc_remote(f"mkdir -p {shlex.quote(remote_assets_dir)}")],
                            stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT,
                            text=True,
                            check=False,
                            env=dispatch_env,
                        )

                        # Upload tarball.
                        p_tar = subprocess.run(
                            prefix + [str(scp_bin)] + scp_args + [str(tar_path), f"{target}:{remote_tar}"],
                            stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT,
                            text=True,
                            check=False,
                            env=dispatch_env,
                        )
                        (local_host_dir / "scp_prestage_sglang.log").write_text(p_tar.stdout or "", encoding="utf-8")

                        if p_tar.returncode == 0:
                            # Extract into $HOME so install_sglang.sh's default src_dir='sglang' works.
                            extract_cmd = (
                                "set -e; "
                                + f"TAR={shlex.quote(remote_tar)}; "
                                + "DEST=$HOME; "
                                + "rm -rf \"$DEST/sglang\" >/dev/null 2>&1 || true; "
                                + "if command -v tar >/dev/null 2>&1; then "
                                + "  tar -xzf \"$TAR\" -C \"$DEST\"; "
                                + "elif command -v python3 >/dev/null 2>&1; then "
                                + "  python3 - <<'PY'\n"
                                + "import os, tarfile\n"
                                + "tar_path=os.environ.get('TAR')\n"
                                + "dest=os.environ.get('DEST')\n"
                                + "with tarfile.open(tar_path, 'r:gz') as tf: tf.extractall(dest)\n"
                                + "print('[ok] extracted sglang tarball')\n"
                                + "PY\n"
                                + "else "
                                + "  echo '[error] neither tar nor python3 available for extraction' >&2; exit 1; "
                                + "fi"
                            )
                            p_ext = subprocess.run(
                                prefix + ["ssh"] + ssh_args + [target, _ssh_bash_lc_remote(extract_cmd)],
                                stdout=subprocess.PIPE,
                                stderr=subprocess.STDOUT,
                                text=True,
                                check=False,
                                env=dispatch_env,
                            )
                            (local_host_dir / "remote_extract_prestage_sglang.log").write_text(
                                p_ext.stdout or "", encoding="utf-8"
                            )

                            if p_ext.returncode == 0:
                                remote_run_retry_log = local_host_dir / "remote_run.retry.log"
                                rc_retry = _run_stream_to_log(run_cmd, log_path=remote_run_retry_log, echo=bool(tee))
                                did_retry = True
                                rc3 = int(rc_retry)

            # Reconcile rc with remote-emitted sweep_rc (authoritative).
            rc_log_path = (local_host_dir / "remote_run.retry.log") if did_retry else remote_run_log
            eff_rc3, sweep_rc3 = _effective_dispatch_rc(ssh_rc=int(rc3), remote_run_log=rc_log_path)
            if rc3 != eff_rc3 and sweep_rc3 is not None:
                print(
                    f"[warn] remote dispatch rc mismatch (host={host}): "
                    f"ssh_rc={rc3} but sweep_rc={sweep_rc3}. "
                    f"Using sweep_rc as authoritative. See {rc_log_path}"
                )
            rc3 = int(eff_rc3)

            if rc3 != 0:
                if did_retry:
                    print(
                        f"[error] remote run failed after prestage retry (host={host}, rc={rc3}). "
                        f"See {local_host_dir / 'remote_run.retry.log'} (and {remote_run_log})"
                    )
                else:
                    print(f"[error] remote run failed (host={host}, rc={rc3}). See {remote_run_log}")
                if not scale.continue_on_error:
                    return int(rc3)

        # 3b) capture server info (best-effort)
        try:
            _capture_remote_lscpu(local_host_dir=local_host_dir, server=server, dispatch_env=dispatch_env)
        except Exception:
            pass

        # 4) rsync back (best-effort). Even when the remote run fails, try to
        # copy back whatever was written under remote_run_dir (including
        # dispatch_debug server logs).
        rsync_src = f"{target}:{remote_run_dir.rstrip('/')}/"
        if rsync_bin and not dry_run:
            # Avoid a noisy rsync error if the directory truly does not exist.
            chk_cmd = prefix + ["ssh"] + ssh_args + [target, _ssh_bash_lc_remote(f"test -d {shlex.quote(remote_run_dir)} && echo OK || echo MISSING")]
            pchk = subprocess.run(
                chk_cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                check=False,
                env=dispatch_env,
            )
            (local_host_dir / "remote_run_dir_check.log").write_text(pchk.stdout or "", encoding="utf-8")
            if "OK" not in (pchk.stdout or ""):
                (local_host_dir / "rsync_back.log").write_text(
                    f"remote_run_dir missing on host (rc3={rc3}); skipped rsync_back.\n",
                    encoding="utf-8",
                )
            else:
                rsync_ssh = _rsync_ssh_command(server)
                rsync_cmd = [str(rsync_bin), "-az", "-e", rsync_ssh, rsync_src, str(local_host_dir) + "/"]
                p4 = subprocess.run(
                    rsync_cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    check=False,
                    env=dispatch_env,
                )
                (local_host_dir / "rsync_back.log").write_text(p4.stdout or "", encoding="utf-8")
                if p4.returncode != 0:
                    print(f"[error] rsync failed (host={host}, rc={p4.returncode}). See {local_host_dir / 'rsync_back.log'}")
                    if not scale.continue_on_error:
                        return int(p4.returncode)
                    continue

        # 5) load host aggregate and rewrite paths to local copies
        host_agg = local_host_dir / "aggregate.csv"
        host_rows = _read_csv_rows(host_agg)
        for r in host_rows:
            r2: Dict[str, Any] = dict(r)
            r2["server_host"] = host
            for k in ["log_path", "metrics_path", "emon_output_path", "emon_summary_xlsx", "emon_process_log"]:
                if k in r2:
                    r2[k] = _rewrite_remote_path_to_local(
                        val=str(r2.get(k, "")),
                        remote_run_dir=remote_run_dir,
                        local_host_dir=local_host_dir,
                    )
            all_rows.append(r2)

    if dry_run:
        print("[ok] dry-run complete (no remote execution, no results copied)")
        return 0

    # Write combined aggregate in the local run root.
    agg_csv = out_dir / "aggregate.csv"
    _write_aggregate_csv(out_csv=agg_csv, rows=all_rows)
    print(f"[ok] Wrote aggregate: {agg_csv}")
    return 0


def _generate_auto_test_config(
    *,
    scale: ScaleConfig,
    result_dir: Path,
    cpu_expr: str,
    sglang_max_total_tokens: str,
    variant_tag: str = "",
) -> Dict[str, Any]:
    template = _load_json(scale.template_auto_test_config)
    if not isinstance(template, dict):
        raise SystemExit(f"template auto-test config is not an object: {scale.template_auto_test_config}")

    cfg = dict(template)
    defaults = dict(cfg.get("defaults") or {})
    defaults["result_dir"] = str(result_dir)
    # Let the scale-test control warmup via per-job warmup_runs.
    cfg["defaults"] = defaults

    jt = scale.job_template
    backend = str(jt.get("backend") or "sglang").strip()
    model = str(jt.get("model") or "").strip()
    if not model:
        raise SystemExit("job_template.model is required")
    model_dir = str(jt.get("model_dir") or "").strip()
    base_url = str(jt.get("base_url") or "").strip()

    host = str(jt.get("host") or "0.0.0.0").strip()
    port = int(jt.get("port") or 30000)
    served_model_name = str(jt.get("served_model_name") or jt.get("model_id") or "").strip()
    model_id = str(jt.get("model_id") or served_model_name or model).strip()

    mode = str(jt.get("mode") or "token_len").strip()
    max_samples = int(jt.get("max_samples") or 1000)
    batch_size_default = int(jt.get("batch_size") or 100)
    dtype = str(jt.get("dtype") or "bfloat16").strip()

    restart_servers = bool(jt.get("restart_servers", True))
    stop_server_after_job = bool(jt.get("stop_server_after_job", True))
    emon_enable = bool(jt.get("emon_enable", False))
    extra_env = jt.get("extra_env") or {}
    if not isinstance(extra_env, dict):
        extra_env = {}

    # -------------------------
    # Auto NUMA binding (best-effort)
    #
    # Goal: if the user constrains CPUs (e.g. 0-31), we can often infer the
    # NUMA node and bind server + benchmark to local memory for better
    # locality and more predictable behavior under MemoryMax.
    #
    # We only apply this when:
    # - scale.cpu_expr is set
    # - user did not explicitly specify NUMACTL_* / SERVER_NUMACTL_* in extra_env
    # - server spec does not already configure servers.<backend>.numactl
    #
    # We only auto-bind CPU cores (numactl -C). We intentionally do NOT auto-add
    # --cpunodebind/--membind because:
    # - it's redundant when -C already pins to a node-local core subset
    # - it can be harmful if the CPU list spans multiple nodes
    # Users can still explicitly set NUMACTL_CPUNODEBIND/NUMACTL_MEMBIND (and
    # SERVER_* variants) in extra_env when they want strict node/memory binding.
    # -------------------------
    inferred_node: Optional[int] = None
    if (cpu_expr or "").strip():
        inferred_node = _infer_single_numa_node_from_cpu_expr(cpu_expr)

    # If the chosen backend uses an HTTP server managed by auto-test, we may
    # need to forward extra env vars into the server start script.
    # In particular, the sglang server script supports SGLANG_PYTHON to pick
    # the correct interpreter (important when running under systemd-run).
    try:
        servers = cfg.get("servers") or {}
        if isinstance(servers, dict) and backend in servers and isinstance(servers.get(backend), dict):
            s = dict(servers[backend])
            env_from_job = dict(s.get("env_from_job") or {})
            for k in extra_env.keys():
                ks = str(k)
                if ks.startswith("SGLANG_"):
                    env_from_job[ks] = ks
            # Also forward sweep-provided SGLANG settings that may not exist in extra_env.
            if str(sglang_max_total_tokens or "").strip():
                env_from_job["SGLANG_MAX_TOTAL_TOKENS"] = "SGLANG_MAX_TOTAL_TOKENS"
            s["env_from_job"] = env_from_job

            # Auto-populate server NUMA binding from scale.cpu_expr when safe.
            # (Server numactl is configured in the server spec, not via env vars.)
            user_server_override = any(
                str(extra_env.get(k) or "").strip()
                for k in ["SERVER_NUMACTL_CORES", "SERVER_NUMACTL_CPUNODEBIND", "SERVER_NUMACTL_MEMBIND"]
            )
            user_job_override = any(
                str(extra_env.get(k) or "").strip() for k in ["NUMACTL_CORES", "NUMACTL_CPUNODEBIND", "NUMACTL_MEMBIND"]
            )

            # Only touch server numactl if user didn't opt into explicit controls.
            # IMPORTANT: the auto-test template config may pin the server to a fixed core range
            # (e.g. 0-15). For scale-test we want the server to follow the scale CPU constraint.
            if (cpu_expr or "").strip() and not user_server_override and not user_job_override:
                numactl = s.get("numactl")
                numactl_obj: Dict[str, Any] = dict(numactl) if isinstance(numactl, dict) else {}

                # Always bind server cores to the scale CPU expr.
                numactl_obj["cores"] = cpu_expr.strip()

                # Do not auto-add cpunodebind/membind; keep only core pinning.

                s["numactl"] = numactl_obj

            servers[backend] = s
            cfg["servers"] = servers
    except Exception:
        pass

    jobs: List[Dict[str, Any]] = []
    repeats = max(1, int(scale.repeats))
    for bs in scale.batch_sizes:
        for tl in scale.token_lens:
            name = f"scale_fix_token_len_tok{int(tl)}_bs{int(bs)}_{model}_{backend}"
            if (variant_tag or "").strip():
                name = f"{name}__{_safe_name(variant_tag)}"

            env = {
                "MODEL": model,
                "BACKEND": backend,
                "MODEL_DIR": model_dir,
                "SERVED_MODEL_NAME": served_model_name,
                "MODEL_ID": model_id,
                "HOST": host,
                "PORT": str(port),
                "BASE_URL": base_url,
                "MODE": mode,
                "MAX_SAMPLES": str(max_samples),
                "BATCH_SIZE": str(int(bs) if int(bs) > 0 else batch_size_default),
                "SYNTHETIC_TOKEN_LEN": str(int(tl)),
                "DTYPE": dtype,
            }

            # For HTTP backends, large batch sizes on CPU can exceed the default
            # 120s request timeout during warmup/first runs.
            #
            # Use job_template.embedding_http_timeout_sec / http_timeout_sec to
            # override; otherwise default to 900s.
            http_timeout_sec: Optional[int] = None
            for k in ["embedding_http_timeout_sec", "http_timeout_sec", "embedding_http_timeout", "http_timeout"]:
                v = jt.get(k)
                if v is None or str(v).strip() == "":
                    continue
                try:
                    http_timeout_sec = int(float(v))
                    break
                except Exception:
                    continue
            if http_timeout_sec is None:
                http_timeout_sec = 900
            env.setdefault("EMBEDDING_HTTP_TIMEOUT", str(http_timeout_sec))

            # Merge user-provided env first (can override any defaults).
            for k, v in extra_env.items():
                env[str(k)] = str(v)

            # Optional sweep override (server + job): SGLANG_MAX_TOTAL_TOKENS.
            if str(sglang_max_total_tokens or "").strip():
                env["SGLANG_MAX_TOTAL_TOKENS"] = str(sglang_max_total_tokens).strip()

            # Auto-populate benchmark NUMA binding when safe.
            # (This wraps the job command with numactl in run_auto_test.py.)
            if (cpu_expr or "").strip() and not any(
                str(extra_env.get(k) or "").strip()
                for k in [
                    "NUMACTL_CORES",
                    "NUMACTL_CPUNODEBIND",
                    "NUMACTL_MEMBIND",
                    "SERVER_NUMACTL_CORES",
                    "SERVER_NUMACTL_CPUNODEBIND",
                    "SERVER_NUMACTL_MEMBIND",
                ]
            ):
                if not str(env.get("NUMACTL_CORES") or "").strip():
                    env["NUMACTL_CORES"] = cpu_expr.strip()
                # Also bind the *server* side explicitly. This is important
                # because some environments accept systemd-run properties
                # but do not actually enforce cpusets, and sglang can
                # otherwise fan out across all host CPUs.
                if not str(env.get("SERVER_NUMACTL_CORES") or "").strip():
                    env["SERVER_NUMACTL_CORES"] = cpu_expr.strip()
                # Do not auto-add NUMACTL_CPUNODEBIND/NUMACTL_MEMBIND.

            jobs.append(
                {
                    "name": name,
                    "script": "run_fix_token_len",
                    "args": {},
                    "warmup_runs": int(scale.warmup_runs),
                    "repeats": int(1 if scale.repeat_threshold_sec is not None else repeats),
                    "repeat_threshold_sec": float(scale.repeat_threshold_sec) if scale.repeat_threshold_sec is not None else None,
                    "repeat_max_repeats": int(scale.repeat_max_repeats),
                    "restart_servers": restart_servers,
                    "stop_server_after_job": stop_server_after_job,
                    "emon_enable": emon_enable,
                    "env": env,
                }
            )

    cfg["jobs"] = jobs
    return cfg


def _run_auto_test(
    *,
    auto_test_config_path: Path,
    work_dir: Path,
    tee: bool,
    dry_run: bool,
    cpu_expr: str,
    mem_gb: Optional[float],
    stdout_log: Path,
    only_jobs: Optional[List[str]] = None,
    skip_jobs: Optional[List[str]] = None,
) -> int:
    runner = (REPO_ROOT / "scripts/auto-test/embedding/run_auto_test.py").resolve()
    if not runner.exists():
        raise SystemExit(f"Missing runner: {runner}")

    def _wrap_with_scope_probe(inner_cmd: List[str]) -> List[str]:
        # Runs inside the constrained scope / taskset wrapper (if any) so we can
        # see what CPU/mem restrictions actually apply to the workload.
        script = (
            "echo '[scope] pid='$$; "
            "if [[ -r /proc/self/status ]]; then "
            "awk '/^Cpus_allowed_list:|^Mems_allowed_list:|^Cpus_allowed:/ {print \"[scope] \" $0}' /proc/self/status || true; "
            "fi; "
            "if command -v taskset >/dev/null 2>&1; then taskset -pc $$ || true; fi; "
            "exec \"$@\""
        )
        return ["bash", "-lc", script, "bash"] + inner_cmd

    cmd = [sys.executable, str(runner), "--config", str(auto_test_config_path)]
    if only_jobs:
        for name in only_jobs:
            n = str(name or "").strip()
            if n:
                cmd.extend(["--only", n])
    if skip_jobs:
        for name in skip_jobs:
            n = str(name or "").strip()
            if n:
                cmd.extend(["--skip", n])
    if tee:
        cmd.append("--tee")
    if dry_run:
        cmd.append("--dry-run")

    mem_bytes = _bytes_from_gb(mem_gb)
    want_constraints = bool((cpu_expr or "").strip() or (mem_bytes and mem_bytes > 0))
    cmd2 = _wrap_with_scope_probe(cmd) if want_constraints else cmd
    full_cmd, method = _constrained_cmd(cmd=cmd2, cpu_expr=cpu_expr, mem_bytes=mem_bytes)
    if method == "none" and ((cpu_expr or "").strip() or (mem_bytes and mem_bytes > 0)):
        print("[warn] No resource limiter found (systemd-run/taskset/prlimit). Running without enforcement.")
    else:
        print(f"[info] Resource limiting method: {method}")
        if (cpu_expr or "").strip():
            print(f"[info] CPU cores: {cpu_expr}")
        if mem_bytes and mem_bytes > 0:
            print(f"[info] MemoryMax(bytes): {mem_bytes}")

    stdout_log.parent.mkdir(parents=True, exist_ok=True)
    with stdout_log.open("w", encoding="utf-8") as f:
        env = os.environ.copy()
        if (cpu_expr or "").strip():
            # Let the child runner self-apply affinity and print evidence.
            env["AUTO_TEST_CPU_EXPR"] = cpu_expr.strip()
        p = subprocess.Popen(
            full_cmd,
            cwd=str(work_dir),
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )
        assert p.stdout is not None
        for line in p.stdout:
            f.write(line)
            if tee:
                try:
                    sys.stdout.write(line)
                except BrokenPipeError:
                    # If output is piped (e.g. `| head`) and the pipe closes early,
                    # keep running but stop teeing to stdout.
                    tee = False
        return int(p.wait())


def _find_single_run_id(result_dir: Path) -> str:
    # run_auto_test creates a subdir named run_id inside result_dir.
    candidates = [p.name for p in result_dir.iterdir() if p.is_dir()]
    # Filter to compact UTC IDs: 20260101T010203Z
    candidates = [c for c in candidates if len(c) == 16 and c.endswith("Z") and "T" in c]
    if not candidates:
        raise SystemExit(f"No run_id directories found under {result_dir}")
    return sorted(candidates)[-1]


def _iter_summary_rows(summary_csv: Path) -> Iterable[Dict[str, str]]:
    with summary_csv.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield {k: (v if v is not None else "") for k, v in row.items()}


def _iter_summary_csv_paths(result_dir: Path) -> List[Path]:
    # Auto-test writes summary_<run_id>.csv at the suite root (result_dir).
    out: List[Path] = []
    try:
        for p in result_dir.glob("summary_*.csv"):
            if p.is_file():
                out.append(p)
    except Exception:
        return []
    return sorted(out)


def _extract_run_id_from_summary_path(p: Path) -> str:
    # summary_<run_id>.csv
    name = p.name
    if not name.startswith("summary_") or not name.endswith(".csv"):
        return ""
    return name[len("summary_") : -len(".csv")]


def _parse_int(s: Any) -> Optional[int]:
    try:
        if s is None:
            return None
        ss = str(s).strip()
        if ss == "":
            return None
        return int(float(ss))
    except Exception:
        return None


def _completed_ok_jobs_from_summaries(result_dir: Path) -> set[str]:
    ok: set[str] = set()
    for p in _iter_summary_csv_paths(result_dir):
        try:
            for row in _iter_summary_rows(p):
                job = str(row.get("job_name") or "").strip()
                if not job:
                    continue
                ec = _parse_int(row.get("exit_code"))
                if ec == 0:
                    ok.add(job)
        except Exception:
            continue
    return ok


def _best_rows_by_job_from_summaries(result_dir: Path) -> Dict[str, Dict[str, str]]:
    # Choose the latest successful row per job across multiple summary_<run_id>.csv files.
    # If there is no successful row, fall back to the latest row we saw.
    best: Dict[str, Tuple[str, bool, Dict[str, str]]] = {}
    for p in _iter_summary_csv_paths(result_dir):
        run_id = _extract_run_id_from_summary_path(p)
        try:
            for row in _iter_summary_rows(p):
                job = str(row.get("job_name") or "").strip()
                if not job:
                    continue
                ec = _parse_int(row.get("exit_code"))
                ok = ec == 0

                prev = best.get(job)
                if prev is None:
                    best[job] = (run_id, ok, row)
                    continue

                prev_run_id, prev_ok, _prev_row = prev
                # Prefer successful over failed.
                if prev_ok and not ok:
                    continue
                if ok and not prev_ok:
                    best[job] = (run_id, ok, row)
                    continue
                # Same success state: prefer lexicographically latest run_id (UTC compact sorts).
                if run_id and prev_run_id and run_id > prev_run_id:
                    best[job] = (run_id, ok, row)
                elif run_id and not prev_run_id:
                    best[job] = (run_id, ok, row)
        except Exception:
            continue

    return {job: rec[2] for job, rec in best.items()}


def _parse_dispatch_sweep_rc(remote_run_log: Path) -> Optional[int]:
    # Look for the last line like: [dispatch] sweep_rc=0
    try:
        if not remote_run_log.exists():
            return None
        last_rc: Optional[int] = None
        with remote_run_log.open("r", encoding="utf-8", errors="replace") as f:
            for line in f:
                if "[dispatch] sweep_rc=" not in line:
                    continue
                try:
                    tail = line.strip().split("sweep_rc=", 1)[1]
                    val = tail.split()[0].strip()
                    rc = int(val)
                    last_rc = rc
                except Exception:
                    continue
        return last_rc
    except Exception:
        return None


def _process_emon_dirs(*, rows: List[Dict[str, Any]], tee: bool, process_cmd: List[str], expected_output: str) -> None:
    emon_bin = _which("emon")
    if not emon_bin:
        print("[warn] emon not found in PATH; skipping emon -process-pyedp")
        return
    if not process_cmd:
        return

    for rec in rows:
        emon_out = str(rec.get("emon_output_path") or "").strip()
        if not emon_out:
            continue
        emon_dat = Path(emon_out)
        if not emon_dat.exists():
            continue
        emon_dir = emon_dat.parent
        out_xlsx = emon_dir / expected_output
        if out_xlsx.exists():
            rec["emon_summary_xlsx"] = str(out_xlsx)
            continue

        cmd = [process_cmd[0]] + process_cmd[1:]
        # Ensure we invoke the resolved emon binary if user used just "emon".
        if cmd and cmd[0] == "emon":
            cmd[0] = emon_bin
        if tee:
            print(f"[emon] processing in {emon_dir}: {' '.join(cmd)}")
        try:
            p = subprocess.run(cmd, cwd=str(emon_dir), stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, check=False)
            rec["emon_process_rc"] = int(p.returncode)
            rec["emon_process_log"] = str((emon_dir / "emon_process.log").resolve())
            (emon_dir / "emon_process.log").write_text(p.stdout or "", encoding="utf-8")
        except Exception as e:
            rec["emon_process_rc"] = 999
            rec["emon_process_err"] = str(e)

        if out_xlsx.exists():
            rec["emon_summary_xlsx"] = str(out_xlsx)


def _maybe_extract_emon_key_values(*, rows: List[Dict[str, Any]]) -> None:
    # Best-effort: if openpyxl is installed, read the first sheet of summary.xlsx
    # and extract simple key/value pairs (two-column layouts are common).
    try:
        import openpyxl  # type: ignore
    except Exception:
        return

    for rec in rows:
        xlsx = str(rec.get("emon_summary_xlsx") or "").strip()
        if not xlsx:
            continue
        p = Path(xlsx)
        if not p.exists():
            continue
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
            ws = wb[wb.sheetnames[0]]
            kv: Dict[str, Any] = {}
            # Read first 200 rows; look for (key, value) pairs.
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), start=1):
                if not row:
                    continue
                if len(row) < 2:
                    continue
                k = row[0]
                v = row[1]
                if k is None or v is None:
                    continue
                ks = str(k).strip()
                if not ks:
                    continue
                # Avoid huge keys.
                if len(ks) > 80:
                    continue
                if ks in kv:
                    continue
                kv[ks] = v
                if len(kv) >= 80:
                    break
            if kv:
                rec["emon_kv"] = kv
        except Exception:
            continue


def _write_aggregate_csv(*, out_csv: Path, rows: List[Dict[str, Any]]) -> None:
    # Flatten emon_kv: keep it as JSON string to avoid explosion.
    fields = [
        "variant",
        "job_name",
        "backend",
        "model",
        "model_id",
        "batch_size",
        "token_len",
        "tps",
        "latency_sec",
        "avg_batch_time_sec",
        "count",
        "num_batches",
        "exit_code",
        "started_at_utc",
        "ended_at_utc",
        "log_path",
        "metrics_path",
        "emon_enabled",
        "emon_output_path",
        "emon_summary_xlsx",
        "emon_process_rc",
        "emon_process_log",
        "resource_cpu",
        "resource_cpu_count",
        "resource_mem_gb",
        "sglang_max_total_tokens",
        "tps_per_cpu",
        "emon_kv_json",
        "server_host",
    ]
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            emon_kv = r.get("emon_kv")
            r2 = dict(r)
            if emon_kv is not None:
                try:
                    r2["emon_kv_json"] = json.dumps(emon_kv, ensure_ascii=False)
                except Exception:
                    r2["emon_kv_json"] = ""
            else:
                r2["emon_kv_json"] = ""

            # Convenience: allow comparing runs with different CPU caps.
            try:
                cpu_count = int(r2.get("resource_cpu_count") or 0)
            except Exception:
                cpu_count = 0
            try:
                tps = float(r2.get("tps") or 0.0)
            except Exception:
                tps = 0.0
            r2["tps_per_cpu"] = (tps / cpu_count) if cpu_count > 0 else ""

            w.writerow({k: r2.get(k, "") for k in fields})
    pass

def _run_post_analyze(*, run_dir: Path, tee: bool) -> None:
    analyzer = (Path(__file__).resolve().parent / "analyze_run.py").resolve()
    if not analyzer.exists():
        print(f"[warn] analyzer not found; skipping: {analyzer}")
        return

    out_dir = run_dir / "analysis"
    out_dir.mkdir(parents=True, exist_ok=True)
    log_path = out_dir / "autogen_analyze_run.log"

    venv_py = (REPO_ROOT / ".venv" / "bin" / "python").resolve()
    py = str(venv_py) if venv_py.exists() else sys.executable
    cmd = [py, str(analyzer), str(run_dir)]
    try:
        proc = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            check=False,
            timeout=15 * 60,
        )
    except subprocess.TimeoutExpired:
        msg = f"[warn] analyze_run.py timed out; see log: {log_path}"
        print(msg)
        try:
            log_path.write_text(msg + "\n", encoding="utf-8")
        except Exception:
            pass
        return
    except Exception as e:
        msg = f"[warn] analyze_run.py failed to execute: {e}"
        print(msg)
        try:
            log_path.write_text(msg + "\n", encoding="utf-8")
        except Exception:
            pass
        return

    out = proc.stdout or ""
    try:
        log_path.write_text(out, encoding="utf-8")
    except Exception:
        pass

    if tee and out:
        print(out, end="" if out.endswith("\n") else "\n")

    if proc.returncode != 0:
        print(f"[warn] post-run analysis failed (rc={proc.returncode}). See: {log_path}")
        return

    print(f"[ok] Wrote analysis: {out_dir}")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True, help="Path to scale-test config JSON")
    ap.add_argument("--tee", action="store_true", help="Stream auto-test output to console")
    ap.add_argument("--no-tee", action="store_true", help="Do not stream output to console (still logs to files)")
    ap.add_argument("--dry-run", action="store_true", help="Do not execute jobs; print commands only")
    ap.add_argument("--no-analyze", action="store_true", help="Skip post-run analysis generation")
    ap.add_argument(
        "--remote-clean-repo",
        nargs="?",
        const="true",
        default="false",
        help=(
            "If true, delete the remote_repo_dir on remote hosts before syncing the local repo via rsync "
            "(default: false). Accepts optional value: true/false."
        ),
    )
    ap.add_argument(
        "--no-ssh-dispatch",
        action="store_true",
        help="Ignore config.run.servers and run locally (useful on remote workers)",
    )
    ap.add_argument(
        "--resume",
        action="store_true",
        help="Resume an existing run directory by skipping jobs/hosts already completed successfully",
    )
    ap.add_argument("--scale-id", default="", help="Override scale_id (default: utc timestamp)")
    ap.add_argument(
        "--result-root",
        default="",
        help="Override result_root directory (useful for remote dispatch)",
    )
    args = ap.parse_args()

    tee = bool(args.tee) and not bool(args.no_tee)

    cfg_path_in = Path(args.config)
    cfg_path: Path
    if cfg_path_in.is_absolute():
        cfg_path = cfg_path_in
    else:
        # Prefer resolving relative paths against the user's current working directory.
        # This is important when running from scripts/scale-test/embedding.
        cand_cwd = (Path.cwd() / cfg_path_in).resolve()
        if cand_cwd.exists():
            cfg_path = cand_cwd
        else:
            cand_repo = (REPO_ROOT / cfg_path_in).resolve()
            cfg_path = cand_repo
    if not cfg_path.exists():
        raise SystemExit(
            "Config not found. Tried:\n"
            f"- { (Path.cwd() / cfg_path_in).resolve() }\n"
            f"- { (REPO_ROOT / cfg_path_in).resolve() }\n"
        )

    scale = _parse_scale_config(cfg_path, resolve_ssh_passwords=not bool(args.no_ssh_dispatch))

    # Allow overriding result_root from CLI.
    result_root = scale.result_root
    if str(args.result_root or "").strip():
        rr_in = Path(str(args.result_root)).expanduser()
        if rr_in.is_absolute():
            result_root = rr_in.resolve()
        else:
            result_root = (REPO_ROOT / rr_in).resolve()

    scale_id = str(args.scale_id or "").strip() or _utc_compact()

    # Multi-host dispatch mode: run the same sweep on all servers via SSH and
    # copy results back under <result_root>/<scale_id>/hosts/<host>/.
    if scale.dispatch.servers and not bool(args.no_ssh_dispatch):
        dispatch_rc = _dispatch_multi_host(
            cfg_path=cfg_path,
            scale=scale,
            result_root=result_root,
            scale_id=scale_id,
            tee=tee,
            dry_run=bool(args.dry_run),
            remote_clean_repo=_parse_bool(args.remote_clean_repo),
            resume=bool(args.resume),
        )

        # Generate local analysis artifacts for the aggregated run directory so
        # the web UI can discover it (it keys off <run_dir>/analysis/).
        if not bool(args.no_analyze):
            run_dir = (result_root / scale_id).resolve()
            if run_dir.exists() and (run_dir / "aggregate.csv").exists():
                _run_post_analyze(run_dir=run_dir, tee=tee)
        return int(dispatch_rc)

    out_dir = result_root / scale_id
    out_dir.mkdir(parents=True, exist_ok=True)

    # Capture local server CPU info for this run (best-effort).
    if not args.dry_run:
        try:
            _capture_local_lscpu(out_dir=out_dir)
        except Exception:
            pass

    print(f"[info] scale_id={scale_id}")

    cpu_exprs = [str(x or "").strip() for x in (scale.cpu_exprs or [""])]
    if not cpu_exprs:
        cpu_exprs = [""]
    kv_list = [str(x or "").strip() for x in (scale.sglang_max_total_tokens or [""])]
    if not kv_list:
        kv_list = [""]

    single_variant = (len(cpu_exprs) * len(kv_list)) == 1
    variants: List[Dict[str, Any]] = []
    for cpu_expr in cpu_exprs:
        for kv in kv_list:
            variant_name = f"cpu_{_safe_name(cpu_expr or 'none')}__kv_{_safe_name(kv or 'auto')}"
            variant_dir = out_dir if single_variant else (out_dir / variant_name)
            variants.append(
                {
                    "variant": variant_name,
                    "cpu_expr": cpu_expr,
                    "sglang_max_total_tokens": kv,
                    "dir": str(variant_dir),
                }
            )

    if not single_variant:
        _dump_json(out_dir / "variants.json", {"scale_id": scale_id, "variants": variants})

    rows: List[Dict[str, Any]] = []
    for v in variants:
        variant_name = str(v["variant"])
        cpu_expr = str(v.get("cpu_expr") or "")
        kv = str(v.get("sglang_max_total_tokens") or "")
        variant_dir = Path(str(v.get("dir") or ""))
        variant_dir.mkdir(parents=True, exist_ok=True)

        auto_test_cfg = _generate_auto_test_config(
            scale=scale,
            result_dir=variant_dir,
            cpu_expr=cpu_expr,
            sglang_max_total_tokens=kv,
            variant_tag=variant_name if not single_variant else "",
        )
        auto_test_cfg_path = variant_dir / "auto_test_config.generated.json"
        _dump_json(auto_test_cfg_path, auto_test_cfg)

        expected_jobs: List[str] = []
        try:
            for j in (auto_test_cfg.get("jobs") or []):
                if isinstance(j, dict) and str(j.get("name") or "").strip():
                    expected_jobs.append(str(j.get("name") or "").strip())
        except Exception:
            expected_jobs = []

        print(f"[info] variant={variant_name}")
        print(f"[info] auto-test config: {auto_test_cfg_path}")

        cpu_count = _cpu_count(cpu_expr)
        if (cpu_expr or "").strip():
            print(f"[info] CPU expr: {cpu_expr}")
            if cpu_count > 0:
                print(f"[info] CPU count: {cpu_count}")
        if (kv or "").strip():
            print(f"[info] SGLANG_MAX_TOTAL_TOKENS: {kv}")

        if bool(args.resume) and not bool(args.dry_run) and expected_jobs:
            ok_jobs = _completed_ok_jobs_from_summaries(variant_dir)
            missing = [j for j in expected_jobs if j not in ok_jobs]
            if not missing:
                print(f"[resume] all jobs already complete; skipping runner (variant={variant_name})")
                rc = 0
            else:
                print(f"[resume] running missing jobs: {len(missing)}/{len(expected_jobs)} (variant={variant_name})")
                rc = _run_auto_test(
                    auto_test_config_path=auto_test_cfg_path,
                    work_dir=REPO_ROOT,
                    tee=tee,
                    dry_run=bool(args.dry_run),
                    cpu_expr=cpu_expr,
                    mem_gb=scale.mem_gb,
                    stdout_log=variant_dir / "auto_test_stdout.log",
                    only_jobs=missing,
                )
        else:
            rc = _run_auto_test(
                auto_test_config_path=auto_test_cfg_path,
                work_dir=REPO_ROOT,
                tee=tee,
                dry_run=bool(args.dry_run),
                cpu_expr=cpu_expr,
                mem_gb=scale.mem_gb,
                stdout_log=variant_dir / "auto_test_stdout.log",
            )
        if rc != 0:
            print(
                f"[error] auto-test runner failed (variant={variant_name}, rc={rc}). "
                f"See {variant_dir / 'auto_test_stdout.log'}"
            )
            if not scale.continue_on_error:
                return rc

        if args.dry_run:
            continue

        summary_found = False
        try:
            best_rows = _best_rows_by_job_from_summaries(variant_dir)
            if best_rows:
                summary_found = True
                for row in best_rows.values():
                    # Pull emon output path from metrics.json (auto-test stores it there).
                    mp = Path(str(row.get("metrics_path") or ""))
                    emon_enabled = ""
                    emon_output_path = ""
                    batch_size = row.get("batch_size") or ""
                    if mp.exists():
                        try:
                            rec = _load_json(mp)
                            emon = rec.get("emon") or {}
                            if isinstance(emon, dict):
                                emon_enabled = str(bool(emon.get("enabled")))
                                emon_output_path = str(emon.get("output_path") or "")
                            env = rec.get("env") or {}
                            if isinstance(env, dict):
                                batch_size = str(env.get("BATCH_SIZE") or batch_size)
                        except Exception:
                            pass

                    merged: Dict[str, Any] = dict(row)
                    merged["variant"] = variant_name
                    merged["emon_enabled"] = emon_enabled
                    merged["emon_output_path"] = emon_output_path
                    merged["batch_size"] = batch_size
                    merged["resource_cpu"] = cpu_expr
                    merged["resource_cpu_count"] = cpu_count if cpu_count > 0 else ""
                    merged["resource_mem_gb"] = scale.mem_gb if scale.mem_gb is not None else ""
                    merged["sglang_max_total_tokens"] = kv
                    rows.append(merged)
        except Exception:
            summary_found = False

        # If the variant failed and we didn't get per-job rows, still record a marker row.
        if rc != 0 and not summary_found:
            rows.append(
                {
                    "variant": variant_name,
                    "job_name": "",
                    "backend": str(scale.job_template.get("backend") or ""),
                    "model": str(scale.job_template.get("model") or ""),
                    "model_id": str(scale.job_template.get("model_id") or ""),
                    "batch_size": "",
                    "token_len": "",
                    "tps": "",
                    "latency_sec": "",
                    "avg_batch_time_sec": "",
                    "count": "",
                    "num_batches": "",
                    "exit_code": str(rc),
                    "started_at_utc": "",
                    "ended_at_utc": "",
                    "log_path": "",
                    "metrics_path": "",
                    "emon_enabled": "",
                    "emon_output_path": "",
                    "emon_summary_xlsx": "",
                    "emon_process_rc": "",
                    "emon_process_log": "",
                    "resource_cpu": cpu_expr,
                    "resource_cpu_count": cpu_count if cpu_count > 0 else "",
                    "resource_mem_gb": scale.mem_gb if scale.mem_gb is not None else "",
                    "sglang_max_total_tokens": kv,
                }
            )

    if args.dry_run:
        print("[ok] dry-run complete (no results generated)")
        return 0

    if scale.emon_process_after_run and rows:
        _process_emon_dirs(
            rows=rows,
            tee=tee,
            process_cmd=scale.emon_process_cmd,
            expected_output=scale.emon_expected_output,
        )
        _maybe_extract_emon_key_values(rows=rows)

    agg_csv = out_dir / "aggregate.csv"
    _write_aggregate_csv(out_csv=agg_csv, rows=rows)
    print(f"[ok] Wrote aggregate: {agg_csv}")

    if (not args.no_analyze) and rows:
        _run_post_analyze(run_dir=out_dir, tee=tee)

    if single_variant:
        run_id = _find_single_run_id(out_dir)
        summary_csv = out_dir / f"summary_{run_id}.csv"
        if summary_csv.exists():
            print(f"[ok] Auto-test summary: {summary_csv}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
